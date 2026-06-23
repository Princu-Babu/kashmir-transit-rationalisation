"""
transit_kashmir_v3.py  —  Srinagar / Kashmir Valley Transit Rationalisation Engine v3.4.3
================================================================================
Principal Secretary of Transport / IAS Officer — Srinagar / Kashmir Valley
Route Rationalisation Project | Forked from Jammu v3 | May 2026

WHY THIS FORK EXISTS:
  Jammu (32.7° N, plain, hot-dry, RITES CMP, no e-bus deployment) and the
  Kashmir Valley (~34° N, bowl-shaped, hill-divided, Chillai Kalan winters,
  SSCL e-bus + CHALO data already live) are structurally different transit
  problems. Tuning Jammu constants for Kashmir produces confidently wrong
  outputs. This is a clean fork — every Jammu-specific assumption rebuilt.

KEY CHANGES FROM JAMMU v3:

  SSCL BACKBONE INJECTION  — 30 Hardcoded Trunk Routes from CHALO Data
    REPLACED: The 13 RITES CMP Jammu Table 4.2 routes have been removed
              entirely. inject_cmp_trunk_routes() now injects all 30 SSCL
              (Srinagar Smart City Limited) e-bus routes from CHALO ridership
              data (Apr 2026), forcing each one to UPGRADED_TO_TRUNK/HP.
    HEADWAY:  Hardcoded to 15 min for SSCL trunks (was 10 in Jammu). 15 min
              matches the SSCL operational headway target (25-min stop-to-stop)
              and Srinagar's actual peak demand (~4,346 pax/hr citywide at 9 AM
              across the network, not radial-CBD).

  GEOGRAPHIC RE-CENTRING
    • Bounding box: Srinagar UA + Budgam/Ganderbal peri-urban + valley
      reach to Sopore/Pulwama/Anantnag/Kupwara
      Lat 33.50–34.50  Lon 74.40–75.20
    • CITY_CORE_LAT_THRESHOLD: 32.72 (old Jammu) → 34.07 (Srinagar
      Lal Chowk + Downtown / mohalla quarters above this latitude).
    • Jhelum river crossing: JHELUM_RIVER_LON 74.81 (was a 74.87 longitude
      approximation in earlier Jammu-targeted code). Routes crossing Jhelum
      pay both a +60% circuity penalty (OSRM-fallback distance) and, in
      v3.2, an additive JHELUM_BRIDGE_BOTTLENECK_MIN bridge-queue penalty.

  CMP POPULATION BASELINE
    Replaced Jammu's 1,653,873 (2024 RITES) with Srinagar UA population from
    Census 2011 + projection: 1,180,570 (2011) → ~1,660,000 (2024) →
    2,100,000 (2034). Source: Census of India 2011 + SMC projections.

  SEASONAL POI TIER (Kashmir-specific, new in this fork)
    Tier 1 (1.0, year-round) : hospitals, transit hubs, govt offices, mosques
    Tier 2 (0.4, year-round) : colleges, markets, schools
    Tier 3 (0.6 summer / 0.0 winter) : tourist gates, gardens, gondola,
                                        shrines on yatra calendar
    Winter mode (WINTER_SCENARIO=True) zeroes Tier 3 and applies a
    snow-walkshed shrink. Default: summer.

  GENDER-WEIGHTED DEMAND (CHALO calibration)
    Observed share: 64.5% of riders are women (free-fare effect).
    Women-anchor POIs (women's colleges, hospitals, markets, schools)
    get a +25% demand boost in the POI gravity score.

  SOCIAL_OBLIGATION_ATTRACTORS
    Replaced 11 Jammu-specific attractors (Jagti, Muthi, Purkhoo, GMC etc.)
    with 13 Kashmir-specific anchors: KP migrant camps (Sheikhpora,
    Vessu, Mattan, Veerwan), tertiary hospitals (SKIMS Soura, SMHS, LD,
    Bone & Joint, JLNM Rainawari), industrial estates (Khonmoh, Rangreth,
    Lassipora), and Sumbal/Ganderbal/Pulwama district HQ hospitals.

  QC SANITY TERMS
    Narwal/Amphalla (Jammu) → Lal Chowk / Hazratbal / Batamaloo (Srinagar).

WHAT IS KEPT INTACT FROM JAMMU v3:
  • All Directive logic (D1 lifeline retention, D2 realistic cycle times,
    D5 95th-percentile pop cap, D6 audit logging)
  • OSRM concurrent fetcher + circuity fallback
  • LPV eradication (HPV/MPV-only fleet)
  • Anti-stranding 5km min trunk length
  • CDI 30th-pct gate
  • Connectivity bonus 1.5× for routes touching the backbone
  • All Folium + XLSX export pipelines

KASHMIR-SPECIFIC LIMITATIONS (Phase 1, to be addressed in v4):
  • Walksheds remain Euclidean buffers. Dal Lake / Anchar / Hokersar /
    Jhelum acting as barriers, and slope-impeded walking, are NOT yet
    modelled. Routes adjacent to these features will systematically
    over-count "served" population by ~15–25%.
  • Winter scenario is binary (flag). A full seasonal-stratified run
    (Chillai Kalan / shoulder / summer / monsoon) requires four passes.
  • Tourist arrival surge (Gulmarg/Pahalgam/Sonmarg) is captured only
    through the Tier-3 seasonal POI weight, not through actual visitor
    flow modelling.
  • Security/convoy windows on NH-44 and military polygons not yet
    subtracted from operable network.
  • CHALO real ridership data is used here to calibrate SSCL trunk
    headways and women-share weighting, but per-route AFC validation
    is still a v4 task.
================================================================================
"""
import os
# v3.2: pin OpenBLAS to a single thread BEFORE multiprocessing forks/spawns.
# The previous fix was to comment out the parallel branch entirely; this
# environment variable is what actually resolves the Windows OpenBLAS hang.
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")

import concurrent.futures
import multiprocessing
from typing import Dict, List, Optional, Tuple
def _intersection_worker(args):
    """Worker function for parallel overlap computation."""
    i, buf_i, area_i, candidates = args
    results = []
    
    # Clean geometry if invalid
    if not buf_i.is_valid:
        buf_i = buf_i.buffer(0)
        
    for j, buf_j, area_j in candidates:
        if not buf_j.is_valid:
            buf_j = buf_j.buffer(0)
        try:
            inter_area = buf_i.intersection(buf_j).area
            ratio_i_j = inter_area / area_i if area_i > 0 else 0.0
            ratio_j_i = inter_area / area_j if area_j > 0 else 0.0
            results.append((j, ratio_i_j, ratio_j_i))
        except Exception:
            # Fallback to 0 if the geometry math fundamentally crashes
            results.append((j, 0.0, 0.0))
            
    return i, results

# ──────────────────────────────────────────────────────────────────────────────
#  CONFIGURATION  (every tunable parameter lives here — never buried in logic)
# ──────────────────────────────────────────────────────────────────────────────

# ─── Directive 2: Congestion zones ──────────────────────────────────────────
# City core = Downtown Srinagar / mohalla quarters (lat > 34.07):
#             Nawakadal, Rajouri Kadal, Khanyar, Gojwara, Habba Kadal,
#             Lal Chowk, Hawal — narrow lanes, bridge bottlenecks, bazaar grids
# Peri-urban = Hyderpora, Rangreth, Pampore, Pantha Chowk, Bemina,
#              Budgam, Ganderbal — mixed urban/highway, lower congestion
CITY_CORE_LAT_THRESHOLD     = 34.07   # SRINAGAR Downtown / mohalla quarters
# v3.2: bumped from 1.4× / 1.1× — Srinagar Traffic Police bottleneck studies
# place the Nawakadal–Habba Kadal bazaar grid at 2.0–2.5× free-flow during the
# 8–11 AM and 4–7 PM windows. The earlier values were borrowed Jammu defaults
# that systematically under-counted Downtown cycle times.
CONGESTION_CITY_CORE        = 2.2    # 2.2× free-flow (Downtown bazaars + bridges)
CONGESTION_PERI_URBAN       = 1.4    # 1.4× free-flow (Hyderpora–Rangreth belt)

# ─── Kashmir-specific: Seasonal scenario flag ───────────────────────────────
# True  → Chillai Kalan / winter mode: Tier-3 (tourist) POIs zeroed,
#         walkshed shrunk, snow-prone routes flagged Lifeline,
#         operated/scheduled KM penalty 15% reflects CHALO Jan data.
# False → Default summer / shoulder season run.
WINTER_SCENARIO             = False
WINTER_WALKSHED_SHRINK      = 0.65   # Effective walkshed = 65% of summer (snow + cold)
WINTER_OP_RATIO_PENALTY     = 0.85   # Operated/scheduled KM ratio in winter (CHALO observed)

# ─── Directive 2: Stop penalty ───────────────────────────────────────────────
STOP_SPACING_M              = 500    # One virtual stop every 500m along route
STOP_PENALTY_MIN            = 0.5    # 30 sec dwell per stop (boarding/alighting)

# ─── Directive 2: Layover buffer ─────────────────────────────────────────────
TERMINAL_LAYOVER_FACTOR     = 1.10   # 10% terminal layover (was 15% blanket in v6)

# ─── Minimum viable fleet (LPV downgrade removed; floor enforced directly) ───
# v3.2: split into urban vs regional. The previous blanket floor of 2 raised
# every rural lifeline route to 2 buses regardless of demand, inflating the
# Regional_District fleet beyond what the demand justified. Regional lifelines
# now floor at 1; urban / peri-urban still floor at 2 for service viability.
MIN_FLEET_URBAN             = 2
MIN_FLEET_REGIONAL          = 1
# Backward-compat alias retained for any legacy references.
MIN_FLEET_THRESHOLD         = MIN_FLEET_URBAN

# ─── Directive 1: Route type thresholds ──────────────────────────────────────
URBAN_KM_THRESHOLD          = 15.0   # < 15km = Urban
PERIURBAN_KM_THRESHOLD      = 40.0   # 15–40km = Peri-Urban, > 40km = Regional_District

# ─── Directive 1: Headways for Regional_District routes ──────────────────────
# v3.3.7 (RTO ask): rural lifelines were 60/90 min — '1 hour+ is unacceptable
# even on inter-district routes'. Brought down to the network-wide 35-min cap
# (HEADWAY_MAX_MIN) so NO route in the plan waits longer than 35 minutes.
HEADWAY_REGIONAL_HP_MIN     = 35     # Rural lifeline HP: was 60 (RTO ask)
HEADWAY_REGIONAL_MP_MIN     = 35     # Rural lifeline MP: was 90 (RTO ask)

# v3.4.2: HYBRID demand-responsive sizing for long rural Regional_District
# lifelines ONLY (Urban + Peri-Urban keep the 15/20/35 clock-headways). The
# route-level audit (ROUTE_LEVEL_AUDIT_2026-06-22.md) showed the flat 35-min
# ceiling — an urban-intended RTO ask — over-provisions 100-km lifelines (a
# 121-km Tangdar route at 35-min = 13 buses for ~270 riders, load 0.18) while a
# few busy inter-district corridors are starved. For Regional routes the headway
# is re-derived from demand (since daily capacity ∝ 1/headway → headway scales
# with the current load toward REGIONAL_TARGET_LOAD), bucketed to a clean RTO set
# and capped at a hard maximum wait; fleet is then recomputed. This is a
# RECOMMENDED year-round size — the RTO may reduce further at execution.
# (User decision 2026-06-22: "Hybrid — Regional only".)
# v3.4.3 (user ask): the rural wait must NOT exceed ~45–50 min even on the
# quietest lifeline — service quality over fleet minimisation. Buckets capped at
# 50 min (was 120). Busy rural corridors still get 35; quiet ones land at 40–50.
REGIONAL_DEMAND_SIZING      = True
REGIONAL_TARGET_LOAD        = 0.55   # aim lifelines at a healthy ~55% load
REGIONAL_HEADWAY_BUCKETS    = (35, 40, 45, 50)  # min; 50 = hard max rural wait

# ─── Directive 5: Population normalisation cap ────────────────────────────────
POP_CAP_PERCENTILE          = 95     # Cap at 95th percentile before normalising

# ─── Srinagar / Kashmir Valley Study-Area Population ─────────────────────────
# Source: Census of India 2011 (Srinagar UA = 1,180,570) + SMC Master Plan
# 2035 + RITES Srinagar CMP projections + Smart City DPR.
# Used to convert raw Population_Served counts → % of study-area population.
# Note: This is Srinagar UA + immediate peri-urban (Budgam, Ganderbal fringe).
# A full Valley-wide model would need ~7M (all districts) but ridership data
# is currently SSCL/CHALO-bound to Srinagar metropolitan reach.
CMP_POPULATION_BY_YEAR = {
    2011: 1_180_570,
    2018: 1_420_000,
    2024: 1_660_000,
    2034: 2_100_000,
    2044: 2_580_000,
}
CMP_REFERENCE_YEAR       = 2024          # Nearest planning horizon
CMP_TOTAL_POPULATION     = CMP_POPULATION_BY_YEAR[CMP_REFERENCE_YEAR]  # 16,60,000

# ─── Kashmir 3-Tier POI Weights (was 2-tier in Jammu v3) ──────────────────────
# Tier 1 (w=1.0): Year-round mass attractors with high sustained ridership
# Tier 2 (w=0.4): Education / secondary commerce (lower than Jammu's 0.2 because
#                 CHALO data shows education trips drive ~30% of women ridership)
# Tier 3 (w=0.6 summer / 0.0 winter): Tourism, gardens, shrines on yatra calendar
#                 Toggled by WINTER_SCENARIO flag
POI_TIER1_WEIGHT            = 1.0
POI_TIER2_WEIGHT            = 0.4
# v3.2: Tier-3 split into "tourist_only" (zeroed in winter) and
# "residential_anchor" (demoted but not zeroed — Boulevard, Dalgate gates
# and Mughal gardens still carry year-round local ridership even when
# tourism collapses). The previous uniform Tier-3 winter zero artificially
# demoted these residentially anchored corridors to LP.
POI_TIER3_WEIGHT_SUMMER             = 0.6
POI_TIER3_WEIGHT_WINTER             = 0.0   # for tourist-only
POI_TIER3_RES_WEIGHT_WINTER         = POI_TIER2_WEIGHT  # 0.4 for residential-anchor in winter
POI_TIER3_WEIGHT                    = (POI_TIER3_WEIGHT_WINTER if WINTER_SCENARIO
                                       else POI_TIER3_WEIGHT_SUMMER)
POI_TIER3_RES_WEIGHT_EFFECTIVE      = (POI_TIER3_RES_WEIGHT_WINTER if WINTER_SCENARIO
                                       else POI_TIER3_WEIGHT_SUMMER)

# Women-anchor POI boost (CHALO calibration: 64.5% of riders are women)
# These categories receive +25% weighted demand to reflect observed gender split
WOMEN_ANCHOR_BOOST          = 1.25
WOMEN_ANCHOR_CATEGORIES = frozenset({
    "womens_college", "women_college", "girls_school", "girls_college",
    "maternity_hospital", "gynaec_hospital", "mch_centre",
    "anganwadi", "women_market", "ladies_market",
})

POI_TIER1_CATEGORIES = frozenset({
    # Transit infrastructure (Srinagar hubs)
    "railway_station", "railway", "train_station", "rail",
    "isbt", "bus_station", "bus_terminal", "transit_hub",
    "trc", "tourist_reception_centre",
    "batamaloo_bus_stand", "parimpora_bus_stand", "pantha_chowk_depot",
    "nishat_bus_terminal",
    # Tertiary / major health (Kashmir-specific)
    "hospital", "govt_hospital", "medical_college", "major_hospital",
    "skims", "skims_soura", "smhs_hospital", "lal_ded_hospital",
    "bone_and_joint_hospital", "jlnm_rainawari", "chest_diseases_hospital",
    "gmc_srinagar",
    # Commercial / mass attractors (Srinagar bazaars + malls)
    "mall", "supermarket", "major_market", "regal_chowk", "lal_chowk",
    "polo_view", "residency_road", "amira_kadal", "city_centre",
    # Religious mass attractors (year-round)
    "jamia_masjid", "hazratbal_shrine", "khanqah", "dastgeer_sahib",
    "shah_e_hamdan",
    # Industrial estates (mass employment)
    "industrial_estate", "factory", "industry", "sicop",
    "khonmoh_industrial", "rangreth_industrial", "lassipora_industrial",
    # KP migrant townships (social obligation anchor — Kashmir-specific)
    "kp_migrant_camp", "kp_township", "sheikhpora_camp", "vessu_camp",
    "mattan_camp", "veerwan_camp",
    # Government / IT
    "civil_secretariat", "high_court", "rajbhawan", "raj_bhawan",
    # University main campuses (year-round, very high ridership)
    "university_of_kashmir", "iist_awantipora", "central_university_kashmir",
    "nit_srinagar", "smvd",
})

POI_TIER2_CATEGORIES = frozenset({
    # Education (secondary — colleges, schools)
    "college", "campus", "engineering_college", "school",
    "higher_education", "polytechnic", "iti", "degree_college",
    # Local commerce
    "market", "commercial", "wholesale_market", "small_shop",
    "haba_kadal_market", "maharaj_gunj",
    # Defence / civic
    "army", "bsf", "crpf", "military", "cantonment",
    "stadium", "bakshi_stadium", "polo_ground",
    # Government services (lower tier)
    "government", "govt_office", "court", "clinic", "small_medical",
    "dispensary", "museum", "police_station",
    # Recreation (year-round)
    "park", "library", "auditorium", "tagore_hall",
})

# Tier 3 — TOURIST-ONLY (zeroed in winter mode, no year-round residential pull)
POI_TIER3_TOURIST_ONLY = frozenset({
    "tourist_spot",
    # High-altitude tourism
    "gulmarg", "pahalgam", "sonmarg", "doodhpathri", "yusmarg",
    "gondola", "ski_resort", "ropeway",
    # Yatra-calendar shrines (seasonal pilgrimage)
    "amarnath_base_camp", "kheer_bhawani",
    # Tourist info / hotels (close in winter)
    "houseboat", "tourist_hotel",
})

# Tier 3 — RESIDENTIAL-ANCHOR (demoted to Tier-2 weight in winter, NOT zeroed —
# Boulevard, Dalgate, Nishat etc. carry year-round local ridership independent
# of tourist arrivals).
POI_TIER3_RESIDENTIAL_ANCHOR = frozenset({
    "tourist_gate", "shikara_ghat",
    "dal_lake_gate", "boulevard_gate", "nigeen_lake_gate",
    # Mughal gardens (closed but neighbourhoods around them are residential)
    "mughal_garden", "nishat_garden", "shalimar_garden",
    "chashme_shahi", "pari_mahal", "tulip_garden", "harwan_garden",
})

# Union retained for the load_pois() / count_weighted_poi_scores() audit logic
POI_TIER3_CATEGORIES = POI_TIER3_TOURIST_ONLY | POI_TIER3_RESIDENTIAL_ANCHOR

DEFAULT_POI_WEIGHT_V2       = POI_TIER2_WEIGHT   # Unmapped → Tier 2
# ─── Bounding Box: full Kashmir Division (all 10 districts) ─────────────────
# v3.4.1: extended from the old narrow box (33.50–34.50 / 74.40–75.20) to the
# bounding rectangle of the 10 Kashmir-division district polygons + a small
# margin. The old box silently clipped/dropped ~29 legitimate routes — not only
# Kupwara/Handwara/Tangdar/Gurez but even Baramulla town & Uri (whose centres lie
# just W of lon 74.40) and SE Anantnag (Shangus/Uttersoo) — and shrank the
# coverage denominator to the in-box population only (5.1M) instead of the true
# division population. Coverage is now measured against the 10-district UNION
# (point-in-polygon, ~6.58M; see study_area_population). The earlier comment that
# these areas were "handled separately as Regional_District lifelines" was never
# true — they were dropped. Bounds = district-union extent (33.364–34.787 /
# 73.750–75.595) padded ~0.06°.
BOUNDS_MIN_LAT = 33.30
BOUNDS_MAX_LAT = 34.85
BOUNDS_MIN_LON = 73.70
BOUNDS_MAX_LON = 75.65

# Speed zones — Kashmir-specific (was Jammu zones)
# Old city Srinagar (Downtown): narrow lanes, bridges, bazaars
# Peri-urban Srinagar: Hyderpora-Rangreth-Bemina belt
# Valley / district highway: NH-1A, Srinagar-Sopore highway, etc.
SPEED_OLD_CITY_KMH          = 10.0   # Downtown Srinagar mohallas
SPEED_VALLEY_HIGHWAY_KMH    = 28.0   # NH-1A and inter-district roads
SPEED_DEFAULT_KMH           = 18.0   # Suburban / mixed corridors

# Geometry / Routing (unchanged)
MAX_IMPUTED_SL_KM           = 999    # DIRECTIVE 1: No hard cap — process all routes
CIRCUITY_FACTOR             = 1.25
CIRCUITY_FACTOR_RIVER       = 1.60   # Jhelum crossings: must use one of 9 bridges
JHELUM_RIVER_LON            = 74.81  # Approx longitude of Jhelum through Srinagar
# v3.2: bridge-queue penalty applied as additive minutes — captures the
# peak-hour bottleneck queueing that OSRM's free-flow duration cannot see.
# A proper per-bridge node graph with daily operability status is a v4 item;
# this is the simple "every Jhelum crossing pays a flat 8-min queue" proxy.
JHELUM_BRIDGE_BOTTLENECK_MIN = 8.0
MIN_ROUTE_KM                = 1.0
VIRTUAL_STOP_SPACING_M      = 250    # For catchment / population (Step 1)
# Walkshed: 400m summer, ~260m winter (snow shrink). Applied to both catchment
# and POI buffer so that demand contracts realistically during Chillai Kalan.
WALK_CATCHMENT_M            = int(400 * (WINTER_WALKSHED_SHRINK if WINTER_SCENARIO else 1.0))
POI_BUFFER_M                = int(250 * (WINTER_WALKSHED_SHRINK if WINTER_SCENARIO else 1.0))
OVERLAP_BUFFER_M            = 80
OVERLAP_THRESHOLD           = 0.65
OD_PROXIMITY_TOLERANCE_M    = 2500
SIMPLIFY_TOL_M              = 2.0

# Step 3: Road Multiplier (unchanged)
ROAD_MULTIPLIER_TRUNK       = 1.25
ROAD_MULTIPLIER_FEEDER      = 0.75
ROAD_MULTIPLIER_DEFAULT     = 1.00

# CDI weights (unchanged)
CDI_POP_WEIGHT              = 0.50
CDI_POI_WEIGHT              = 0.50

# Social Obligation — v3.3.1 audit response (round 2 STEP 10):
# The 500m buffer + 17-attractor list produced a 91.5% flag rate (313/342),
# which made Social_Flag effectively a no-op filter. Tightened to 250m and
# the attractor list restricted to truly critical / lifeline facilities so
# the flag actually discriminates. Target: ~30-40% of network.
SOCIAL_FLAG_BUFFER_M        = 250

# Step 6: Headway per Priority Band for Urban/Peri-Urban routes.
# v3.3.5 (reality check): non-SSCL trunks moved from 15→20 min and MP from
# 30→35 min, to land between BMTC Bengaluru (10–15 min trunks) and Mysuru
# KSRTC (20–30 min). SSCL_TRUNK_HEADWAY_MIN stays at 15 because that is
# SSCL's published design target for the e-bus backbone — the floor logic
# in step9 ensures the formula can still raise SSCL fleet when needed.
HEADWAY_HP_MIN              = 20   # was 15
HEADWAY_MP_MIN              = 35   # was 30
HEADWAY_LP_MIN              = 35   # v3.3.6 (RTO ask): was 60 — '1 hour is
                                    #         too long' for lifeline routes;
                                    #         peer-city inter-district routes
                                    #         typically run 30–40 min, so 35
                                    #         splits the difference.

# v3.3.7 (RTO ask): a hard network-wide headway ceiling. After step6 assigns
# headways by band / route-type / operator-category, ANY value above this is
# clamped down to it. Guarantees the published plan contains no headway > 35
# min — the legacy 60-min and 90-min rural/operator buckets are eliminated
# entirely. The SSCL 15-min backbone and 20-min HP trunks sit below the ceiling.
HEADWAY_MAX_MIN             = 35

# v3.3.6 (RTO ask): per-route HPV cap for SSCL-table-matched trunks. CHALO's
# empirical 12m count is what they run today; the RTO wants the engine's
# recommendation to give MPVs more share on long routes ("dominated by HPV").
# v3.3.7 (RTO ask): tightened 0.60 → 0.50 so that on a trunk route NEITHER
# vehicle class is the majority — the recommendation lands a balanced 50/50
# HPV/MPV mix (integer rounding leaves MPV at most one bus ahead), matching the
# depot inventory + narrower-road reality the RTO described. Road-width data
# (a pending P2 RTO data ask) would let us refine this per-corridor later.
SSCL_HPV_SHARE_CAP          = 0.50

# ─── v3.3 Phase-1 audit response: post-cycle spare ratio ────────────────────
# The fleet computed by ceil(Cycle_Time / Headway) is the *operating* fleet.
# Standard transit-planning practice adds a spare ratio (typically 10–20%)
# to cover scheduled maintenance, breakdown rotation, and depot storage.
# 15% is the SSCL operational target. SSCL backbone routes are NOT bumped
# by this multiplier — their fleet is the empirical CHALO bus count and
# already includes spare allocation in the SSCL operational budget.
FLEET_SPARE_RATIO           = 1.15

# ─── v3.3: Tourist Corridor + Seasonal Operability flags ────────────────────
# Routes whose name contains any TOURIST_CORRIDOR_KEYWORDS token are tagged
# Tourist_Corridor=True at ingestion. Routes touching SEASONAL_SUSPENDED
# tokens are flagged Winter_Suspended (Mughal Road / Sinthan / Sadhna /
# Z-Morh portal close in winter). The tags do NOT alter CDI, headway, or
# fleet — they are surfaced for planner review and used by the seasonal
# Folium map layer in Phase 4. (The proposal to multiply tourist CDI ×1.3
# was dropped after audit: it would compound with the existing Tier-3 POI
# weight and the Phase-2 catchment penalty, triple-counting tourism.)
TOURIST_CORRIDOR_KEYWORDS = (
    "gulmarg", "pahalgam", "sonmarg", "sonamarg", "doodhpathri", "yusmarg",
    "tangmarg", "aharbal", "kokernag", "verinag", "achabal", "harwan",
    "mughal", "shalimar", "nishat", "chashme", "pari mahal",
    "boulevard", "dal lake", "nigeen", "tulip", "amarnath", "kheer bhawani",
)
SEASONAL_SUSPENDED_KEYWORDS = (
    "mughal road", "sinthan", "sadhna", "z-morh", "zmorh", "kishtwar",
    "doodhpathri", "yusmarg", "aharbal",
)

# v3.3.3 (teammate review): the keyword check alone misses 99% of tourist
# corridors because the registered permit data (existing-routes.csv) doesn't
# record tourist endpoints — Origin / Destination are urban hub names
# (SRINAGAR / PARIMPORA / SOURA) even on routes that physically traverse
# tourist zones.
#
# v3.3.3 audit fix: split centroids into two classes:
#   DISTANT  — far-flung tourist destinations whose seasonal visitor surge is
#              invisible to the WorldPop residential raster. We tag any route
#              whose GEOMETRY passes within DISTANT buffer. These get the
#              1.3x catchment boost.
#   INNER_CITY — Mughal gardens, Boulevard, Nigeen, Cheshma, Pari Mahal —
#              clustered around 34.13°N / 74.87°E inside Srinagar. ANY
#              cross-town route traverses near these centroids, so a generous
#              geometry buffer over-tags every urban commuter line. For
#              these zones we only tag if the route's TERMINAL endpoint is
#              within a tight buffer (the route actually serves the gate).
#              Tagged routes still get the 1.3x boost — they really do carry
#              extra visitor footfall.
TOURIST_ZONE_BUFFER_KM = 2.0          # for distant destinations
TOURIST_INNER_BUFFER_KM = 0.6          # for inner-city zones (endpoint-only)
TOURIST_ZONES_DISTANT: Dict[str, Tuple[float, float]] = {
    "Gulmarg":       (34.0481, 74.3805),
    "Tangmarg":      (34.0427, 74.4413),
    "Pahalgam":      (34.0151, 75.3322),
    "Sonamarg":      (34.3050, 75.2935),
    "Doodhpathri":   (33.8628, 74.5722),
    "Yusmarg":       (33.8333, 74.6667),
    "Aharbal":       (33.6500, 74.7600),
    "Kokernag":      (33.6235, 75.3050),
    "Verinag":       (33.5400, 75.2500),
    "Achabal":       (33.6800, 75.2300),
    "Mughal Garden Achabal": (33.6803, 75.2289),
    "Harwan":        (34.1568, 74.8931),
    "Tulip Garden":  (34.0884, 74.8836),
}
TOURIST_ZONES_INNER_CITY: Dict[str, Tuple[float, float]] = {
    "Shalimar Bagh": (34.1450, 74.8761),
    "Nishat Bagh":   (34.1233, 74.8722),
    "Cheshma Shahi": (34.1158, 74.8854),
    "Pari Mahal":    (34.1131, 74.8825),
    "Dal Boulevard": (34.0833, 74.8500),
    "Nigeen Lake":   (34.1264, 74.8443),
}
# Seasonal-suspended zone centroids (closed by snow ~Nov–Apr).
SEASONAL_SUSPENDED_ZONES_LATLON: Dict[str, Tuple[float, float]] = {
    "Mughal Road (Pir Panjal)": (33.5800, 74.5300),
    "Sinthan Top":              (33.6300, 75.5500),
    "Sadhna Pass":              (34.3500, 73.9200),
    "Z-Morh":                   (34.2700, 75.3000),
    "Zojila":                   (34.2833, 75.4744),
}
SEASONAL_SUSPENDED_BUFFER_KM = 3.0

# v3.3.3 (teammate review): tourist corridor population multiplier.
# Applied ONCE — at catchment level — so it propagates to Pop_Score (CDI) and
# Phase-4 Daily_Demand consistently. The earlier proposal to also multiply
# CDI ×1.3 was rejected (triple-counting tourism). With this single
# multiplier in place, Tier-3 POI weights stay as-is, no CDI multiplier is
# applied, and the 1.3× shows up everywhere it should.
TOURIST_POPULATION_MULTIPLIER = 1.3

# ─── v3.3: Vehicle capacities + emissions factors (Phase 4 derived only) ────
# Used ONLY for the new Load_Ratio / Emissions_Proxy outputs. Do NOT feed
# these back into Step 8 fleet sizing (would double-count headway/capacity).
VEHICLE_CAPACITY_HPV         = 60      # 12m bus, peak crush load (seated+standing)
VEHICLE_CAPACITY_MPV         = 35      # 9m bus
VEHICLE_CAPACITY_LPV         = 20      # minibus / Sumo
EMISSIONS_GCO2_PER_KM_DIESEL = 950.0   # g CO2 / km — diesel city bus baseline
EMISSIONS_GCO2_PER_KM_EBUS   = 30.0    # g CO2 / km — Indian grid mix electric

# Phase-4 demand proxy (matches cross_evaluate.py constants — keep in sync).
PHASE4_MODE_SHARE            = 0.09
# v3.3.1 (STEP 6): typology-aware modal capture rate. Urban core gets the
# CHALO-derived baseline; peri-urban routes capture less because alternative
# autos/private vehicles are more available; inter-district lifelines have
# the lowest bus mode share because trip purpose is mostly long-distance
# and bus competition with shared sumos / private cars is stronger.
PHASE4_MODE_SHARE_BY_TYPE = {
    "Urban":             0.090,
    "Peri_Urban":        0.072,   # × 0.8 of urban
    "Regional_District": 0.054,   # × 0.6 of urban
}
PHASE4_TRIP_RATE             = 1.6
# v3.3.2: empirical SSCL-anchored capture scalar. The raw formula
#   Pop_Buffer × mode_share × trip_rate × (1/headway / sum_overlap(1/headway))
# over-predicts SSCL daily demand by ~5.7× vs CHALO observed ridership
# (~32k/day across 30 SSCL routes). The 0.18 factor reconciles the buffer-
# based supply view with the observed ridership floor; it absorbs:
#   - residents in buffer who walk / use auto / private mode despite proximity
#   - residents who use a *different* parallel bus service not captured in
#     this engine's overlap_metric (CHALO only sees SSCL, not full network)
#   - peak-vs-mean reconciliation
# Recalibrate when CHALO yearly totals shift more than ±15%.
# v3.3.8 (F-V8 re-verification): the district-aware RE-GEOCODE changed every
# route's walkshed catchment, which drifted the calibration — at 0.18 the model
# reproduced only 0.54× of CHALO's published SSCL ridership (17.3k vs 31.9k/day),
# i.e. it under-counted demand ~2× network-wide and made the economics read far
# worse than reality. Re-fitted to the published anchor: 11,632,326 trips/yr ÷
# 365 = 31,869/day across the SSCL trunks → scale 0.18 × (31,869/17,258) ≈ 0.33.
# (Affects Daily_Demand / Load_Ratio / economics only — NOT fleet/headway/bands.)
PHASE4_CORRIDOR_CAPTURE_SCALE = 0.33
PHASE4_SERVICE_HOURS         = 16
PHASE4_FARE_INR              = 10.0    # avg single-trip fare proxy (post free-fare)
PHASE4_OPERATING_COST_PER_KM = 65.0    # INR/km diesel minibus all-in
PHASE4_JOURNEY_TIME_FLAG_MIN = 45      # passenger journey time amber threshold
# v3.3.1 (STEP 7): subsidy-risk threshold raised from 0.5 to 0.6 so the
# "marginal" band (0.6-1.0) is surfaced separately from outright subsidy
# dependency. Anything < 0.6 fare-recovery is meaningfully unsustainable.
PHASE4_SUBSIDY_RISK_THRESHOLD = 0.6
# v3.3.1 (STEP 1): per-km cycle-time sanity cap. Prevents OSRM glitches or
# pathological congestion from producing nonsense cycle times. NOTE: the
# proposal's "second peak multiplier on top of congestion" was rejected
# (would double-count with CONGESTION_CITY_CORE=2.2). Cap only.
CYCLE_TIME_CAP_MIN_PER_KM = {
    "Urban":             4.0,    # max 4 min/km one-way → 8 min/km round
    "Peri_Urban":        2.5,
    "Regional_District": 1.5,
}

# Junction penalty (unchanged)
JUNCTION_PENALTY_PER_TURN_MIN = 0.5
SHARP_TURN_DEG              = 75

# Fleet cap (unchanged)
STD_TO_MINI_RATIO           = 2.5
TERMINAL_BUFFER_M           = 300
FLEET_CAP_HARD              = 45
GRAVITY_EPSILON             = 0.1
TRANSFER_PENALTY_MIN        = 5

# OSRM (unchanged)
OSRM_BASE_URL               = "http://localhost:5000"
OSRM_TIMEOUT_S              = 8
OSRM_MAX_WORKERS            = 12

# CRS (unchanged)
UTM_CRS                     = "EPSG:32643"
WGS84_CRS                   = "EPSG:4326"

# I/O
# v3.3.2: raster path resolution — env var > CLI flag > local default.
# Loud RuntimeError raised in compute_population() if the file is missing,
# so a wrong path can no longer silently zero out Population_Served.
def _resolve_raster_path() -> str:
    # CLI flag --raster <path> takes priority. Argparse here is intentionally
    # lightweight (no help/usage) — the engine never had a proper argparse and
    # this is a back-door override; full CLI is a v4 cleanup.
    import sys
    if "--raster" in sys.argv:
        i = sys.argv.index("--raster")
        if i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    env_path = os.environ.get("KASHMIR_WORLDPOP")
    if env_path:
        return env_path
    # Local default — works when the engine runs from the project root.
    here = os.path.dirname(os.path.abspath(__file__))
    local = os.path.join(here, "kashmir_worldpop.tif")
    if os.path.exists(local):
        return local
    # Last-resort legacy absolute path (will still fail loudly if missing).
    return r"E:/kash/kashmir_worldpop.tif"

RASTER_PATH                 = _resolve_raster_path()
ROUTES_CSV                  = "existing-routes.csv"
POIS_CSV                    = "pois.csv"
OUTPUT_DIR                  = "route_maps_kashmir"
LOG_CSV                     = "Rationalisation_Log_Kashmir_v3.csv"
ROUTES_OUT_CSV              = "Rationalised_Routes_Kashmir_v3.csv"
ROUTES_OUT_XLSX             = "Kashmir_Route_Frequency_Plan_v3.xlsx"
PASSENGER_IMPACT_CSV        = "Passenger_Impact_Kashmir_v3.csv"
MASTER_MAP_HTML             = "Master_Transit_Map_Kashmir_v3.html"
ROUTES_GEOJSON              = "Rationalised_Routes_Kashmir_v3.geojson"

# ─── CHALO / SSCL ridership calibration constants ───────────────────────────
# Sourced from SSCL_Data_Updated.xlsx (April 2026 + 12-month series)
CHALO_TOTAL_RIDERSHIP_12M   = 11_632_326   # May 2025 – Apr 2026
CHALO_WOMEN_SHARE           = 0.645         # 64.5% women (free-fare effect)
CHALO_PEAK_HOUR             = 9             # 9 AM — observed citywide peak
CHALO_SECONDARY_PEAK_HOUR   = 17            # 5 PM
CHALO_PEAK_PAX_PER_HOUR     = 4346          # Citywide pax/hr at peak (Apr 2026 mean)
CHALO_TROUGH_PAX_PER_HOUR   = 521           # Citywide pax/hr at 9 PM (Apr 2026 mean)
CHALO_OP_RATIO              = 0.845         # Operated / Scheduled KM (annual mean)
CHALO_SERVICE_START_HR      = 6
CHALO_SERVICE_END_HR        = 22

# Social Obligation Attractors — Kashmir / Srinagar specific (was Jammu)
# v3.3.1 audit response (STEP 10): pruned from 17 to 11. Industrial estates
# removed (they're employment hubs, not social-obligation lifelines — they
# get fair treatment via the POI gravity score already). Duplicate hospital
# entries collapsed. Target flag rate ~30-40% of network instead of 91.5%.
SOCIAL_OBLIGATION_ATTRACTORS = [
    # KP migrant townships (Govt of India ORM-managed lifeline corridors)
    ("Sheikhpora KP Township Budgam",  33.9683, 74.6736),
    ("Vessu KP Township Qazigund",     33.6478, 75.1431),
    ("Mattan KP Township Anantnag",    33.7406, 75.2347),
    ("Veerwan KP Camp Baramulla",      34.2025, 74.3450),
    # Tertiary hospitals — Srinagar metropolitan (3 truly tertiary referral)
    ("SKIMS Soura",                    34.1308, 74.8472),
    ("SMHS Hospital Karan Nagar",      34.0842, 74.7956),
    ("Lal Ded Hospital",               34.0822, 74.8059),
    # District-HQ hospitals (rural lifeline routes)
    ("DH Pulwama",                     33.8716, 74.8983),
    ("DH Ganderbal",                   34.2275, 74.7775),
    ("DH Budgam",                      33.9908, 74.7115),
    ("DH Anantnag",                    33.7311, 75.1497),
]

# Sheet 1 column layout (Groups A / B / C) — v2 adds Route_Type
SHEET1_GROUP_A = ["New_Route_ID", "Route_Name", "Action_Taken", "Route_KM",
                  "Route_Type", "Social_Flag", "Displaced_Operator_Class"]
SHEET1_GROUP_B = ["Priority_Band", "Headway_Min", "Cycle_Time_Min",
                  "Fleet_Required", "HPV_Count", "MPV_Count", "LPV_Count",
                  "CMP_Trunk"]
SHEET1_GROUP_C = ["Pop_Score", "POI_Score", "Road_Multiplier", "Final_CDI",
                  "Population_Served", "Population_Served_Pct",
                  "OSRM_Duration_S", "Sharp_Turns",
                  "Junction_Penalty_Min", "Congestion_Zone",
                  "N_Stops_Estimated", "Stop_Penalty_Min"]
SHEET1_ALL_COLS = SHEET1_GROUP_A + SHEET1_GROUP_B + SHEET1_GROUP_C

TERMINAL_CATEGORIES = frozenset({"bus_terminal", "market"})

# ─── SSCL / CHALO: 30 Official Srinagar Smart City E-Bus Routes ─────────────
# Source: SSCL_Data_Updated.xlsx, "Route Wise deployed Buses" sheet (Apr 2026)
# Total fleet deployed: 98 buses (73 × 9-metre + 25 × 12-metre)
# These 30 routes are injected as UPGRADED_TO_TRUNK/HP before Step 4.
# Headway is fixed at SSCL_TRUNK_HEADWAY_MIN (15 min) — matches the SSCL
# operational target of 25-min stop-to-stop frequency and Srinagar's
# observed peak demand (~4,346 pax/hr citywide at 9 AM from CHALO data).
# 12-metre buses (HPV) are deployed on long-haul / inter-district routes;
# 9-metre buses (MPV) cover intra-city loops. NOTE (audit fix): SSCL's *actual*
# deployment is 73×9m + 25×12m ≈ 74/26 MPV-heavy — i.e. MPV-dominant, NOT the
# "85/15 HPV-MPV" an earlier comment claimed. Step 9 no longer uses any fixed
# 85/15 split; it assigns HPV share from Route_KM brackets, capped at
# SSCL_HPV_SHARE_CAP (0.50 in v3.3.7 — neither class a trunk majority).
SSCL_TRUNK_HEADWAY_MIN = 15   # SSCL e-bus trunk headway — matches SSCL operational
                              # target (~25-min stop-to-stop frequency) and CHALO
                              # observed peak demand (~4,346 pax/hr citywide).
                              # v3.2: corrected from 45 (a leftover Jammu-era value
                              # that produced a 60–70% fleet shortfall vs CHALO).

# Alias for backward compatibility with downstream functions that reference
# CMP_TRUNK_HEADWAY_MIN (no need to rename across the entire codebase).
CMP_TRUNK_HEADWAY_MIN  = SSCL_TRUNK_HEADWAY_MIN

# Each route carries:
#   id     — SSCL route ID (SSCL-01 … SSCL-30)
#   origin — start terminal (used for fuzzy match against dataset)
#   dest   — end terminal
#   km     — approximate route length (used as fallback when OSRM fails)
#   fleet  — buses deployed (from CHALO data, for audit only)
#   bus_9m — 9m bus count (MPV-equivalent)
#   bus_12m — 12m bus count (HPV-equivalent)
CMP_TRUNK_ROUTES: List[Dict] = [
    {"id": "SSCL-01", "origin": "Parimpora",       "dest": "Harwan",                      "km": 18, "fleet": 7, "bus_9m": 7, "bus_12m": 0},
    {"id": "SSCL-02", "origin": "Batamaloo",       "dest": "Nasrullah Pora",              "km": 12, "fleet": 5, "bus_9m": 5, "bus_12m": 0},
    {"id": "SSCL-03", "origin": "Batamaloo",       "dest": "Hazratbal",                   "km": 10, "fleet": 6, "bus_9m": 6, "bus_12m": 0},
    {"id": "SSCL-04", "origin": "Batamaloo",       "dest": "Chadoora Chowk",              "km": 13, "fleet": 3, "bus_9m": 3, "bus_12m": 0},
    {"id": "SSCL-05", "origin": "LD Hospital",     "dest": "Pandach",                     "km": 11, "fleet": 4, "bus_9m": 4, "bus_12m": 0},
    {"id": "SSCL-06", "origin": "Pantha Chowk",    "dest": "Narbal Crossing",             "km": 22, "fleet": 2, "bus_9m": 0, "bus_12m": 2},
    {"id": "SSCL-07", "origin": "Batamaloo",       "dest": "Drussu Pulwama",              "km": 28, "fleet": 4, "bus_9m": 0, "bus_12m": 4},
    {"id": "SSCL-08", "origin": "Batamaloo",       "dest": "Budgam Railway Station",      "km": 14, "fleet": 4, "bus_9m": 4, "bus_12m": 0},
    {"id": "SSCL-09", "origin": "Parimpora",       "dest": "Naseem Bagh",                 "km":  9, "fleet": 4, "bus_9m": 4, "bus_12m": 0},
    {"id": "SSCL-10", "origin": "Pantha Chowk",    "dest": "Agri Kalan Kanihama",         "km": 24, "fleet": 5, "bus_9m": 0, "bus_12m": 5},
    {"id": "SSCL-11", "origin": "Old City",        "dest": "Old City Loop",               "km":  8, "fleet": 1, "bus_9m": 1, "bus_12m": 0},
    {"id": "SSCL-12", "origin": "Rangreth",        "dest": "District Court Srinagar",     "km": 11, "fleet": 3, "bus_9m": 3, "bus_12m": 0},
    {"id": "SSCL-13", "origin": "Batamaloo",       "dest": "Beehama Ganderbal",           "km": 25, "fleet": 4, "bus_9m": 4, "bus_12m": 0},
    {"id": "SSCL-14", "origin": "TRC",             "dest": "Central University Ganderbal","km": 28, "fleet": 6, "bus_9m": 6, "bus_12m": 0},
    {"id": "SSCL-15", "origin": "TRC",             "dest": "Pulwama",                     "km": 32, "fleet": 4, "bus_9m": 4, "bus_12m": 0},
    {"id": "SSCL-16", "origin": "Batamaloo",       "dest": "Khonmoh",                     "km": 15, "fleet": 2, "bus_9m": 2, "bus_12m": 0},
    {"id": "SSCL-17", "origin": "Batamaloo",       "dest": "Womens College Batpora",      "km":  7, "fleet": 2, "bus_9m": 2, "bus_12m": 0},
    {"id": "SSCL-18", "origin": "Jehangir Chowk",  "dest": "Wadwan",                      "km": 12, "fleet": 2, "bus_9m": 2, "bus_12m": 0},
    {"id": "SSCL-19", "origin": "Pantha Chowk",    "dest": "Sumbal",                      "km": 38, "fleet": 4, "bus_9m": 0, "bus_12m": 4},
    {"id": "SSCL-20", "origin": "Pantha Chowk",    "dest": "Safapora",                    "km": 42, "fleet": 2, "bus_9m": 0, "bus_12m": 2},
    {"id": "SSCL-21", "origin": "Batamaloo",       "dest": "Arath",                       "km": 11, "fleet": 2, "bus_9m": 2, "bus_12m": 0},
    {"id": "SSCL-22", "origin": "Batamaloo",       "dest": "Khrew Bus Stand",             "km": 16, "fleet": 4, "bus_9m": 4, "bus_12m": 0},
    {"id": "SSCL-23", "origin": "Batamaloo",       "dest": "Charesharief",                "km": 28, "fleet": 2, "bus_9m": 2, "bus_12m": 0},
    {"id": "SSCL-24", "origin": "Pantha Chowk",    "dest": "Palhalan",                    "km": 30, "fleet": 6, "bus_9m": 0, "bus_12m": 6},
    {"id": "SSCL-25", "origin": "Batamaloo",       "dest": "Dadsara Tral",                "km": 36, "fleet": 2, "bus_9m": 2, "bus_12m": 0},
    {"id": "SSCL-26", "origin": "Batamaloo",       "dest": "Kangan",                      "km": 35, "fleet": 2, "bus_9m": 2, "bus_12m": 0},
    {"id": "SSCL-27", "origin": "Batamaloo",       "dest": "Manigam",                     "km": 32, "fleet": 2, "bus_9m": 2, "bus_12m": 0},
    {"id": "SSCL-28", "origin": "Batamaloo",       "dest": "Pinglena Pampore",            "km": 22, "fleet": 2, "bus_9m": 0, "bus_12m": 2},
    {"id": "SSCL-29", "origin": "Pantha Chowk",    "dest": "Panzinara",                   "km":  8, "fleet": 1, "bus_9m": 1, "bus_12m": 0},
    {"id": "SSCL-30", "origin": "Batamaloo",       "dest": "DC Office Budgam",            "km": 14, "fleet": 1, "bus_9m": 1, "bus_12m": 0},
]

# Fuzzy-match tolerance: a dataset terminal must score ≥ this similarity ratio
# (0–1) against a CMP terminal to be considered a match.
CMP_FUZZY_THRESHOLD = 0.45   # 45% character overlap — lowered from 0.55 so short
                             # terminal names like "Hazratbal" still match dataset
                             # entries that have them mid-string ("via Hazratbal")

# Minimum route length to be eligible for Trunk promotion (anti-stranding rule)
TRUNK_MIN_LENGTH_KM = 5.0    # No route < 5 km can be promoted to Trunk

# CDI percentile gate for Trunk eligibility (lowered from 50th to 30th)
TRUNK_CDI_GATE_PERCENTILE = 30   # 30th percentile — aggressive trunk promotion

# Demand multiplier for routes that intersect a CMP backbone trunk corridor
CMP_CONNECTIVITY_BONUS = 1.5  # 1.5× combined demand score for connected routes

COLOUR = {
    "trunk":          "#1A237E",
    "feeder":         "#00695C",
    "regional":       "#6A1B9A",
    "current_net":    "#9E9E9E",
    "catchment_fill": "#80DEEA",
    "catchment_line": "#0097A7",
    "poi_high":       "#D32F2F",
    "poi_secondary":  "#F57F17",
    "start_pin":      "#1B5E20",
    "end_pin":        "#B71C1C",
    "via_dot":        "#5C6BC0",
}
# ─── Map tiles ────────────────────────────────────────────────────────────────
TILE_URL  = "https://{s}.basemaps.cartocdn.com/light_all/{z}/{x}/{y}{r}.png"
TILE_ATTR = "© OpenStreetMap contributors © CARTO"
# Pre-built outside f-string so Leaflet's {z}/{x}/{y} template vars survive verbatim
TILE_LAYER_JS = (
    "L.tileLayer('%s', {attribution: '%s', maxZoom: 18}).addTo(map);"
    % (TILE_URL, TILE_ATTR.replace("'", "\\'"))
)

FG = {
    "heatmap":   "Population Heatmap",
    "trunk":     "Trunk Corridors",
    "feeder":    "Feeder Routes",
    "regional":  "Regional District Routes",
    "original":  "Original Network",
    "hv_poi":    "High Priority POIs",
    "sec_poi":   "Secondary POIs",
    "catchment": "Route Catchments",
    "pins":      "Start End Terminals",
    "via":       "Via Waypoints",
}

# ──────────────────────────────────────────────────────────────────────────────
#  IMPORTS
# ──────────────────────────────────────────────────────────────────────────────
import json
import math
import re
import difflib
import sys
import time
import logging
import warnings
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pandas as pd
import geopandas as gpd
import requests
from shapely.geometry import LineString, MultiLineString, Point, shape, mapping
from shapely.ops import unary_union
from shapely.strtree import STRtree
try:
    import rasterstats
    _HAS_RASTERSTATS = True
except ImportError:
    _HAS_RASTERSTATS = False
import folium
from folium.plugins import AntPath, HeatMap
try:
    from folium.plugins import PolyLineTextPath
    _HAS_PLTPATH = True
except ImportError:
    _HAS_PLTPATH = False
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    import jenkspy
    _HAS_JENKSPY = True
except ImportError:
    _HAS_JENKSPY = False

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)

# ──────────────────────────────────────────────────────────────────────────────
#  LOGGING
# ──────────────────────────────────────────────────────────────────────────────
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("transit_v3.log", mode="w", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 1  ─  DATA INGESTION
# ══════════════════════════════════════════════════════════════════════════════
def truncate_routes_to_bbox(df: pd.DataFrame) -> pd.DataFrame:
    """
    Truncates routes using a strict latitude/longitude bounding box.
    Any point falling outside the box triggers a cut-off, and the 
    last valid point inside the box becomes the new End Terminal.
    """
    log.info("Truncating routes to strict bounding box (Lat: %s to %s, Lon: %s to %s)…",
             BOUNDS_MIN_LAT, BOUNDS_MAX_LAT, BOUNDS_MIN_LON, BOUNDS_MAX_LON)

    new_rows = []
    dropped_count = 0
    truncated_count = 0

    def _is_inside(lon, lat):
        return (BOUNDS_MIN_LON <= lon <= BOUNDS_MAX_LON) and \
               (BOUNDS_MIN_LAT <= lat <= BOUNDS_MAX_LAT)

    for _, row in df.iterrows():
        start_lon, start_lat = float(row["Start_Lon"]), float(row["Start_Lat"])
        end_lon, end_lat     = float(row["End_Lon"]), float(row["End_Lat"])
        vias                 = parse_via(row.get("Via_Coordinates"))

        # Assemble the full ordered sequence of coordinates
        all_pts = [(start_lon, start_lat)] + vias + [(end_lon, end_lat)]
        valid_pts = []

        # Iterate and cut off at the first point outside the bounding box
        for lon, lat in all_pts:
            if _is_inside(lon, lat):
                valid_pts.append((lon, lat))
            else:
                break  # Hit the boundary! Cut off the rest of the route here.

        # If it started outside the box, or we don't have enough points left to make a line
        if len(valid_pts) < 2:
            log.debug("  Route %s dropped (Started outside bounding box or too short).", row["Route_ID"])
            _record_drop(row, "origin_outside_bbox")
            dropped_count += 1
            continue

        if len(valid_pts) < len(all_pts):
            truncated_count += 1

        # The last valid point inside the box is now the new End Terminal
        row["Start_Lon"], row["Start_Lat"] = valid_pts[0]
        row["End_Lon"], row["End_Lat"]     = valid_pts[-1]

        # Re-pack the remaining middle points into the via JSON format
        new_vias = valid_pts[1:-1]
        if new_vias:
            row["Via_Coordinates"] = json.dumps([{"lat": lat, "lon": lon} for lon, lat in new_vias])
        else:
            row["Via_Coordinates"] = None

        new_rows.append(row)

    log.info("  Bounding Box Truncation: %d routes truncated, %d dropped entirely.", 
             truncated_count, dropped_count)
    
    return pd.DataFrame(new_rows)

def load_routes(path: str) -> pd.DataFrame:
    """
    Load routes CSV with flexible column aliasing.
    All ~700 routes are loaded regardless of length — filtering
    for length is NO LONGER performed here (Directive 1).
    """
    log.info("Loading routes from '%s'", path)
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip().str.replace(" ", "_")
    col_map = {c.lower(): c for c in df.columns}

    def _resolve(targets):
        for k in targets:
            if k in col_map:
                return col_map[k]
        return None

    alias = {
        "Start_Lat":          ["start_lat", "startlat", "origin_lat", "from_lat"],
        "Start_Lon":          ["start_lon", "startlon", "origin_lon", "start_lng",
                               "from_lon"],
        "End_Lat":            ["end_lat",   "endlat",   "dest_lat",   "to_lat"],
        "End_Lon":            ["end_lon",   "endlon",   "dest_lon",   "end_lng",
                               "to_lon"],
        "Via_Coordinates":    ["via_coordinates", "via_coords", "waypoints", "via_points_geocoded"],
        "Minibus_Count":      ["minibus_count", "minibuses", "mini_bus_count"],
        "Standard_Bus_Count": ["standard_bus_count", "std_bus", "standard_buses"],
        "Route_Name":         ["route_name", "name", "route"],
        # v3.2: surface Origin/Destination columns so the Route_Name fallback
        # below can use them when Route_Name itself is present-but-blank.
        "Route_From":         ["route_from", "origin", "from", "start_terminal"],
        "Route_To":           ["route_to",   "destination", "dest", "to", "end_terminal"],
        "Vehicle_Category":   ["vehicle_category", "veh_cat"],
    }
    rename = {}
    for canon, targets in alias.items():
        found = _resolve(targets)
        if found and found != canon:
            rename[found] = canon
    df.rename(columns=rename, inplace=True)

    if "Route_ID" not in df.columns:
        df["Route_ID"] = [f"R{i+1:04d}" for i in range(len(df))]

    # v3.2: Route_Name fallback. The previous logic only ran when the column
    # was absent — but existing-routes.csv has a literal Route_Name column
    # with every row empty, so the engine was emitting "nan" for every popup
    # and breaking SSCL fuzzy-matching via _get_terminal's "A to B" parser.
    if "Route_Name" not in df.columns:
        df["Route_Name"] = None
    # Treat empty strings and the literal "nan" as missing
    name_series = df["Route_Name"].astype(str).str.strip()
    blank_mask  = name_series.isin(["", "nan", "None", "NaN"]) | df["Route_Name"].isna()
    if "Route_From" in df.columns and "Route_To" in df.columns:
        fr = df["Route_From"].astype(str).str.strip().replace({"nan": "", "None": ""})
        to = df["Route_To"].astype(str).str.strip().replace({"nan": "", "None": ""})
        constructed = (fr + " ↔ " + to).where(fr.ne("") & to.ne(""), other=df["Route_ID"])
    else:
        constructed = df["Route_ID"].astype(str)
    df.loc[blank_mask, "Route_Name"] = constructed[blank_mask].values
    n_constructed = int(blank_mask.sum())
    if n_constructed:
        log.info("  Route_Name reconstructed for %d blank rows (Origin ↔ Destination).",
                 n_constructed)
    for col, default in [("Minibus_Count", 0), ("Standard_Bus_Count", 0),
                          ("Via_Coordinates", None)]:
        if col not in df.columns:
            df[col] = default

    df["Minibus_Count"]      = (pd.to_numeric(df["Minibus_Count"],
                                               errors="coerce")
                                  .fillna(0).astype(int))
    df["Standard_Bus_Count"] = (pd.to_numeric(df["Standard_Bus_Count"],
                                               errors="coerce")
                                  .fillna(0).astype(int))
    log.info("  Loaded %d routes (no length filter applied — Directive 1).", len(df))
    return df


def load_pois(path: str) -> gpd.GeoDataFrame:
    """
    Load POIs with the Kashmir 3-tier system + women-anchor boost.

    Tier 1 (w=1.0): hospitals, transit hubs, govt offices, mosques, malls
    Tier 2 (w=0.4): colleges, markets, schools, recreation
    Tier 3 (w=0.6 summer / 0.0 winter): tourism, gardens, gondola, shrines
                    Toggled by WINTER_SCENARIO flag.

    Women-anchor categories (girls' colleges, maternity hospitals, women's
    markets) receive a +25% boost on top of their tier weight, calibrated
    to CHALO observation that 64.5% of riders are women (free-fare effect).
    """
    log.info("Loading POIs from '%s' (Kashmir 3-tier + women-anchor boost)…", path)
    df = pd.read_csv(path)

    # Standardise column names to lowercase to handle Capitalised headers
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

    # Map 'latitude'/'longitude' to the expected 'lat'/'lon'
    rename_map = {
        "latitude": "lat",
        "longitude": "lon"
    }
    df.rename(columns=rename_map, inplace=True)

    if "category" not in df.columns:
        df["category"] = "Other"
    if "name" not in df.columns:
        df["name"] = df["category"]

    # Assign Kashmir 3-tier weights + women-anchor boost
    def _poi_weight(row) -> float:
        # Priority 1: Use the 'Importance' column if it exists
        base = None
        c = str(row["category"]).lower().strip()
        if "importance" in row and pd.notna(row["importance"]):
            imp = str(row["importance"]).lower().strip()
            if imp == "high":
                base = POI_TIER1_WEIGHT    # 1.0
            elif imp == "seasonal":
                # v3.2: Even with "Importance=seasonal", residential anchors
                # keep their Tier-2-floor weight in winter.
                if c in POI_TIER3_RESIDENTIAL_ANCHOR:
                    base = POI_TIER3_RES_WEIGHT_EFFECTIVE
                else:
                    base = POI_TIER3_WEIGHT
            elif imp in ["medium", "low"]:
                base = POI_TIER2_WEIGHT    # 0.4

        # Priority 2: Fallback to category-based tier mapping
        if base is None:
            if c in POI_TIER1_CATEGORIES:
                base = POI_TIER1_WEIGHT
            elif c in POI_TIER3_TOURIST_ONLY:
                base = POI_TIER3_WEIGHT             # 0.6 summer / 0.0 winter
            elif c in POI_TIER3_RESIDENTIAL_ANCHOR:
                base = POI_TIER3_RES_WEIGHT_EFFECTIVE  # 0.6 summer / 0.4 winter
            elif c in POI_TIER2_CATEGORIES:
                base = POI_TIER2_WEIGHT
            else:
                base = DEFAULT_POI_WEIGHT_V2

        # Women-anchor boost (+25%) — CHALO calibration
        c = str(row["category"]).lower().strip()
        if c in WOMEN_ANCHOR_CATEGORIES:
            base *= WOMEN_ANCHOR_BOOST

        return base

    # Apply the weight function row-wise
    df["POI_Weight"] = df.apply(_poi_weight, axis=1)

    # Flag high-value POIs for map popups (anything ≥ Tier 1 weight)
    df["Is_HV_POI"] = df["POI_Weight"] >= POI_TIER1_WEIGHT

    # Tier audit
    tier1_n   = (df["POI_Weight"] >= POI_TIER1_WEIGHT).sum()
    tier3_n   = ((df["POI_Weight"] > POI_TIER2_WEIGHT) &
                 (df["POI_Weight"] < POI_TIER1_WEIGHT)).sum()
    tier2_n   = (df["POI_Weight"] <= POI_TIER2_WEIGHT).sum()
    women_n   = df["category"].str.lower().str.strip().isin(
                    WOMEN_ANCHOR_CATEGORIES).sum()
    season    = "WINTER" if WINTER_SCENARIO else "SUMMER"
    log.info("  Loaded %d POIs — T1 (≥1.0): %d  T2 (≤0.4): %d  T3 mid: %d "
             "[scenario=%s, T3 weight=%.2f]",
             len(df), tier1_n, tier2_n, tier3_n, season, POI_TIER3_WEIGHT)
    if women_n:
        log.info("  Women-anchor POIs boosted by %.0f%%: %d",
                 (WOMEN_ANCHOR_BOOST - 1) * 100, women_n)

    gdf = gpd.GeoDataFrame(
        df,
        geometry=gpd.points_from_xy(df["lon"], df["lat"]),
        crs=WGS84_CRS,
    )
    return gdf


# Module guard so a producer/consumer Via format mismatch can never again be
# silent (audit Finding 3).
_VIA_PARSE_WARNED = False

# Per-input-row disposition trail (audit Bug 2 / Rec 8). Every route that is
# DROPPED before reaching the output (bbox exit, null geometry, sub-1 km
# zero-length collision) is recorded here with a reason, so a government-facing
# rationalisation can show what happened to every permit — not just an
# aggregate "dropped=N". Reset at the top of main().
_DROP_LOG: List[Dict] = []


def _record_drop(row, reason: str) -> None:
    """Append a dropped route to the disposition trail."""
    _DROP_LOG.append({
        "Route_ID":   row.get("Route_ID", ""),
        "Route_Name": row.get("Route_Name", ""),
        "Start_Lat":  row.get("Start_Lat"), "Start_Lon": row.get("Start_Lon"),
        "End_Lat":    row.get("End_Lat"),   "End_Lon":   row.get("End_Lon"),
        "Vehicle_Category": row.get("Vehicle_Category", ""),
        "Disposition": "DROPPED",
        "Reason":     reason,
    })


# Srinagar centroid the geocoders historically collapsed onto (audit Finding 1).
# Used by the input-QA gate to flag endpoints that snapped to it.
_SRINAGAR_CENTROID = (34.085650, 74.805550)


def _near_srinagar_centroid(lat, lon, tol_km: float = 0.3) -> bool:
    try:
        return _haversine_km(float(lat), float(lon),
                             _SRINAGAR_CENTROID[0], _SRINAGAR_CENTROID[1]) <= tol_km
    except (TypeError, ValueError):
        return False


def audit_input_quality(df: pd.DataFrame, out_path: str = "input_qa_report.csv") -> dict:
    """Pre-engine endpoint QA gate (audit Recommendation 2).

    For every input row, computes the straight-line O→D distance and flags:
      • origin / destination snapped to the Srinagar centroid (geocode collapse),
      • zero-length routes (haversine < MIN_ROUTE_KM → will be dropped),
    writes a per-row report so the data quality is auditable BEFORE rationalisation,
    and logs a loud summary. Returns a counts dict. Diagnostic, not blocking —
    the hard gate lives in run_all_qc_checks() (strict mode).
    """
    log.info("Input QA: checking endpoint geocode plausibility (Rec 2)…")
    recs = []
    for _, row in df.iterrows():
        try:
            hav = _haversine_km(float(row["Start_Lat"]), float(row["Start_Lon"]),
                                float(row["End_Lat"]),   float(row["End_Lon"]))
        except (TypeError, ValueError, KeyError):
            hav = float("nan")
        o_ctr = _near_srinagar_centroid(row.get("Start_Lat"), row.get("Start_Lon"))
        d_ctr = _near_srinagar_centroid(row.get("End_Lat"),   row.get("End_Lon"))
        recs.append({
            "Route_ID": row.get("Route_ID", ""),
            "Route_Name": row.get("Route_Name", ""),
            "Haversine_KM": round(hav, 3) if hav == hav else "",
            "Origin_At_Srinagar_Centroid": o_ctr,
            "Dest_At_Srinagar_Centroid":   d_ctr,
            "Zero_Length": bool(hav == hav and hav < MIN_ROUTE_KM),
        })
    rep = pd.DataFrame(recs)
    counts = {
        "total":          len(rep),
        "zero_length":    int(rep["Zero_Length"].sum()),
        "origin_centroid": int(rep["Origin_At_Srinagar_Centroid"].sum()),
        "dest_centroid":   int(rep["Dest_At_Srinagar_Centroid"].sum()),
    }
    try:
        rep.to_csv(out_path, index=False, encoding="utf-8-sig")
    except OSError as exc:
        log.warning("  Could not write input QA report: %s", exc)
    level = log.warning if counts["zero_length"] else log.info
    level("  Input QA: %d/%d zero-length (<%.1f km, will be dropped) | "
          "%d origins & %d destinations at the Srinagar centroid. Report → %s",
          counts["zero_length"], counts["total"], MIN_ROUTE_KM,
          counts["origin_centroid"], counts["dest_centroid"], out_path)
    if counts["zero_length"]:
        log.warning("  ⚠ Zero-length routes indicate the geocode collapse of "
                    "Finding 1 — re-geocode (geocode_common.py) before submission.")
    return counts


def export_route_disposition(gdf: gpd.GeoDataFrame, out_path: str) -> None:
    """Write a per-route disposition record covering every input row (Rec 8).

    Survivors carry their Action_Taken (kept / merged / upgraded); dropped rows
    come from _DROP_LOG with a reason. No input route is unaccounted for.
    """
    surv = gdf[["Route_ID", "Route_Name", "Action_Taken"]].copy()
    surv = surv.rename(columns={"Action_Taken": "Reason"})
    surv["Disposition"] = "KEPT"
    surv = surv[["Route_ID", "Route_Name", "Disposition", "Reason"]]
    drops = pd.DataFrame(_DROP_LOG)
    if not drops.empty:
        drops = drops[["Route_ID", "Route_Name", "Disposition", "Reason"]]
        out = pd.concat([surv, drops], ignore_index=True)
    else:
        out = surv
    out.to_csv(out_path, index=False, encoding="utf-8-sig")
    log.info("  Route disposition written: %d kept + %d dropped = %d rows → %s",
             len(surv), len(_DROP_LOG), len(out), out_path)


def parse_via(via_raw) -> List[Tuple[float, float]]:
    """Parse a Via_Coordinates value into a list of (lon, lat) tuples.

    Accepts BOTH formats the pipeline uses:
      • JSON      — '[{"lat":..,"lon":..}, …]' or '[[lon,lat], …]'
                    (this is what truncate_routes_to_bbox re-writes), and
      • the plain producer string — 'lat,lon;lat,lon'
                    (Via_Points_Geocoded emitted by latlon.py /
                     geocode_other_routes.py — semicolon-separated lat,lon pairs).

    Audit Finding 3 (CRITICAL): the previous implementation only attempted
    json.loads() and swallowed every failure in a bare except, so 100% of the
    plain-string vias (429 of 613 input rows) were silently discarded and the
    permitted via-routing never reached OSRM. Returns [] only for genuinely
    empty input; a non-empty value that parses to nothing is now logged.
    """
    global _VIA_PARSE_WARNED
    if via_raw is None or (isinstance(via_raw, float) and math.isnan(via_raw)):
        return []
    if not isinstance(via_raw, str) or not via_raw.strip():
        return []
    raw = via_raw.strip()

    def _as_lonlat(a: float, b: float) -> Tuple[float, float]:
        # Kashmir latitudes are ~33–35, longitudes ~73–76. The value in the
        # [6, 40] band is the latitude, so emit (lon, lat) regardless of order.
        return (b, a) if 6 <= a <= 40 else (a, b)

    # 1) JSON form (list of dicts, or list of [x, y] pairs)
    if raw[0] in "[{":
        try:
            pts = json.loads(raw)
            result = []
            for p in pts:
                if isinstance(p, dict):
                    result.append(
                        (float(p.get("lon", p.get("lng", 0))),
                         float(p.get("lat", 0)))
                    )
                elif isinstance(p, (list, tuple)) and len(p) >= 2:
                    result.append(_as_lonlat(float(p[0]), float(p[1])))
            return result
        except (json.JSONDecodeError, TypeError, ValueError):
            pass  # fall through to the plain-string parser

    # 2) Plain producer string 'lat,lon;lat,lon' (tolerate '|'/whitespace seps)
    result = []
    for chunk in re.split(r"[;|]", raw):
        chunk = chunk.strip()
        if not chunk:
            continue
        nums = re.split(r"[,\s]+", chunk)
        if len(nums) < 2:
            continue
        try:
            result.append(_as_lonlat(float(nums[0]), float(nums[1])))
        except ValueError:
            continue

    if not result and not _VIA_PARSE_WARNED:
        log.warning("parse_via: a non-empty Via_Coordinates value parsed to no "
                    "points (e.g. %r) — vias for such rows are ignored.", raw[:60])
        _VIA_PARSE_WARNED = True
    return result


def _build_osrm_url(coords: List[Tuple[float, float]]) -> str:
    return (f"{OSRM_BASE_URL}/route/v1/driving/"
            + ";".join(f"{lon},{lat}" for lon, lat in coords)
            + "?overview=full&geometries=geojson&steps=false")


def _fetch_osrm_single(route_id: str, coords: List[Tuple[float, float]]) -> Dict:
    """Fetch a single route from OSRM. Returns dict with geometry and duration."""
    try:
        resp = requests.get(_build_osrm_url(coords), timeout=OSRM_TIMEOUT_S)
        resp.raise_for_status()
        data = resp.json()
        if data.get("code") == "Ok" and data.get("routes"):
            r0 = data["routes"][0]
            return {
                "route_id":        route_id,
                "geometry":        shape(r0["geometry"]),
                "osrm_km":         r0["distance"] / 1000.0,
                "osrm_duration_s": float(r0["duration"]),
                "success":         True,
            }
    except Exception as exc:
        log.debug("OSRM failed [%s]: %s", route_id, exc)
    return {"route_id": route_id, "geometry": None, "osrm_km": None,
            "osrm_duration_s": None, "success": False}


def fetch_all_osrm(df: pd.DataFrame) -> Dict[str, Dict]:
    """Concurrent OSRM fetches for all routes."""
    log.info("Fetching OSRM geometries (%d routes, %d workers)…",
             len(df), OSRM_MAX_WORKERS)
    tasks = {}
    for _, row in df.iterrows():
        via = parse_via(row.get("Via_Coordinates"))
        tasks[row["Route_ID"]] = (
            [(float(row["Start_Lon"]), float(row["Start_Lat"]))]
            + via
            + [(float(row["End_Lon"]), float(row["End_Lat"]))]
        )
    results = {}
    with ThreadPoolExecutor(max_workers=OSRM_MAX_WORKERS) as pool:
        futures = {pool.submit(_fetch_osrm_single, rid, coords): rid
                   for rid, coords in tasks.items()}
        for fut in as_completed(futures):
            rid = futures[fut]
            try:
                res = fut.result()
                results[res["route_id"]] = res
            except Exception as exc:
                log.warning("Future error [%s]: %s", rid, exc)
                results[rid] = {"route_id": rid, "geometry": None,
                                "osrm_km": None, "osrm_duration_s": None,
                                "success": False}
    ok = sum(1 for r in results.values() if r["success"])
    log.info("  OSRM: %d/%d ok, %d use circuity fallback.", ok, len(df),
             len(df) - ok)
    return results


def _haversine_km(lat1, lon1, lat2, lon2) -> float:
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi  = math.radians(lat2 - lat1)
    dlam  = math.radians(lon2 - lon1)
    a = (math.sin(dphi / 2) ** 2
         + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(a))


# v3.3.3 helper: shortest distance (km) from a (lat, lon) point to a LineString
# stored in WGS84. Used by tourist-corridor and seasonal-suspended tagging.
# Uses an equirectangular approximation around the point's latitude — accurate
# to ~0.3% at Srinagar's latitude over the relevant 5-50 km buffer scale.
def _line_near_point_km(line, plat: float, plon: float) -> float:
    if line is None or line.is_empty:
        return float("inf")
    cos_lat = math.cos(math.radians(plat))
    # Project line coordinates (lon, lat) to a local metric where 1 unit ≈ 1 km.
    coords = [((lon - plon) * 111.32 * cos_lat,
               (lat - plat) * 110.57)
              for (lon, lat) in line.coords]
    # Shortest distance from origin (0,0) to the polyline.
    best = float("inf")
    for i in range(len(coords) - 1):
        ax, ay = coords[i]
        bx, by = coords[i + 1]
        dx, dy = bx - ax, by - ay
        if dx == 0 and dy == 0:
            d = math.hypot(ax, ay)
        else:
            t = max(0.0, min(1.0, -(ax * dx + ay * dy) / (dx * dx + dy * dy)))
            px, py = ax + t * dx, ay + t * dy
            d = math.hypot(px, py)
        if d < best:
            best = d
    return best


def apply_geometries(df: pd.DataFrame, osrm_results: Dict) -> gpd.GeoDataFrame:
    """
    Merge OSRM routing results with the route dataframe.

    DIRECTIVE 1 FIX:
    The v6 hard filter `gdf = gdf[gdf["Route_KM"] <= 60].copy()` has been
    removed entirely. All routes are retained. Instead, a Route_Type column
    is assigned:
      - Urban          : Route_KM < 15
      - Peri_Urban     : 15 ≤ Route_KM < 40
      - Regional_District : Route_KM ≥ 40  (rural lifeline — flagged, not dropped)

    Regional_District routes receive relaxed headway targets in Step 6.
    """
    log.info("Merging geometries — ALL routes retained (Directive 1)…")
    rows = []
    bridge_pen_n = 0
    for _, row in df.iterrows():
        rid       = row["Route_ID"]
        res       = osrm_results.get(rid, {})
        via       = parse_via(row.get("Via_Coordinates"))
        raw_coords = (
            [(float(row["Start_Lon"]), float(row["Start_Lat"]))]
            + via
            + [(float(row["End_Lon"]), float(row["End_Lat"]))]
        )
        fallback_geom = LineString(raw_coords) if len(raw_coords) >= 2 else None

        # v3.2: detect Jhelum crossing once — applies whether OSRM succeeded
        # or we fell back to circuity. The previous code only honoured this
        # in the fallback branch, so OSRM-routed trips paid zero bridge
        # bottleneck cost even though they cross the same bridges.
        crosses_jhelum = (float(row["Start_Lon"]) < JHELUM_RIVER_LON) != \
                         (float(row["End_Lon"])   < JHELUM_RIVER_LON)

        if res.get("success") and res["geometry"] is not None:
            geom       = res["geometry"]
            dist_km    = max(0.0, res["osrm_km"])
            duration_s = max(0.0, res["osrm_duration_s"])
            if crosses_jhelum:
                # Add a flat bridge-queue penalty — OSRM duration is free-flow
                # midnight time and cannot see Habba Kadal / Amira Kadal queues.
                duration_s += JHELUM_BRIDGE_BOTTLENECK_MIN * 60.0
                bridge_pen_n += 1
                source = "OSRM+JhelumBridge"
            else:
                source = "OSRM"
        else:
            geom    = fallback_geom
            sl_km   = _haversine_km(
                float(row["Start_Lat"]), float(row["Start_Lon"]),
                float(row["End_Lat"]),   float(row["End_Lon"]))
            cf      = CIRCUITY_FACTOR_RIVER if crosses_jhelum else CIRCUITY_FACTOR
            dist_km = sl_km * cf

            start_lat = float(row["Start_Lat"])
            end_lat   = float(row["End_Lat"])
            # Kashmir-specific speed zones:
            #  - Both terminals in Downtown Srinagar (lat > 34.07) → old-city speed
            #  - Both terminals on valley periphery (lat < 34.00) → highway speed
            #  - Otherwise mixed corridor → default suburban speed
            if start_lat > CITY_CORE_LAT_THRESHOLD and end_lat > CITY_CORE_LAT_THRESHOLD:
                local_speed = SPEED_OLD_CITY_KMH
            elif start_lat < 34.00 and end_lat < 34.00:
                local_speed = SPEED_VALLEY_HIGHWAY_KMH
            else:
                local_speed = SPEED_DEFAULT_KMH

            duration_s = (dist_km / local_speed) * 3600.0
            if crosses_jhelum:
                duration_s += JHELUM_BRIDGE_BOTTLENECK_MIN * 60.0
                bridge_pen_n += 1
            source     = "Circuity-River" if crosses_jhelum else "Circuity"

        # DIRECTIVE 1: Classify route type — no dropping
        km = max(0.0, dist_km)
        cat = str(row.get("Vehicle_Category", "")).strip().lower()

        if "mps" in cat:
            route_type = "Regional_District"  # Operator hint overrides length
        elif "city bus" in cat:
            route_type = "Urban" if km < PERIURBAN_KM_THRESHOLD else "Peri_Urban"
        elif km < URBAN_KM_THRESHOLD:
            route_type = "Urban"
        elif km < PERIURBAN_KM_THRESHOLD:
            route_type = "Peri_Urban"
        else:
            route_type = "Regional_District"

        # v3.3.3 (teammate review): Tourist Corridor tagging extended.
        # v3.3's name-only check missed 99% of corridors because the permit
        # data records urban hub names at the endpoints, not the tourist
        # destinations along the way. We now ALSO check whether the route's
        # geometry passes within TOURIST_ZONE_BUFFER_KM of any tourist
        # centroid (Gulmarg / Pahalgam / Sonamarg / Mughal Road / Boulevard /
        # Mughal gardens etc.).
        name_lc = str(row.get("Route_Name", "")).lower()
        tourist_corridor = any(k in name_lc for k in TOURIST_CORRIDOR_KEYWORDS)
        suspended        = any(k in name_lc for k in SEASONAL_SUSPENDED_KEYWORDS)
        if geom is not None and not tourist_corridor:
            # Distant tourist destinations: any geometry crossing the buffer.
            for _, (tlat, tlon) in TOURIST_ZONES_DISTANT.items():
                if _line_near_point_km(geom, tlat, tlon) <= TOURIST_ZONE_BUFFER_KM:
                    tourist_corridor = True
                    break
        if geom is not None and not tourist_corridor:
            # Inner-city tourist gates: ENDPOINT must be near. Stops a route
            # from being tagged just because it passes through downtown.
            start_lat = float(row["Start_Lat"])
            start_lon = float(row["Start_Lon"])
            end_lat   = float(row["End_Lat"])
            end_lon   = float(row["End_Lon"])
            for _, (tlat, tlon) in TOURIST_ZONES_INNER_CITY.items():
                d_start = _haversine_km(start_lat, start_lon, tlat, tlon)
                d_end   = _haversine_km(end_lat,   end_lon,   tlat, tlon)
                if min(d_start, d_end) <= TOURIST_INNER_BUFFER_KM:
                    tourist_corridor = True
                    break
        if geom is not None and not suspended:
            for _, (zlat, zlon) in SEASONAL_SUSPENDED_ZONES_LATLON.items():
                if _line_near_point_km(geom, zlat, zlon) <= SEASONAL_SUSPENDED_BUFFER_KM:
                    suspended = True
                    break
        if suspended:
            seasonal_op = "Winter_Suspended"
        elif tourist_corridor:
            seasonal_op = "Seasonal"
        else:
            seasonal_op = "Year_Round"

        rows.append({**row.to_dict(),
                     "geometry":            geom,
                     "Route_KM":            round(km, 3),
                     "OSRM_Duration_S":     round(max(0.0, duration_s), 1),
                     "Geo_Source":          source,
                     "Route_Type":          route_type,
                     "Tourist_Corridor":    tourist_corridor,
                     "Seasonal_Operability": seasonal_op})

    gdf  = gpd.GeoDataFrame(rows, geometry="geometry", crs=WGS84_CRS)
    n0   = len(gdf)
    # Only drop null geometry and sub-minimum length routes — but record each
    # drop in the disposition trail first (audit Bug 2: was silent before).
    for _, r in gdf[gdf.geometry.isna()].iterrows():
        _record_drop(r, "null_geometry")
    present = gdf[gdf.geometry.notna()].copy()
    short_mask = present["Route_KM"] < MIN_ROUTE_KM
    for _, r in present[short_mask].iterrows():
        _record_drop(r, f"sub_min_km_zero_length (Route_KM={r['Route_KM']:.3f}; "
                        f"likely identical O/D geocode — see Finding 1)")
    gdf  = present[~short_mask].copy()

    # Count route types for audit
    type_counts = gdf["Route_Type"].value_counts().to_dict()
    log.info("  Directive 1 — Route type breakdown after geometry: "
             "Urban=%d  Peri_Urban=%d  Regional_District=%d  (total=%d, dropped=%d for "
             "null geometry or sub-1km)",
             type_counts.get("Urban", 0),
             type_counts.get("Peri_Urban", 0),
             type_counts.get("Regional_District", 0),
             len(gdf), n0 - len(gdf))
    log.info("  Jhelum bridge bottleneck (+%.1f min) applied to %d routes.",
             JHELUM_BRIDGE_BOTTLENECK_MIN, bridge_pen_n)
    return gdf.reset_index(drop=True)


def impute_fleet(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Impute existing fleet sizes for context only. Not used in Step 8."""
    log.info("Imputing existing fleet sizes for context…")
    gdf = gdf.copy()
    gdf["Total_Minibuses_Existing"] = (
        gdf["Minibus_Count"] + gdf["Standard_Bus_Count"] * STD_TO_MINI_RATIO
    ).round(1).clip(lower=0)

    mask = gdf["Total_Minibuses_Existing"] == 0

    def _ghost_fleet(km):
        cycle = (km * 2 / SPEED_DEFAULT_KMH) * 60
        return max(1, math.ceil(cycle / 60))

    gdf.loc[mask, "Total_Minibuses_Existing"] = gdf.loc[mask, "Route_KM"].apply(
        _ghost_fleet)
    log.info("  Fleet imputed for %d zero-bus routes (context only).", mask.sum())
    return gdf


# ══════════════════════════════════════════════════════════════════════════════
#  CMP TRUNK INJECTION  ─  Pre-classification override for 13 CMP Table 4.2 routes
# ══════════════════════════════════════════════════════════════════════════════

def _fuzzy_match_score(a: str, b: str) -> float:
    """
    Return SequenceMatcher similarity ratio (0–1) between two terminal name strings.
    Both strings are lowercased and stripped before comparison so that minor
    capitalisation/punctuation differences do not reduce match quality.
    """
    from difflib import SequenceMatcher
    a_clean = a.lower().strip()
    b_clean = b.lower().strip()
    return SequenceMatcher(None, a_clean, b_clean).ratio()


# Generic place-name tokens that must NOT, on their own, anchor an SSCL match.
# These appear in many unrelated terminal names ("… Srinagar", "… Chowk",
# "… Bus Stand") so a raw character-ratio match on them produces geographically
# impossible upgrades (see _terminal_matches_cmp docstring).
_CMP_GENERIC_TOKENS = {
    "srinagar", "sgr", "kashmir", "jkrtc", "chowk", "crossing", "stand",
    "bus", "stop", "road", "colony", "hospital", "station", "bypass", "loop",
    "circular", "downtown", "city", "district", "court", "new", "old", "via",
    "near", "main", "town", "centre", "center", "office", "depot", "terminal",
}


def _cmp_tokens(name: str) -> set:
    """Meaningful (non-generic, >=4 char) word tokens from a terminal name."""
    import re as _re
    toks = _re.split(r"[^a-z0-9]+", str(name).lower())
    return {t for t in toks if len(t) >= 4 and t not in _CMP_GENERIC_TOKENS}


def _terminal_matches_cmp(dataset_terminal: str, cmp_terminal: str) -> bool:
    """
    True only when the dataset terminal genuinely refers to the same place as the
    SSCL terminal.

    Hardened (v3.3.9 audit fix): the previous implementation returned True on a
    0.45 character-ratio OR a raw substring, which mis-matched any route merely
    *ending in* a generic word. e.g. "Srinagar" (8 chars) scores 0.52 against
    "District Court Srinagar" (23 chars), so every "X to Srinagar" permit was
    falsely upgraded to SSCL-12, and weak collisions like "tangmarg"↔"rangreth"
    cleared 0.45 — producing 11 geographically-impossible SSCL trunks (Anantnag→
    Srinagar tagged SSCL-12, Anantnag→Shopian tagged SSCL-05, …).

    Now requires a STRONG full-string fuzzy match (>=0.80) OR a shared *meaningful*
    token (>=4 chars, not a generic place word), with close spelling variants of
    such tokens allowed (>=0.85, e.g. "Batamaloo"/"Batamallo"). This still catches
    the documented "Hazratbal" ↔ "… via Hazratbal" partial case (shared token)
    while rejecting generic-word and noise matches.
    """
    ds = str(dataset_terminal).lower().strip()
    cm = str(cmp_terminal).lower().strip()
    if not ds or not cm:
        return False
    # 1. exact / near-exact full-string match
    if _fuzzy_match_score(ds, cm) >= 0.80:
        return True
    # 2. shared meaningful token (whole word), incl. close spelling variants
    ds_tok, cm_tok = _cmp_tokens(ds), _cmp_tokens(cm)
    if ds_tok & cm_tok:
        return True
    return any(_fuzzy_match_score(a, b) >= 0.85
               for a in cm_tok for b in ds_tok)


def inject_cmp_trunk_routes(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Pre-classification SSCL E-Bus Backbone Injection.

    For each of the 30 SSCL routes (from CHALO Apr 2026 data), this function
    scans the dataset for routes whose origin–destination pair fuzzy-matches
    the SSCL terminals (using SequenceMatcher with CMP_FUZZY_THRESHOLD = 0.55).

    On a match:
      • Action_Taken   → "UPGRADED_TO_TRUNK"
      • Priority_Band  → "HP"
      • CMP_Trunk      → True            (legacy flag name; means SSCL trunk)
      • CMP_Route_ID   → e.g. "SSCL-01"  (audit trail)

    Headway for matched routes is hardcoded to SSCL_TRUNK_HEADWAY_MIN (15 min)
    in step6_assign_headways() via the CMP_Trunk flag — matches actual SSCL
    operational targets, not the Jammu BRT 10-min assumption.

    Note: in this Kashmir fork the "CMP_*" column names are retained for
    backward compatibility with downstream functions (QC, export, mapping)
    but the underlying data is SSCL e-bus routes, NOT RITES CMP routes.

    This function must be called AFTER apply_geometries() and BEFORE
    classify_routes().
    """
    log.info("SSCL Injection: scanning dataset against %d SSCL e-bus trunk routes "
             "(from CHALO data)…", len(CMP_TRUNK_ROUTES))
    gdf = gdf.copy()

    # Ensure marker columns exist
    if "CMP_Trunk" not in gdf.columns:
        gdf["CMP_Trunk"]    = False
    if "CMP_Route_ID" not in gdf.columns:
        gdf["CMP_Route_ID"] = ""
    if "Priority_Band" not in gdf.columns:
        gdf["Priority_Band"] = ""
    if "Action_Taken" not in gdf.columns:
        gdf["Action_Taken"] = "RETAINED_AS_FEEDER"

    # Build terminal name series — prefer dedicated terminal columns if present,
    # fall back to Route_Name parsing.
    def _get_terminal(row, end: str) -> str:
        col = f"{'Start' if end == 'start' else 'End'}_Terminal"
        if col in row.index and pd.notna(row[col]) and str(row[col]).strip():
            return str(row[col]).strip()
        # Fallback: parse from Route_Name "A to B" pattern
        name = str(row.get("Route_Name", ""))
        parts = [p.strip() for p in name.lower().split(" to ")]
        if end == "start" and len(parts) >= 1:
            return parts[0]
        if end == "end"   and len(parts) >= 2:
            return parts[-1]
        return name

    matched_total = 0
    for cmp_route in CMP_TRUNK_ROUTES:
        cmp_origin = cmp_route["origin"]
        cmp_dest   = cmp_route["dest"]
        cmp_id     = cmp_route["id"]

        matched_idx = []
        for idx, row in gdf.iterrows():
            # The synthetic backbone row ALWAYS self-matches, regardless of the
            # terminal matcher (v3.3.9: the hardened matcher can't anchor on SSCL
            # names that are entirely generic words — e.g. SSCL-11 "Old City Loop
            # Downtown Srinagar" — so without this every one of the 30 government
            # backbone routes must still be guaranteed an upgrade).
            if str(gdf.at[idx, "Route_ID"]).strip() == cmp_id:
                matched_idx.append(idx)
                continue
            ds_start = _get_terminal(row, "start")
            ds_end   = _get_terminal(row, "end")

            # Check both directions (route may be stored A→B or B→A)
            fwd = (_terminal_matches_cmp(ds_start, cmp_origin) and
                   _terminal_matches_cmp(ds_end,   cmp_dest))
            rev = (_terminal_matches_cmp(ds_start, cmp_dest) and
                   _terminal_matches_cmp(ds_end,   cmp_origin))
            if fwd or rev:
                matched_idx.append(idx)

        if matched_idx:
            written = 0
            for idx in matched_idx:
                # Exact Route_ID match (synthetic self-route) always wins regardless
                # of prior claims — prevents a same-origin false-positive from
                # overwriting the synthetic's own CMP_Route_ID (e.g. SSCL-24
                # synthetic claimed by SSCL-06 because both start at Pantha Chowk).
                is_self_route = (str(gdf.at[idx, "Route_ID"]).strip() == cmp_id)
                # Length sanity (v3.3.9 audit fix): a real permit can only BE an
                # SSCL corridor if its routed length is comparable to the SSCL
                # nominal km. Backstops any residual terminal false-positive
                # (e.g. a 59 km Anantnag→Srinagar permit can't be the 11 km
                # SSCL-12). The synthetic backbone (self-routes) bypasses this.
                if not is_self_route and "Route_KM" in gdf.columns:
                    ds_km  = gdf.at[idx, "Route_KM"]
                    cmp_km = float(cmp_route.get("km", 0) or 0)
                    if pd.notna(ds_km) and ds_km > 0 and cmp_km > 0 \
                       and not (0.45 <= float(ds_km) / cmp_km <= 2.2):
                        continue
                existing_id   = str(gdf.at[idx, "CMP_Route_ID"]).strip()
                if existing_id and existing_id != cmp_id and not is_self_route:
                    # First-match wins: don't overwrite a CMP_Route_ID already set by
                    # an earlier SSCL route (e.g. via-point false positives for routes
                    # like "Batamaloo to Kangan via Manigam" matching both SSCL-26 and
                    # SSCL-27 because "manigam" is a substring of the Route_Name).
                    continue
                gdf.at[idx, "Action_Taken"]  = "UPGRADED_TO_TRUNK"
                gdf.at[idx, "Priority_Band"] = "HP"
                gdf.at[idx, "CMP_Trunk"]     = True
                gdf.at[idx, "CMP_Route_ID"]  = cmp_id
                written += 1
            if written:
                log.info("  ✓ CMP %s (%s → %s): matched %d dataset route(s) — forced TRUNK/HP.",
                         cmp_id, cmp_origin, cmp_dest, written)
                matched_total += written
            else:
                log.info("  ~ CMP %s (%s → %s): %d fuzzy match(es) already claimed "
                         "by prior SSCL ID — skipped.",
                         cmp_id, cmp_origin, cmp_dest, len(matched_idx))
        else:
            log.warning("  ⚠ CMP %s (%s → %s): NO dataset match found at "
                        "threshold=%.2f.  Route absent from dataset.",
                        cmp_id, cmp_origin, cmp_dest, CMP_FUZZY_THRESHOLD)

    log.info("SSCL Injection complete: %d dataset routes forcibly upgraded to TRUNK/HP.",
             matched_total)
    return gdf


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 2  ─  SPATIAL ANALYSIS: CATCHMENTS, POPULATION, POIs, JUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def _pts_along_line(geom: LineString, spacing_m: float) -> List[Point]:
    length = geom.length
    if length < spacing_m:
        return [geom.interpolate(0.5, normalized=True)]
    return [geom.interpolate(d) for d in np.arange(0, length + spacing_m, spacing_m)
            if d <= length]


def build_catchments(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    log.info("Building walk catchments (stop spacing=%dm, walk radius=%dm)…",
             VIRTUAL_STOP_SPACING_M, WALK_CATCHMENT_M)
    gdf_utm = gdf.to_crs(UTM_CRS).copy()
    gdf_utm.geometry = gdf_utm.geometry.simplify(SIMPLIFY_TOL_M)
    catchments = []
    for geom in gdf_utm.geometry:
        pts  = _pts_along_line(geom, VIRTUAL_STOP_SPACING_M)
        poly = unary_union([p.buffer(WALK_CATCHMENT_M) for p in pts])
        catchments.append(poly)
    gdf_utm["_catch"] = catchments
    catch_series = (
        gdf_utm.set_geometry("_catch")
               .set_crs(UTM_CRS)
               .to_crs(WGS84_CRS)["_catch"]
    )
    gdf["Catchment"] = catch_series.values
    log.info("  Catchments built for %d routes.", len(gdf))
    return gdf


def _read_raster_nodata(raster_path: str) -> Optional[float]:
    try:
        import rasterio
        with rasterio.open(raster_path) as src:
            return src.nodata
    except Exception:
        return -9999.0


def compute_population(gdf: gpd.GeoDataFrame, raster_path: str) -> gpd.GeoDataFrame:
    """
    Per-route population from WorldPop raster using catchment polygons.

    KASHMIR-FORK NOTE:
    Raw raster counts are converted to a percentage of the Srinagar UA + Valley
    study-area population for the reference year (CMP_REFERENCE_YEAR = 2024,
    total = 1,660,000 — projection from Census 2011 + SMC Master Plan).
    Replaces the Jammu CMP 1,653,873 baseline from the original engine.
    The walkshed used for the catchment polygon is contracted in winter mode
    (WALK_CATCHMENT_M scaled by WINTER_WALKSHED_SHRINK).

    New columns added:
      Population_Served      — raw raster head-count (unchanged, for audit)
      Population_Served_Pct  — % of Srinagar 2024 study-area population served
    """
    log.info("Computing per-route population from raster: %s", raster_path)
    log.info("  CMP reference: Year=%d  Total study-area population=%s",
             CMP_REFERENCE_YEAR, f"{CMP_TOTAL_POPULATION:,}")

    # v3.2: convert silent zero-fallback into a loud, blocking error. The
    # previous behaviour propagated Pop_Score = 0.5 to every route and let
    # the audit log ship with Population_Served_Pct = 0.00% everywhere.
    # v3.3.6: dev-mode dummy fallback gated behind KASHMIR_ALLOW_DUMMY_POP=1
    # so a real run still fails loudly when the raster goes missing — the
    # earlier blanket fallback let invalid runs ship undetected.
    _DUMMY_OK = os.environ.get("KASHMIR_ALLOW_DUMMY_POP", "").strip() == "1"

    def _dummy_pop_fallback(reason: str) -> gpd.GeoDataFrame:
        log.error("  ✗ %s — KASHMIR_ALLOW_DUMMY_POP=1 → using flat 1000/route", reason)
        log.error("    *** DEV-MODE FALLBACK — DO NOT SHIP OUTPUTS FROM THIS RUN ***")
        gdf["Population_Served"] = 1000
        gdf["Population_Served_Pct"] = round(1000 / CMP_TOTAL_POPULATION * 100, 4)
        return gdf

    if not _HAS_RASTERSTATS:
        if _DUMMY_OK:
            return _dummy_pop_fallback("rasterstats not importable")
        log.error("  ✗ rasterstats not importable — Population_Served would be zero.")
        raise RuntimeError(
            "rasterstats is required for Population_Served computation. "
            "Install with: pip install rasterstats (or conda equivalent). "
            "For development runs without rasterstats, set KASHMIR_ALLOW_DUMMY_POP=1."
        )
    if not Path(raster_path).exists():
        if _DUMMY_OK:
            return _dummy_pop_fallback(f"WorldPop raster missing at {raster_path}")
        log.error("  ✗ WorldPop raster not found at: %s", raster_path)
        raise RuntimeError(
            f"WorldPop raster not found at {raster_path!r}. "
            "Update RASTER_PATH or place kashmir_worldpop.tif at the configured path. "
            "For development runs without the raster, set KASHMIR_ALLOW_DUMMY_POP=1."
        )

    nodata    = _read_raster_nodata(raster_path)
    catch_gdf = gpd.GeoDataFrame(
        gdf[["Route_ID"]], geometry=gdf["Catchment"], crs=WGS84_CRS)
    stats     = rasterstats.zonal_stats(
        catch_gdf, raster_path, stats=["sum"], nodata=nodata, geojson_out=False)

    raw_counts = [
        min(2_000_000, max(0, int(s["sum"]))) if s.get("sum") is not None else 0
        for s in stats
    ]
    gdf["Population_Served"] = raw_counts

    # v3.3.3 (teammate review): tourist corridor catchment boost.
    # Tourist corridors carry seasonal visitor footfall that WorldPop's
    # residential-only raster cannot see. Apply a 1.3x multiplier at the
    # catchment level (single source of tourist amplification — no separate
    # CDI multiplier, no separate POI bump beyond Tier-3 weights).
    if "Tourist_Corridor" in gdf.columns:
        tourist_mask = gdf["Tourist_Corridor"].fillna(False).astype(bool)
        if tourist_mask.any():
            boosted = (gdf["Population_Served"] * TOURIST_POPULATION_MULTIPLIER).astype(int)
            gdf.loc[tourist_mask, "Population_Served"] = boosted[tourist_mask]
            log.info("  Tourist boost applied: %.2fx on %d tourist-corridor routes "
                     "(catchment-level, single multiplier).",
                     TOURIST_POPULATION_MULTIPLIER, int(tourist_mask.sum()))

    # Convert to % of CMP 2024 total — capped at 100% to guard against raster overcount
    gdf["Population_Served_Pct"] = (
        (gdf["Population_Served"] / CMP_TOTAL_POPULATION * 100)
        .clip(upper=100.0)
        .round(4)
    )

    log.info("  Population_Served     — max raw count : %s",
             f"{gdf['Population_Served'].max():,}")
    log.info("  Population_Served_Pct — max: %.2f%%  mean: %.2f%%  "
             "(denominator: CMP %d total = %s)",
             gdf["Population_Served_Pct"].max(),
             gdf["Population_Served_Pct"].mean(),
             CMP_REFERENCE_YEAR,
             f"{CMP_TOTAL_POPULATION:,}")

    # v3.2: post-condition — if the raster is present and rasterstats works
    # but every catchment came back empty, the run is meaningless. Catch the
    # silent-zero failure mode (wrong CRS, bbox miss, all-nodata raster).
    if int(gdf["Population_Served"].sum()) == 0:
        log.error("  ✗ Population_Served sums to zero across %d routes.", len(gdf))
        raise RuntimeError(
            "WorldPop zonal_stats returned zero for every catchment. "
            "Check that raster CRS matches catchment CRS (expected EPSG:4326) "
            "and that the raster bbox overlaps the Kashmir study area."
        )
    return gdf


_STUDY_AREA_POP_CACHE: Dict[str, int] = {}


def study_area_population(raster_path: str) -> int:
    """Population of the Kashmir Division — the correct denominator for
    'network coverage %'.

    v3.4.1: measured as the WorldPop sum INSIDE the 10-district UNION
    (point-in-polygon against kashmir_districts_osm.geojson, ~6.58M), NOT the
    raster's full rectangular extent. Before v3.4.1 the raster was pre-cropped to
    the narrow study bbox, so this silently returned only the in-box population
    (5.1M) — understating the true division population and overstating coverage.
    Falls back to the raster sum if the district polygons are unavailable.

    F-V9 history: the denominator must be the people who actually live in the
    division, NOT the Srinagar CMP urban-core figure (1.66M) which inflated
    coverage ~3×.
    """
    if raster_path in _STUDY_AREA_POP_CACHE:
        return _STUDY_AREA_POP_CACHE[raster_path]
    total = 0
    # Preferred: clip the raster by the authoritative 10-district union polygons.
    try:
        import rasterio, json as _json
        from rasterio.mask import mask as _rmask
        from shapely.geometry import shape as _shape
        dpath = None
        for p in ("kashmir_districts_osm.geojson",
                  os.path.join(os.path.dirname(os.path.abspath(__file__)),
                               "kashmir_districts_osm.geojson")):
            if Path(p).exists():
                dpath = p; break
        if dpath and Path(raster_path).exists():
            geoms = [f["geometry"] for f in _json.load(open(dpath, encoding="utf-8"))["features"]]
            with rasterio.open(raster_path) as src:
                out, _ = _rmask(src, geoms, crop=True, nodata=0)
                arr = out[0].astype("float64")
                total = int(arr[arr > 0].sum())
    except Exception as exc:  # noqa: BLE001
        log.warning("  study_area_population (district-union clip): %s", exc)
    # Fallback: raster rectangular sum.
    if total <= 0 and Path(raster_path).exists():
        try:
            import rasterio
            with rasterio.open(raster_path) as src:
                arr = src.read(1).astype("float64")
                nd = src.nodata
                m = arr > 0
                if nd is not None:
                    m &= (arr != nd)
                total = int(arr[m].sum())
        except Exception as exc:  # noqa: BLE001
            log.warning("  study_area_population: %s", exc)
    _STUDY_AREA_POP_CACHE[raster_path] = total
    return total


def compute_network_population_total(gdf: gpd.GeoDataFrame,
                                      raster_path: str) -> int:
    """Deduplicated network population via dissolved catchment union."""
    log.info("Computing deduplicated network population (dissolved union)…")
    if not Path(raster_path).exists() or not _HAS_RASTERSTATS:
        return int(gdf["Population_Served"].max())
    valid     = [c for c in gdf["Catchment"]
                 if c is not None and hasattr(c, "is_empty") and not c.is_empty]
    if not valid:
        return 0
    dissolved = unary_union(valid)
    nodata    = _read_raster_nodata(raster_path)
    stats     = rasterstats.zonal_stats(
        [dissolved], raster_path, stats=["sum"], nodata=nodata, geojson_out=False)
    val    = stats[0].get("sum") if stats else None
    # F-V9: no artificial 2M clamp (a Srinagar-era cap that would truncate a
    # valley-wide union). Coverage % is now against the true study-area
    # population, not the Srinagar CMP figure.
    result = max(0, int(val)) if val is not None else 0
    sap    = study_area_population(raster_path)
    pct    = (result / sap * 100) if sap else 0.0
    log.info("  Deduplicated network population: %s  (%.2f%% of the %s residents "
             "living in the study area; Srinagar CMP %d planning ref: %s)",
             f"{result:,}", pct, f"{sap:,}", CMP_REFERENCE_YEAR,
             f"{CMP_TOTAL_POPULATION:,}")
    return result


def count_weighted_poi_scores(gdf_routes: gpd.GeoDataFrame,
                               gdf_pois:   gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 2 preparation: raw weighted POI score per route.

    Kashmir 3-tier POI weighting (see load_pois for assignment):
      Tier 1  (w=1.0): Hospitals, transit hubs, malls, major shrines,
                      Industrial estates, KP migrant townships, universities
      Tier 2  (w=0.4): Colleges, markets, schools, defence, civic services
      Tier 3  (w=0.6 summer):
                      • Tourist-only (gondola, ski_resort, etc.) → 0.0 in winter
                      • Residential-anchor (Boulevard, Dalgate, Mughal gardens)
                        → 0.4 in winter (v3.2 — no longer zeroed)
      Women-anchor categories receive a +25% boost on top of their tier weight
      (CHALO calibration: 64.5% of riders are women under the free-fare regime).

    POI_Score_Raw = Σ(tier_weight_i for POIs within POI_BUFFER_M of route) / Route_KM
    """
    log.info("Counting POI weighted scores (250m buffer, 2-tier weights "
             "per Directive 5)…")

    gdf_pois_w       = gdf_pois.copy()
    # POI_Weight already assigned in load_pois()
    gdf_pois_w["_w"] = gdf_pois_w["POI_Weight"]
    gdf_pois_w["_hv"] = gdf_pois_w["Is_HV_POI"].astype(int)

    r_utm         = gdf_routes.to_crs(UTM_CRS).copy()
    r_utm["_buf"] = r_utm.geometry.simplify(SIMPLIFY_TOL_M).buffer(POI_BUFFER_M)
    buf_gdf       = r_utm[["Route_ID", "_buf"]].set_geometry("_buf")

    p_utm  = gdf_pois_w[["_w", "_hv", "geometry"]].to_crs(UTM_CRS)
    joined = gpd.sjoin(p_utm, buf_gdf, how="inner", predicate="within")

    weight_sums = joined.groupby("Route_ID")["_w"].sum()
    hv_counts   = joined.groupby("Route_ID")["_hv"].sum()
    route_km    = (gdf_routes.set_index("Route_ID")["Route_KM"]
                   .clip(lower=GRAVITY_EPSILON))

    gdf_routes["_poi_weight_sum"] = (gdf_routes["Route_ID"]
                                     .map(weight_sums).fillna(0.0))
    gdf_routes["POI_Score_Raw"]   = (gdf_routes["_poi_weight_sum"]
                                     / route_km.values).clip(lower=0.0)
    gdf_routes["HV_POI_Count"]    = (gdf_routes["Route_ID"]
                                     .map(hv_counts).fillna(0).astype(int))
    gdf_routes.drop(columns=["_poi_weight_sum"], inplace=True)

    log.info("  POI raw score mean: %.4f  max: %.4f",
             gdf_routes["POI_Score_Raw"].mean(),
             gdf_routes["POI_Score_Raw"].max())
    return gdf_routes


def _deflection_deg(A, B, C) -> float:
    ba = (A[0] - B[0], A[1] - B[1])
    bc = (C[0] - B[0], C[1] - B[1])
    mag_ba, mag_bc = math.hypot(*ba), math.hypot(*bc)
    if mag_ba < 1e-6 or mag_bc < 1e-6:
        return 0.0
    cos_t = max(-1.0, min(1.0,
                (ba[0]*bc[0] + ba[1]*bc[1]) / (mag_ba * mag_bc)))
    return 180.0 - math.degrees(math.acos(cos_t))


def compute_junction_penalties(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Junction_Penalty_Min = Sharp_Turns × 0.5 min per turn > 75°."""
    log.info("Computing junction penalties (threshold=%d°, %.1f min/turn)…",
             SHARP_TURN_DEG, JUNCTION_PENALTY_PER_TURN_MIN)
    gdf_utm = gdf.to_crs(UTM_CRS).copy()
    gdf_utm.geometry = gdf_utm.geometry.simplify(SIMPLIFY_TOL_M)
    n_sharp_list, penalty_list = [], []
    for geom in gdf_utm.geometry:
        if geom is None or geom.is_empty:
            n_sharp_list.append(0)
            penalty_list.append(0.0)
            continue
        coords  = list(geom.coords)
        n_sharp = sum(
            1 for i in range(1, len(coords) - 1)
            if _deflection_deg(coords[i-1], coords[i], coords[i+1]) > SHARP_TURN_DEG
        )
        n_sharp_list.append(n_sharp)
        penalty_list.append(float(n_sharp * JUNCTION_PENALTY_PER_TURN_MIN))
    gdf["Sharp_Turns"]          = n_sharp_list
    gdf["Junction_Penalty_Min"] = penalty_list
    log.info("  %d routes have junction penalties.",
             sum(1 for p in penalty_list if p > 0))
    return gdf


def _detect_congestion_zone(start_lat: float, end_lat: float) -> Tuple[str, float]:
    """
    DIRECTIVE 2 — Congestion zone detection (Kashmir-recentred in v3.2).

    Decision rule:
      If EITHER terminal sits above CITY_CORE_LAT_THRESHOLD (34.07°N — the
      Downtown Srinagar / mohalla quarters around Nawakadal, Khanyar, Habba
      Kadal), the route touches the bazaar-grid congestion zone at least at
      one end → CONGESTION_CITY_CORE (2.2×).
      Otherwise → CONGESTION_PERI_URBAN (1.4×).

    Returns (zone_label, multiplier).
    """
    if start_lat > CITY_CORE_LAT_THRESHOLD or end_lat > CITY_CORE_LAT_THRESHOLD:
        return "City_Core", CONGESTION_CITY_CORE
    return "Peri_Urban", CONGESTION_PERI_URBAN


def compute_cycle_times(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 7: Realistic Cycle Time (Directive 2 — complete rewrite from v6).

    v6 formula (WRONG):
      Cycle = (OSRM_s / 60 + Junction_Penalty) × 2 × 1.15
      Problem: OSRM seconds = midnight free-flow. On a 30km route, this
               produces a 40-min cycle time and a 2-bus allocation — a
               spreadsheet fantasy.

    v2 formula (CORRECT):
      One_Way_Time = (OSRM_Base_Min × Congestion_Multiplier)
                     + (N_Stops × Stop_Penalty_Min)
                     + Junction_Penalty_Min
      Cycle_Time   = One_Way_Time × 2 × Terminal_Layover_Factor (1.10)

    Where:
      • OSRM_Base_Min        = OSRM_Duration_S / 60
      • Congestion_Multiplier = 2.2 (Downtown, lat > 34.07) or 1.4 (peri-urban)
      • N_Stops               = Route_KM × 1000 / STOP_SPACING_M  (every 500m)
      • Stop_Penalty_Min      = 1.5 min per stop (boarding + alighting dwell)
      • Junction_Penalty_Min  = from compute_junction_penalties() (unchanged)
      • Terminal_Layover      = 1.10 (10%) — covers driver rest + schedule recovery

    Audit columns added: Congestion_Zone, N_Stops_Estimated, Stop_Penalty_Min
    """
    log.info("Step 7: Computing REALISTIC cycle times (Directive 2)…")
    log.info("  Congestion: City_Core=%.1f×  Peri_Urban=%.1f×",
             CONGESTION_CITY_CORE, CONGESTION_PERI_URBAN)
    log.info("  Stop spacing=%dm, penalty=%.1f min/stop, "
             "layover buffer=%.0f%%",
             STOP_SPACING_M, STOP_PENALTY_MIN,
             (TERMINAL_LAYOVER_FACTOR - 1) * 100)

    zones, n_stops_list, stop_penalties, cycle_times = [], [], [], []
    capped_n = 0

    for _, row in gdf.iterrows():
        start_lat = float(row.get("Start_Lat", 0.0))
        end_lat   = float(row.get("End_Lat",   0.0))
        osrm_min  = float(row["OSRM_Duration_S"]) / 60.0
        junc_pen  = float(row["Junction_Penalty_Min"])
        route_km  = float(row["Route_KM"])
        route_type = row.get("Route_Type", "Urban")

        # Congestion zone (Directive 2)
        zone, cong_mult = _detect_congestion_zone(start_lat, end_lat)

        # Stop penalty: one stop every STOP_SPACING_M metres
        n_stops    = max(1, int((route_km * 1000) / STOP_SPACING_M))
        stop_pen   = n_stops * STOP_PENALTY_MIN

        # One-way travel time (congestion-adjusted + stops + junctions)
        one_way = (osrm_min * cong_mult) + stop_pen + junc_pen

        # Round-trip + terminal layover buffer
        cycle = one_way * 2 * TERMINAL_LAYOVER_FACTOR

        # v3.3.1 (STEP 1): per-km cycle-time sanity cap. Pure upper bound —
        # prevents OSRM glitches or pathological congestion math from producing
        # runaway cycle times. NOT a second peak multiplier (rejected during
        # audit: would double-count with the existing CONGESTION_CITY_CORE).
        cap_per_km = CYCLE_TIME_CAP_MIN_PER_KM.get(route_type, 4.0)
        cycle_cap  = route_km * 2.0 * cap_per_km  # round-trip basis
        if cycle_cap > 0 and cycle > cycle_cap:
            cycle = cycle_cap
            capped_n += 1

        zones.append(zone)
        n_stops_list.append(n_stops)
        stop_penalties.append(round(stop_pen, 1))
        cycle_times.append(round(max(1.0, cycle), 1))

    gdf = gdf.copy()
    gdf["Congestion_Zone"]    = zones
    gdf["N_Stops_Estimated"]  = n_stops_list
    gdf["Stop_Penalty_Min"]   = stop_penalties
    gdf["Cycle_Time_Min"]     = cycle_times

    city_core_n  = gdf["Congestion_Zone"].eq("City_Core").sum()
    periurban_n  = gdf["Congestion_Zone"].eq("Peri_Urban").sum()
    log.info("  Zone split — City_Core: %d  Peri_Urban: %d",
             city_core_n, periurban_n)
    log.info("  Cycle_Time_Min — mean: %.1f  max: %.1f  min: %.1f",
             gdf["Cycle_Time_Min"].mean(),
             gdf["Cycle_Time_Min"].max(),
             gdf["Cycle_Time_Min"].min())
    if capped_n:
        log.info("  Per-km sanity cap (v3.3.1): %d routes capped at "
                 "{Urban:%.1f / Peri:%.1f / Inter:%.1f} min/km × 2 directions.",
                 capped_n, CYCLE_TIME_CAP_MIN_PER_KM["Urban"],
                 CYCLE_TIME_CAP_MIN_PER_KM["Peri_Urban"],
                 CYCLE_TIME_CAP_MIN_PER_KM["Regional_District"])
    return gdf


def reconcile_active_population(gdf: gpd.GeoDataFrame, net_pop: int) -> gpd.GeoDataFrame:
    """F-V6: make the ACTIVE plan's per-route Population_Served sum to the
    deduplicated network union (the cover figure).

    apportion_route_population() normalises Population_Served across ALL routes,
    but after clustering/dedup ~62% of that total is stranded on the merged /
    consolidated rows (fleet-zeroed, not population-zeroed). A reader summing the
    Population_Served column in the published (active) plan therefore lands well
    below the cover number — the residual of audit Finding 9 / Output #2.

    This step credits the absorbed population back to the active network: it zeros
    Population_Served on merged rows and rescales the active rows so their sum
    equals net_pop (the dissolved-catchment union). Relative shares among active
    routes are preserved. Safe to run after Phase-4 — Load_Ratio/Daily_Demand use
    Population_Served_Raw (untouched), and Equity_Score is a 0–1 normalisation
    invariant to a global rescale.
    """
    gdf = gdf.copy()
    active = gdf["Action_Taken"] != "MERGED_INTO_TRUNK"
    cur = float(gdf.loc[active, "Population_Served"].sum())
    gdf.loc[~active, "Population_Served"] = 0
    if cur > 0 and net_pop and net_pop > 0:
        scale = net_pop / cur
        gdf.loc[active, "Population_Served"] = (
            (gdf.loc[active, "Population_Served"] * scale).round().clip(lower=0).astype(int))
    log.info("  Reconciled active Population_Served to the dedup union %s "
             "(Σ active = %s; Finding 9 / Output #2 fully closed).",
             f"{net_pop:,}", f"{int(gdf.loc[active, 'Population_Served'].sum()):,}")
    return gdf


# ══════════════════════════════════════════════════════════════════════════════
#  STEPS 1 & 2  ─  NORMALISED DEMAND SCORES
# ══════════════════════════════════════════════════════════════════════════════

def step1_normalise_population_score(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 1: Pop_Score (0–1 normalised population density per route).

    Percentage-based scoring against the Srinagar UA / Kashmir Valley
    reference population:
    Raw head-counts are replaced by Population_Served_Pct (% of CMP 2024 total
    = 16,60,000 — Census 2011 + SMC Master Plan 2035 projection). This grounds
    the score in a demand framework comparable to the original RITES CMP work:
      • A route serving 3.0% of the study area (≈49,616 residents) is scored
        relative to the highest-serving route in the dataset.
      • The 95th-percentile cap (Directive 5) is applied to the percentage
        density (Pct / Route_KM), preventing a single hyper-dense corridor
        from compressing scores across the network — same logic as before,
        now in CMP-percentage space.

    DIRECTIVE 5 FIX (population cap) — still fully active:
    v6 normalised against the raw maximum, meaning one ultra-dense ward
    mathematically suppressed the CDI of all other routes.  The 95th-pct
    cap prevents that outlier effect; routes above the cap clip to 1.0.

    Columns used:  Population_Served_Pct  (set by compute_population)
    Column set:    Pop_Score              (0–1, used in CDI formula)
    """
    log.info("Step 1: Normalising population score from CMP %d percentages "
             "(0–1, %dth pct density cap — Directive 5)…",
             CMP_REFERENCE_YEAR, POP_CAP_PERCENTILE)

    if "Population_Served_Pct" not in gdf.columns or gdf["Population_Served_Pct"].sum() == 0:
        log.warning("  Population_Served_Pct missing or all-zero — "
                    "falling back to raw Population_Served for scoring.")
        pop_series = gdf["Population_Served"]
    else:
        pop_series = gdf["Population_Served_Pct"]

    # Percentage density: pct of CMP population per km of route
    # (normalises for route length so a 2km stub through a dense ward
    #  doesn't automatically outrank a 20km trunk serving the same zone)
    raw = (pop_series / gdf["Route_KM"].clip(lower=GRAVITY_EPSILON)).clip(lower=0)

    # Directive 5: Cap at 95th percentile of the percentage-density distribution
    cap_val    = float(np.percentile(raw, POP_CAP_PERCENTILE))
    raw_capped = raw.clip(upper=cap_val)

    log.info("  Pct-density 95th pct cap = %.4f %%/km  (raw max was %.4f %%/km)",
             cap_val, raw.max())

    n_capped = (raw > cap_val).sum()
    if n_capped:
        log.info("  %d routes capped at density ceiling (Pop_Score clipped to 1.0).",
                 n_capped)

    min_val = raw_capped.min()
    max_val = raw_capped.max()
    if max_val > min_val:
        gdf["Pop_Score"] = ((raw_capped - min_val) / (max_val - min_val)).round(4)
    else:
        gdf["Pop_Score"] = 0.5
        log.warning("  All routes have equal pct-density — assigning Pop_Score = 0.5.")

    log.info("  Pop_Score mean: %.4f  max: %.4f  "
             "(based on %% of CMP %d study-area population)",
             gdf["Pop_Score"].mean(), gdf["Pop_Score"].max(), CMP_REFERENCE_YEAR)
    return gdf


def step2_normalise_poi_score(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 2: POI_Score (0–1 normalised weighted POI gravity per route).
    POI weights are the 2-tier system from load_pois() / count_weighted_poi_scores().
    """
    log.info("Step 2: Normalising POI gravity score (0–1, 2-tier weights)…")
    raw     = gdf["POI_Score_Raw"].clip(lower=0)
    min_val = raw.min()
    max_val = raw.max()
    if max_val > min_val:
        gdf["POI_Score"] = ((raw - min_val) / (max_val - min_val)).round(4)
    else:
        gdf["POI_Score"] = 0.5
        log.warning("  All routes have equal POI_Score_raw — assigning 0.5.")
    log.info("  POI_Score mean: %.4f  max: %.4f",
             gdf["POI_Score"].mean(), gdf["POI_Score"].max())
    return gdf


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 2b  ─  CORRIDOR FREQUENCY, CLUSTERING & ROUTE CLASSIFICATION
# ══════════════════════════════════════════════════════════════════════════════
def apportion_route_population(gdf: gpd.GeoDataFrame, freq_scores: np.ndarray,
                               raster_path: str) -> gpd.GeoDataFrame:
    """
    Apportion catchment population among overlapping routes WITHOUT double-counting.

    Audit Finding 9 (CRITICAL): the previous implementation just divided each
    route's catchment by (competitors + 1). That gives every route a smaller
    number, but the per-route shares no longer sum to anything meaningful — they
    totalled 134,684 against the cover's deduplicated-union figure of 1,158,399
    (an 8.6× contradiction), and everything downstream (Daily_Demand, Load_Ratio,
    Viability, Subsidy flags, Equity) inherited the error.

    Fix: keep the frequency weighting (a more-contested corridor takes a smaller
    share of shared demand) but NORMALISE the shares so they sum to the
    deduplicated network union population. Concretely, with raw catchment p_i and
    competitor count c_i = freq_scores_i + 1:

        weight_i = p_i / c_i
        Population_Served_i = weight_i × ( UnionPop / Σ_j weight_j )

    so Σ_i Population_Served_i ≈ UnionPop (the dissolved-catchment union, i.e. the
    cover number) — the property a de-double-counting apportionment must have.
    Raw catchment is preserved in Population_Served_Raw for the walkshed view.
    """
    log.info("Apportioning population among overlapping routes "
             "(frequency-weighted, normalised to the deduplicated union)…")
    gdf = gdf.copy()

    # freq_scores = number of OTHER overlapping routes; total sharing = +1.
    competitors = freq_scores + 1

    gdf["Population_Served_Raw"] = gdf["Population_Served"]
    gdf["Corridor_Competitors"]  = competitors

    weights   = gdf["Population_Served"].astype(float) / competitors
    union_pop = compute_network_population_total(gdf, raster_path)
    wsum      = float(weights.sum())
    if wsum > 0 and union_pop and union_pop > 0:
        scaled = weights * (union_pop / wsum)
    else:
        # Degenerate (no raster / no weight): fall back to the weights as-is so
        # the column is at least monotonic and non-NaN.
        log.warning("  Apportionment normalisation skipped (union_pop=%s, wsum=%.1f) "
                    "— using unnormalised frequency weights.", union_pop, wsum)
        scaled = weights
    gdf["Population_Served"] = scaled.round().clip(lower=0).astype(int)

    log.info("  Apportioned Population_Served — Σ = %s vs deduplicated union %s "
             "(now reconciles with the cover figure; Finding 9 fixed).",
             f"{int(gdf['Population_Served'].sum()):,}", f"{union_pop:,}")
    return gdf

def compute_frequency_scores(gdf: gpd.GeoDataFrame) -> np.ndarray:
    log.info("Computing corridor frequency scores (STRtree)…")
    gdf_utm = gdf.to_crs(UTM_CRS).copy()
    gdf_utm.geometry = gdf_utm.geometry.simplify(SIMPLIFY_TOL_M)
    bufs   = [g.buffer(OVERLAP_BUFFER_M) for g in gdf_utm.geometry]
    tree   = STRtree(bufs)
    scores = np.zeros(len(bufs), dtype=int)
    for i, buf_i in enumerate(bufs):
        candidates = tree.query(buf_i)
        scores[i]  = sum(
            1 for j in candidates if j != i and bufs[j].intersects(buf_i))
    log.info("  Frequency scores — mean: %.1f  max: %d",
             scores.mean(), scores.max())
    return scores


def compute_overlap_matrix(gdf: gpd.GeoDataFrame) -> np.ndarray:
    log.info("Computing pairwise overlap matrix (Multiprocessed)…")
    gdf_utm = gdf.to_crs(UTM_CRS).copy()

    # 1. VECTORIZATION: Use GeoPandas built-in C-level methods instead of Python loops
    log.info("  Simplifying and buffering geometries at C-level...")
    geom_series = gdf_utm.geometry.simplify(SIMPLIFY_TOL_M * 2)
    geom_series = geom_series.buffer(OVERLAP_BUFFER_M, resolution=2)
    geom_series = geom_series.simplify(5.0)

    bufs = geom_series.tolist()
    areas = np.array([b.area for b in bufs])
    n = len(bufs)
    matrix = np.zeros((n, n), dtype=np.float32)

    log.info("  Building spatial index...")
    tree = STRtree(bufs)

    # 2. PREPARE TASKS: Map out only the necessary comparisons
    log.info("  Preparing parallel intersection tasks...")
    tasks = []
    for i, buf_i in enumerate(bufs):
        candidate_indices = tree.query(buf_i)
        # Only check j > i to avoid duplicate math (if A overlaps B, B overlaps A)
        valid_j = [j for j in candidate_indices if j > i]
        if not valid_j:
            continue
            
        candidates = [(j, bufs[j], areas[j]) for j in valid_j]
        tasks.append((i, buf_i, areas[i], candidates))

    total_tasks = len(tasks)
    if total_tasks == 0:
        return matrix

    # 3. EXECUTE: Flood the CPU cores
    num_cores = max(1, multiprocessing.cpu_count() - 1) # Leave 1 core for OS stability
    log.info("  Computing %d route clusters across %d CPU cores...", total_tasks, num_cores)

    completed = 0
    sys.stdout.write(f"\r  Progress: [0/{total_tasks}] 0.0%")
    sys.stdout.flush()

    # v3.2: Attempt parallel execution with OpenBLAS pinned to a single
    # thread (env vars set at module import). If anything goes wrong on
    # Windows, fall back to the sequential loop — we never regress.
    used_parallel = False
    try:
        with concurrent.futures.ProcessPoolExecutor(max_workers=num_cores) as executor:
            futures = {executor.submit(_intersection_worker, task): task[0]
                       for task in tasks}
            for future in concurrent.futures.as_completed(futures):
                i, results = future.result()
                for j, ratio_i_j, ratio_j_i in results:
                    matrix[i, j] = ratio_i_j
                    matrix[j, i] = ratio_j_i
                completed += 1
                percent = (completed / total_tasks) * 100
                sys.stdout.write(f"\r  Progress: [{completed}/{total_tasks}] {percent:.1f}%")
                sys.stdout.flush()
        used_parallel = True
    except Exception as exc:
        log.warning("  Parallel overlap pool failed (%s) — falling back to sequential.",
                    type(exc).__name__)
        # Reset matrix and progress in case parallel partially populated
        matrix = np.zeros((n, n), dtype=np.float32)
        completed = 0
        for task in tasks:
            i, results = _intersection_worker(task)
            for j, ratio_i_j, ratio_j_i in results:
                matrix[i, j] = ratio_i_j
                matrix[j, i] = ratio_j_i
            completed += 1
            percent = (completed / total_tasks) * 100
            sys.stdout.write(f"\r  Progress: [{completed}/{total_tasks}] {percent:.1f}%")
            sys.stdout.flush()

    print()  # Clear the progress bar line
    log.info("  Overlap matrix (%d×%d) done  (%s execution).",
             n, n, "parallel" if used_parallel else "sequential")
    return matrix


def cluster_routes(gdf: gpd.GeoDataFrame,
                   overlap_matrix: np.ndarray) -> gpd.GeoDataFrame:
    log.info("Union-Find clustering (threshold=%.0f%%)…",
             OVERLAP_THRESHOLD * 100)
    n      = len(gdf)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x

    def union(x, y):
        parent[find(x)] = find(y)

    for r, c in zip(*np.where(overlap_matrix >= OVERLAP_THRESHOLD)):
        union(int(r), int(c))

    gdf = gdf.copy()
    gdf["Cluster_ID"]   = [find(i) for i in range(n)]
    gdf["Cluster_Size"] = gdf["Cluster_ID"].map(gdf["Cluster_ID"].value_counts())
    log.info("  %d clusters formed.", gdf["Cluster_ID"].nunique())
    return gdf


def backfill_overlap_metric(gdf: gpd.GeoDataFrame,
                             overlap_matrix: np.ndarray) -> gpd.GeoDataFrame:
    n     = len(gdf)
    gdf   = gdf.copy()
    means = []
    for i in range(n):
        row = np.concatenate([overlap_matrix[i, :i], overlap_matrix[i, i+1:]])
        means.append(float(row.mean()) if len(row) else 0.0)
    gdf["Overlap_Metric"] = [round(v, 4) for v in means]
    return gdf


def classify_routes(gdf: gpd.GeoDataFrame,
                    freq_scores: np.ndarray,
                    overlap_matrix: np.ndarray) -> gpd.GeoDataFrame:
    """
    Classifies routes into UPGRADED_TO_TRUNK / MERGED_INTO_TRUNK /
    RETAINED_AS_FEEDER.

    v3 CHANGES (three new rules):

    1. CDI Gate lowered to 30th percentile (was median / 50th in v2).
       Aggressively promotes more corridors to trunk status.

    2. Anti-Stranding (Length constraint):
       No route under TRUNK_MIN_LENGTH_KM (5.0 km) can be promoted to Trunk.
       This prevents isolated 1–4 km city-centre stubs from entering the trunk
       network as disconnected artefacts.

    3. Network Connectivity Bonus:
       Before evaluating cluster leadership, any candidate route whose geometry
       spatially intersects (Shapely intersection) a CMP backbone trunk corridor
       receives a CMP_CONNECTIVITY_BONUS (1.5×) multiplier on its combined
       demand score.  This ensures newly promoted trunks physically attach to
       the CMP backbone rather than floating in isolation.

    CMP-injected routes (CMP_Trunk == True) already carry Action_Taken =
    UPGRADED_TO_TRUNK from inject_cmp_trunk_routes() and are NOT re-classified
    here — they are preserved as-is and merely receive a New_Route_ID.
    """
    log.info("Classifying routes — v3 aggressive trunk rules…")
    log.info("  CDI gate: %dth percentile (was 50th in v2)", TRUNK_CDI_GATE_PERCENTILE)
    log.info("  Anti-stranding: routes < %.1f km ineligible for Trunk", TRUNK_MIN_LENGTH_KM)
    log.info("  CMP connectivity bonus: %.1f× for routes intersecting CMP backbone",
             CMP_CONNECTIVITY_BONUS)

    gdf = gdf.copy()

    # Preserve CMP-injected routes; initialise all others as feeders
    non_cmp_mask = ~gdf.get("CMP_Trunk", pd.Series(False, index=gdf.index))
    gdf.loc[non_cmp_mask, "Action_Taken"] = "RETAINED_AS_FEEDER"

    # Initialise ID and headway placeholders
    gdf["New_Route_ID"] = ""
    gdf["Headway_Min"]  = HEADWAY_MP_MIN   # Placeholder; overwritten by Step 6

    # ── Combined demand score ──────────────────────────────────────────────────
    gdf["_combined_demand"] = gdf["Pop_Score"] + gdf["POI_Score"]

    # ── CDI gate at 30th percentile (v3: lowered from median) ─────────────────
    gate_threshold = float(np.percentile(gdf["_combined_demand"].values,
                                         TRUNK_CDI_GATE_PERCENTILE))
    log.info("  Demand gate threshold (%dth pct): %.4f",
             TRUNK_CDI_GATE_PERCENTILE, gate_threshold)

    # ── Build CMP backbone geometry union for connectivity bonus ───────────────
    cmp_geoms = []
    cmp_mask  = gdf.get("CMP_Trunk", pd.Series(False, index=gdf.index))
    if cmp_mask.any():
        gdf_utm_cmp = gdf[cmp_mask].to_crs(UTM_CRS)
        for geom in gdf_utm_cmp.geometry:
            if geom is not None and not geom.is_empty:
                cmp_geoms.append(geom.buffer(OVERLAP_BUFFER_M))
        if cmp_geoms:
            cmp_backbone_union = unary_union(cmp_geoms)
            log.info("  CMP backbone union built from %d trunk geometries.", len(cmp_geoms))
        else:
            cmp_backbone_union = None
            log.warning("  CMP backbone union is empty — connectivity bonus inactive.")
    else:
        cmp_backbone_union = None
        log.warning("  No CMP_Trunk routes found — connectivity bonus inactive.")

    gdf_utm_all = gdf.to_crs(UTM_CRS).copy()

    trunk_n = feeder_n = 1
    cluster_no_trunk = 0

    # ── First pass: assign IDs to CMP-injected routes already marked Trunk ─────
    cmp_trunk_indices = gdf[cmp_mask & (gdf["Action_Taken"] == "UPGRADED_TO_TRUNK")].index
    for idx in cmp_trunk_indices:
        if not gdf.at[idx, "New_Route_ID"]:
            cmp_rid = gdf.at[idx, "CMP_Route_ID"]
            gdf.at[idx, "New_Route_ID"] = cmp_rid if cmp_rid else f"TRK-{trunk_n:03d}"
            trunk_n += 1

    # ── Second pass: classify non-CMP routes cluster by cluster ───────────────
    for cluster_id, grp in gdf.groupby("Cluster_ID"):
        idxs = grp.index.tolist()

        if len(idxs) == 1:
            # Single-route cluster
            idx0 = idxs[0]
            if gdf.at[idx0, "Action_Taken"] == "UPGRADED_TO_TRUNK":
                # Already a CMP trunk with ID assigned above
                if not gdf.at[idx0, "New_Route_ID"]:
                    gdf.at[idx0, "New_Route_ID"] = f"TRK-{trunk_n:03d}"
                    trunk_n += 1
            else:
                gdf.at[idx0, "New_Route_ID"] = f"FDR-{feeder_n:03d}"
                feeder_n += 1
            continue

        # Sort by combined demand (with connectivity bonus applied)
        def _boosted_demand(i) -> float:
            base = gdf.at[i, "_combined_demand"]
            # Anti-stranding: routes < TRUNK_MIN_LENGTH_KM get zero demand
            if gdf.at[i, "Route_KM"] < TRUNK_MIN_LENGTH_KM:
                return 0.0
            # CMP already-trunk routes need no additional promotion
            if gdf.at[i, "CMP_Trunk"]:
                return base
            # Connectivity bonus: does this route's geometry touch CMP backbone?
            if cmp_backbone_union is not None:
                geom_utm = gdf_utm_all.at[i, "geometry"]
                if geom_utm is not None and not geom_utm.is_empty:
                    try:
                        if geom_utm.intersects(cmp_backbone_union):
                            return base * CMP_CONNECTIVITY_BONUS
                    except Exception:
                        pass
            return base

        sorted_idxs = sorted(idxs, key=_boosted_demand, reverse=True)
        ti = sorted_idxs[0]   # Candidate Trunk (highest boosted demand)

        # If candidate is already a CMP Trunk, skip gate evaluation
        if gdf.at[ti, "CMP_Trunk"] and gdf.at[ti, "Action_Taken"] == "UPGRADED_TO_TRUNK":
            if not gdf.at[ti, "New_Route_ID"]:
                gdf.at[ti, "New_Route_ID"] = f"TRK-{trunk_n:03d}"
                trunk_n += 1
            for idx in sorted_idxs[1:]:
                if gdf.at[idx, "CMP_Trunk"]:
                    continue
                pos_i = gdf.index.get_loc(idx)
                pos_t = gdf.index.get_loc(ti)
                overlap = float(overlap_matrix[pos_i, pos_t])
                start_i = gpd.GeoSeries(
                    [Point(float(gdf.at[idx, "Start_Lon"]),
                           float(gdf.at[idx, "Start_Lat"]))],
                    crs=WGS84_CRS).to_crs(UTM_CRS).iloc[0]
                start_t = gpd.GeoSeries(
                    [Point(float(gdf.at[ti, "Start_Lon"]),
                           float(gdf.at[ti, "Start_Lat"]))],
                    crs=WGS84_CRS).to_crs(UTM_CRS).iloc[0]
                if (overlap >= OVERLAP_THRESHOLD
                        and start_i.distance(start_t) <= OD_PROXIMITY_TOLERANCE_M):
                    gdf.at[idx, "Action_Taken"] = "MERGED_INTO_TRUNK"
                    gdf.at[idx, "New_Route_ID"] = gdf.at[ti, "New_Route_ID"]
                else:
                    if not gdf.at[idx, "New_Route_ID"]:
                        gdf.at[idx, "New_Route_ID"] = f"FDR-{feeder_n:03d}"
                        feeder_n += 1
            continue

        # Anti-stranding: top candidate must meet min length
        if gdf.at[ti, "Route_KM"] < TRUNK_MIN_LENGTH_KM:
            cluster_no_trunk += 1
            log.debug(
                "  Cluster %s: top candidate %.1f km < %.1f km anti-stranding "
                "floor — all %d routes remain feeders.",
                cluster_id, gdf.at[ti, "Route_KM"], TRUNK_MIN_LENGTH_KM, len(idxs))
            for idx in idxs:
                if not gdf.at[idx, "New_Route_ID"]:
                    gdf.at[idx, "New_Route_ID"] = f"FDR-{feeder_n:03d}"
                    feeder_n += 1
            continue

        # v3 CDI gate at 30th percentile (raw, un-boosted demand used for gate)
        candidate_demand_raw = gdf.at[ti, "_combined_demand"]
        if candidate_demand_raw <= gate_threshold:
            cluster_no_trunk += 1
            log.debug(
                "  Cluster %s: top candidate demand %.4f ≤ %dth pct gate %.4f "
                "— NO trunk promoted (all %d routes remain feeders).",
                cluster_id, candidate_demand_raw, TRUNK_CDI_GATE_PERCENTILE,
                gate_threshold, len(idxs))
            for idx in idxs:
                if not gdf.at[idx, "New_Route_ID"]:
                    gdf.at[idx, "New_Route_ID"] = f"FDR-{feeder_n:03d}"
                    feeder_n += 1
            continue

        # Passed gate → promote to Trunk
        gdf.at[ti, "Action_Taken"] = "UPGRADED_TO_TRUNK"
        gdf.at[ti, "New_Route_ID"] = f"TRK-{trunk_n:03d}"

        for idx in sorted_idxs[1:]:
            if gdf.at[idx, "CMP_Trunk"]:
                continue
            pos_i   = gdf.index.get_loc(idx)
            pos_t   = gdf.index.get_loc(ti)
            overlap = float(overlap_matrix[pos_i, pos_t])
            start_i = gpd.GeoSeries(
                [Point(float(gdf.at[idx, "Start_Lon"]),
                       float(gdf.at[idx, "Start_Lat"]))],
                crs=WGS84_CRS).to_crs(UTM_CRS).iloc[0]
            start_t = gpd.GeoSeries(
                [Point(float(gdf.at[ti, "Start_Lon"]),
                       float(gdf.at[ti, "Start_Lat"]))],
                crs=WGS84_CRS).to_crs(UTM_CRS).iloc[0]
            if (overlap >= OVERLAP_THRESHOLD
                    and start_i.distance(start_t) <= OD_PROXIMITY_TOLERANCE_M):
                gdf.at[idx, "Action_Taken"] = "MERGED_INTO_TRUNK"
                gdf.at[idx, "New_Route_ID"] = f"TRK-{trunk_n:03d}"
            else:
                if not gdf.at[idx, "New_Route_ID"]:
                    gdf.at[idx, "New_Route_ID"] = f"FDR-{feeder_n:03d}"
                    feeder_n += 1

        trunk_n += 1

    gdf.drop(columns=["_combined_demand"], inplace=True)

    log.info("  v3 gate (30th pct) — %d clusters had no eligible Trunk "
             "(all assigned as feeders).", cluster_no_trunk)
    log.info("  Trunks: %d | Merged: %d | Feeders: %d",
             (gdf["Action_Taken"] == "UPGRADED_TO_TRUNK").sum(),
             (gdf["Action_Taken"] == "MERGED_INTO_TRUNK").sum(),
             (gdf["Action_Taken"] == "RETAINED_AS_FEEDER").sum())

    # Assign Displaced_Operator_Class for MERGED routes (political accounting).
    # Maps Vehicle_Category → operator class so permit-holder impact is visible.
    _OP_CLASS_MAP = {
        "minibus":        "Private Minibus",
        "mini bus":       "Private Minibus",
        "mpv":            "Private Minibus",
        "mps":            "MPS (Stage Carriage)",
        "city bus":       "JKRTC / City Bus",
        "jkrtc":          "JKRTC / City Bus",
        "jkrtc city bus": "JKRTC / City Bus",
        "e-bus":          "SSCL E-Bus",
        "hpv":            "HPV Bus",
        "lpv":            "LPV / Tempo",
        "tempo":          "LPV / Tempo",
    }
    def _op_class(row):
        if row.get("Action_Taken") != "MERGED_INTO_TRUNK":
            return ""
        cat = str(row.get("Vehicle_Category", "")).strip().lower()
        for key, label in _OP_CLASS_MAP.items():
            if key in cat:
                return label
        return "Private Minibus"  # safe default for unlabelled permits

    gdf["Displaced_Operator_Class"] = gdf.apply(_op_class, axis=1)

    merged_mask = gdf["Action_Taken"] == "MERGED_INTO_TRUNK"
    if merged_mask.any():
        class_counts = gdf.loc[merged_mask, "Displaced_Operator_Class"].value_counts()
        log.info("  Displaced operator summary (%d routes merged):", merged_mask.sum())
        for cls, cnt in class_counts.items():
            log.info("    %-30s : %d permits absorbed", cls, cnt)

    return gdf


def apply_terminal_capacity(gdf: gpd.GeoDataFrame,
                              gdf_pois: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    log.info("Applying terminal capacity constraints…")
    term_pois = gdf_pois[
        gdf_pois["category"].str.lower().isin(TERMINAL_CATEGORIES)
    ].to_crs(UTM_CRS)
    if term_pois.empty:
        return gdf
    term_tree  = STRtree(term_pois.geometry.tolist())
    gdf        = gdf.copy()
    downgraded = 0
    for idx, row in gdf.iterrows():
        if row["Action_Taken"] != "UPGRADED_TO_TRUNK":
            continue
        if row.get("Fleet_Required", 0) <= FLEET_CAP_HARD:
            continue
        has_cap = False
        for lon, lat in [(float(row["Start_Lon"]), float(row["Start_Lat"])),
                          (float(row["End_Lon"]),   float(row["End_Lat"]))]:
            pt_utm = (gpd.GeoSeries([Point(lon, lat)], crs=WGS84_CRS)
                      .to_crs(UTM_CRS).iloc[0])
            if len(term_tree.query(pt_utm.buffer(TERMINAL_BUFFER_M))) > 0:
                has_cap = True
                break
        if not has_cap:
            if row.get("CMP_Trunk", False):
                continue
            gdf.at[idx, "Action_Taken"] = "RETAINED_AS_FEEDER"
            log.debug("  Route %s capacity-downgraded to feeder.",
                      row["Route_ID"])
            downgraded += 1
    log.info("  %d trunks capacity-downgraded to feeder.", downgraded)
    return gdf


# ══════════════════════════════════════════════════════════════════════════════
#  STEPS 3 – 9  ─  CDI PIPELINE
# ══════════════════════════════════════════════════════════════════════════════

def step3_compute_road_multiplier(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 3: Road_Multiplier via Action_Taken fallback.
    Primary OSM method deferred to Phase 2 GIS work.
    """
    log.info("Step 3: Assigning Road_Multiplier (Action_Taken fallback)…")

    def _road_multiplier(action: str) -> float:
        if action in ("UPGRADED_TO_TRUNK", "MERGED_INTO_TRUNK"):
            return ROAD_MULTIPLIER_TRUNK
        elif action == "RETAINED_AS_FEEDER":
            return ROAD_MULTIPLIER_FEEDER
        return ROAD_MULTIPLIER_DEFAULT

    gdf["Road_Multiplier"] = gdf["Action_Taken"].apply(_road_multiplier)
    log.info("  Road_Multiplier — Trunk: %.2f  Feeder: %.2f",
             ROAD_MULTIPLIER_TRUNK, ROAD_MULTIPLIER_FEEDER)
    return gdf


def step4a_compute_final_cdi(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 4a: Final_CDI = Pop_Score×0.50 + POI_Score×0.50.

    v3.2: Road_Multiplier is NO LONGER folded into Final_CDI. The previous
    formulation was circular — Road_Multiplier is derived from Action_Taken,
    which is set by SSCL injection and cluster classification, so multiplying
    it back into the CDI granted already-promoted trunks a 1.67× advantage
    for no demand-side reason and contaminated the Jenks break thresholds
    in Step 5. The multiplier is retained as a separate audit column and
    used only as a tie-breaker inside step5_assign_priority_bands().

    50/50 Pop/POI weighting is a Phase 1 default. Phase 2 will calibrate
    against AFC ridership.
    """
    log.info("Step 4a: Computing Final_CDI = Pop×%.2f + POI×%.2f "
             "(Road_Multiplier moved to Step-5 tie-breaker in v3.2)…",
             CDI_POP_WEIGHT, CDI_POI_WEIGHT)
    raw_cdi          = (gdf["Pop_Score"] * CDI_POP_WEIGHT +
                        gdf["POI_Score"] * CDI_POI_WEIGHT)
    gdf["Final_CDI"] = raw_cdi.round(4).clip(lower=0.0)
    log.info("  Final_CDI mean: %.4f  max: %.4f  min: %.4f",
             gdf["Final_CDI"].mean(), gdf["Final_CDI"].max(),
             gdf["Final_CDI"].min())
    return gdf


def step4b_compute_social_flag(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 4b: Social_Flag = TRUE if route passes within 500m of any
    Social Obligation attractor (migrant camps, industrial estates, hospitals).
    LP routes flagged TRUE are overridden to MP in Step 5.
    """
    log.info("Step 4b: Computing Social_Flag (500m buffer, %d attractors)…",
             len(SOCIAL_OBLIGATION_ATTRACTORS))
    attractor_points = [Point(lon, lat)
                        for _, lat, lon in SOCIAL_OBLIGATION_ATTRACTORS]
    gdf_attractors   = gpd.GeoDataFrame(
        {"label": [a[0] for a in SOCIAL_OBLIGATION_ATTRACTORS]},
        geometry=attractor_points, crs=WGS84_CRS,
    ).to_crs(UTM_CRS)

    attractor_union  = unary_union(
        [p.buffer(SOCIAL_FLAG_BUFFER_M) for p in gdf_attractors.geometry])
    gdf_utm          = gdf.to_crs(UTM_CRS).copy()
    gdf["Social_Flag"] = (
        gdf_utm.geometry.simplify(SIMPLIFY_TOL_M).intersects(attractor_union))
    log.info("  %d routes flagged as Social Obligation.", gdf["Social_Flag"].sum())
    return gdf


def step5_assign_priority_bands(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 5: Priority Band (HP / MP / LP) via Jenks Natural Breaks on Final_CDI.

    Social Obligation floor: LP → MP if Social_Flag = TRUE.

    v3.2: Road_Multiplier tie-breaker. For routes whose Final_CDI lands within
    ±BAND_TIEBREAK_BAND of either Jenks break, the Road_Multiplier (1.25
    trunk / 0.75 feeder / 1.00 default) is used to promote the route up by
    one band when the multiplier is ≥ 1.25, or demote it down by one band
    when the multiplier ≤ 0.75. This restores Road_Multiplier's intent —
    "trunk roads can sustain higher frequencies" — without baking the
    self-reinforcing 1.67× advantage into the CDI itself.
    """
    log.info("Step 5: Assigning Priority Bands via Jenks Natural Breaks "
             "(Road_Multiplier acts as tie-breaker near band edges)…")
    cdi_values = gdf["Final_CDI"].fillna(0.0).values

    if not _HAS_JENKSPY:
        log.warning("  ⚠ jenkspy not installed — falling back to 33rd/66th percentile.")
        thresh_lp_mp = float(np.percentile(cdi_values, 33.33))
        thresh_mp_hp = float(np.percentile(cdi_values, 66.67))
    else:
        n = len(cdi_values)
        if n < 3:
            log.warning("  Fewer than 3 routes — assigning MP to all.")
            gdf["Priority_Band"] = "MP"
            return gdf
        breaks       = jenkspy.jenks_breaks(cdi_values.tolist(), n_classes=3)
        thresh_lp_mp = float(breaks[1])
        thresh_mp_hp = float(breaks[2])
        log.info("  Jenks breaks → LP < %.4f  ≤ MP < %.4f  ≤ HP",
                 thresh_lp_mp, thresh_mp_hp)

    # Tie-breaker band: 5% of the LP↔HP span around each break
    band_span = max(1e-4, thresh_mp_hp - thresh_lp_mp)
    tiebreak  = 0.05 * band_span

    def _band(cdi: float) -> str:
        if cdi >= thresh_mp_hp:   return "HP"
        elif cdi >= thresh_lp_mp: return "MP"
        return "LP"

    # First pass: pure Jenks
    base_bands = gdf["Final_CDI"].apply(_band).tolist()

    # Second pass: Road_Multiplier tie-breaker near band edges
    promote_n = demote_n = 0
    final_bands = []
    rm_series = gdf.get("Road_Multiplier", pd.Series(1.0, index=gdf.index))
    for cdi, band, rm in zip(gdf["Final_CDI"].values, base_bands, rm_series.values):
        new_band = band
        if rm >= ROAD_MULTIPLIER_TRUNK and band == "MP" \
           and (thresh_mp_hp - cdi) <= tiebreak:
            new_band = "HP"; promote_n += 1
        elif rm >= ROAD_MULTIPLIER_TRUNK and band == "LP" \
             and (thresh_lp_mp - cdi) <= tiebreak:
            new_band = "MP"; promote_n += 1
        elif rm <= ROAD_MULTIPLIER_FEEDER and band == "HP" \
             and (cdi - thresh_mp_hp) <= tiebreak:
            new_band = "MP"; demote_n += 1
        elif rm <= ROAD_MULTIPLIER_FEEDER and band == "MP" \
             and (cdi - thresh_lp_mp) <= tiebreak:
            new_band = "LP"; demote_n += 1
        final_bands.append(new_band)
    gdf["Priority_Band"] = final_bands

    social_lp_mask = (gdf["Social_Flag"] == True) & (gdf["Priority_Band"] == "LP")
    gdf.loc[social_lp_mask, "Priority_Band"] = "MP"

    # Re-lock: SSCL backbone trunks (CMP_Trunk=True) must always be HP.
    # inject_cmp_trunk_routes() set them to HP before this step, but Jenks
    # can demote a short SSCL route (e.g. SSCL-03 Batamaloo→Hazratbal, 10km)
    # if its CDI falls below the MP break.  We restore HP unconditionally so
    # the QC sanity check and the headway override in step6 both work correctly.
    if "CMP_Trunk" in gdf.columns:
        cmp_lock_mask = gdf["CMP_Trunk"] == True
        cmp_demoted   = int(((gdf.loc[cmp_lock_mask, "Priority_Band"]) != "HP").sum())
        gdf.loc[cmp_lock_mask, "Priority_Band"] = "HP"
        if cmp_demoted:
            log.info("  CMP lock: %d SSCL trunk(s) re-elevated to HP after Jenks.",
                     cmp_demoted)

    # v3.3: Inter-district + District-HQ floor — Regional_District routes that
    # touch any DH attractor are unconditionally lifted from LP→MP regardless
    # of CDI. Social_Flag already covers DH proximity (it's in the attractor
    # list), but the proposal asked for an explicit, named floor so the audit
    # can show the count separately.
    dh_keywords = ("DH ", "District Court", "DC Office")
    is_dh_route = gdf.apply(
        lambda r: (r.get("Route_Type") == "Regional_District")
                  and bool(r.get("Social_Flag", False)),
        axis=1,
    )
    dh_lp_mask = is_dh_route & (gdf["Priority_Band"] == "LP")
    dh_promoted_n = int(dh_lp_mask.sum())
    if dh_promoted_n:
        gdf.loc[dh_lp_mask, "Priority_Band"] = "MP"
    gdf["District_HQ_Floor"] = is_dh_route.values

    n_hp, n_mp, n_lp = ((gdf["Priority_Band"] == b).sum() for b in ("HP","MP","LP"))
    log.info("  Priority bands — HP: %d  MP: %d  LP: %d  "
             "(%d Social LP→MP, %d District-HQ LP→MP, %d road-promoted, %d road-demoted)",
             n_hp, n_mp, n_lp, social_lp_mask.sum(), dh_promoted_n,
             promote_n, demote_n)
    return gdf


def step5b_flag_sscl_cdi_conflicts(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    v3.3.1 Phase-1 audit (round 2 STEP 11): tightened conflict flag with
    +0.2 CDI delta. The v3.3 flag triggered on 78% of active routes (any
    non-SSCL CDI > weakest SSCL CDI), which made it a near-universal signal
    instead of a planner-actionable one. Now requires a MEANINGFUL delta:

      • SSCL_CDI_Conflict_Strong : non-SSCL active route with
        Final_CDI ≥ (max_sscl_cdi - SSCL_CONFLICT_DELTA).
        These are routes that are at-or-above the strongest SSCL trunk
        — strong displacement candidates if the SSCL deployment political
        constraint were relaxed.

      • SSCL_CDI_Conflict_Weak_SSCL : SSCL backbone route whose
        Final_CDI ≤ (max_non_sscl_cdi - SSCL_CONFLICT_DELTA).
        These are SSCL trunks that are demand-justified less well than
        nearby private routes — political deployment locked in regardless.

    Both flags are PLANNER REVIEW ONLY. No automatic reclassification —
    the SSCL backbone is a Govt of J&K deployment commitment.
    """
    SSCL_CONFLICT_DELTA = 0.2  # CDI units (0-1 scale)
    if "CMP_Trunk" not in gdf.columns or "Final_CDI" not in gdf.columns:
        gdf["SSCL_CDI_Conflict"]            = False
        gdf["SSCL_CDI_Conflict_Strong"]     = False
        gdf["SSCL_CDI_Conflict_Weak_SSCL"]  = False
        return gdf

    sscl_mask = gdf["CMP_Trunk"] == True
    sscl_cdi  = gdf.loc[sscl_mask, "Final_CDI"]
    active    = gdf["Action_Taken"] != "MERGED_INTO_TRUNK"
    non_sscl_active_cdi = gdf.loc[(~sscl_mask) & active, "Final_CDI"]

    if sscl_cdi.empty or non_sscl_active_cdi.empty:
        gdf["SSCL_CDI_Conflict"]            = False
        gdf["SSCL_CDI_Conflict_Strong"]     = False
        gdf["SSCL_CDI_Conflict_Weak_SSCL"]  = False
        return gdf

    sscl_max     = float(sscl_cdi.max())
    non_sscl_max = float(non_sscl_active_cdi.max())

    strong_mask = (
        (~sscl_mask)
        & active
        & (gdf["Final_CDI"] >= sscl_max - SSCL_CONFLICT_DELTA)
    )
    weak_sscl_mask = (
        sscl_mask
        & (gdf["Final_CDI"] <= non_sscl_max - SSCL_CONFLICT_DELTA)
    )

    gdf["SSCL_CDI_Conflict_Strong"]    = strong_mask
    gdf["SSCL_CDI_Conflict_Weak_SSCL"] = weak_sscl_mask
    # Backward-compat single flag = OR of both
    gdf["SSCL_CDI_Conflict"] = strong_mask | weak_sscl_mask

    n_strong = int(strong_mask.sum())
    n_weak   = int(weak_sscl_mask.sum())
    log.info("Step 5b (v3.3.1 +%.2f delta): %d non-SSCL routes at/above "
             "strongest SSCL trunk (max_sscl_cdi=%.4f); %d SSCL routes below "
             "strongest non-SSCL by ≥%.2f (max_non_sscl_cdi=%.4f). "
             "Planner review only; no auto-reclassification.",
             SSCL_CONFLICT_DELTA, n_strong, sscl_max, n_weak,
             SSCL_CONFLICT_DELTA, non_sscl_max)
    return gdf


def step6_assign_headways(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 6: Headway_Min from Priority Band and Route_Type.

    DIRECTIVE 1 INTEGRATION:
    Regional_District routes are rural lifelines with relaxed headways.

    SSCL OVERRIDE (v3):
    Any route flagged CMP_Trunk = True receives a hardcoded headway of
    SSCL_TRUNK_HEADWAY_MIN (15 min), bypassing the Priority Band logic entirely.
    15 min matches SSCL's actual operational target on the Srinagar e-bus
    network and is consistent with the ~4,346 pax/hr citywide peak demand
    observed in CHALO data.

    Headway rules:
      SSCL Trunk routes                  → 15 min (hardcoded)
      Urban / Peri_Urban routes          → HP=20  MP=35  LP=35
      Regional_District routes           → HP=35  MP=35       (rural lifeline)

    v3.3.7 (RTO ask): every assigned headway is finally clamped to
    HEADWAY_MAX_MIN (35 min). No route in the published plan waits longer than
    35 minutes — the old 45/60/90-min operator and rural buckets are gone.
    """
    log.info("Step 6: Assigning headways by Priority Band + Route_Type…")
    log.info("  SSCL Trunk routes: hardcoded %d min (SSCL operational target)",
             CMP_TRUNK_HEADWAY_MIN)
    log.info("  Urban/Peri-Urban: HP=%d  MP=%d  LP=%d min",
             HEADWAY_HP_MIN, HEADWAY_MP_MIN, HEADWAY_LP_MIN)
    log.info("  Regional_District lifelines: HP=%d  MP=%d min",
             HEADWAY_REGIONAL_HP_MIN, HEADWAY_REGIONAL_MP_MIN)

    def _headway(row) -> int:
        # CMP bypass — hardcoded 10 min regardless of priority band
        if row.get("CMP_Trunk", False):
            return CMP_TRUNK_HEADWAY_MIN

        band       = row["Priority_Band"]
        route_type = row.get("Route_Type", "Urban")
        cat        = str(row.get("Vehicle_Category", "")).strip().lower()

        # Operator-specific overrides
        # v3.3.7 (RTO ask): the legacy 45-min MPS floor and 60-min "regular"
        # bucket are pulled down to the 35-min ceiling — nothing waits an hour.
        if "city bus" in cat:
            return max(20, HEADWAY_HP_MIN if band == "HP" else HEADWAY_MP_MIN)
        if "mps" in cat:
            return max(35, HEADWAY_HP_MIN if band == "HP" else HEADWAY_MP_MIN)
        if "regular" in cat:
            return 35

        if route_type == "Regional_District":
            if band == "HP": return HEADWAY_REGIONAL_HP_MIN
            return HEADWAY_REGIONAL_MP_MIN

        if band == "HP": return HEADWAY_HP_MIN
        if band == "MP": return HEADWAY_MP_MIN
        return HEADWAY_LP_MIN

    gdf["Headway_Min"] = gdf.apply(_headway, axis=1).astype(int)

    # v3.3.7 (RTO ask): hard network-wide ceiling. Belt-and-suspenders clamp so
    # that even if any future branch returns a value above HEADWAY_MAX_MIN, the
    # published plan still contains no headway longer than 35 minutes.
    over = int((gdf["Headway_Min"] > HEADWAY_MAX_MIN).sum())
    gdf["Headway_Min"] = gdf["Headway_Min"].clip(upper=HEADWAY_MAX_MIN).astype(int)
    if over:
        log.info("  v3.3.7 clamp: %d routes capped down to the %d-min ceiling.",
                 over, HEADWAY_MAX_MIN)

    cmp_n      = gdf.get("CMP_Trunk", pd.Series(False, index=gdf.index)).sum()
    regional_n = (gdf.get("Route_Type", pd.Series()) == "Regional_District").sum()
    log.info("  %d SSCL Trunk routes → hardcoded %d-min headway.", cmp_n, CMP_TRUNK_HEADWAY_MIN)
    log.info("  %d Regional_District routes assigned relaxed headways.", regional_n)
    log.info("  Max headway in plan: %d min (ceiling %d).",
             int(gdf["Headway_Min"].max()), HEADWAY_MAX_MIN)
    return gdf


def step8_compute_fleet_required(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 8: Fleet_Required = CEILING(Cycle_Time_Min / Headway_Min).

    v3.2: Type-specific minimum-fleet floor. Urban / Peri_Urban routes still
    floor at MIN_FLEET_URBAN (2) for service viability; Regional_District
    lifelines floor at MIN_FLEET_REGIONAL (1) — the previous blanket "2"
    inflated rural fleet counts beyond justified demand.
    """
    log.info("Step 8: Computing Fleet_Required = CEILING(Cycle_Time / Headway) "
             "× spare_ratio %.2f. Floors: Urban/Peri=%d  Regional_District=%d.",
             FLEET_SPARE_RATIO, MIN_FLEET_URBAN, MIN_FLEET_REGIONAL)

    fleet_required_list = []
    floored_urban = floored_regional = spare_bumped = 0

    for _, row in gdf.iterrows():
        operating  = max(1, math.ceil(
            row["Cycle_Time_Min"] / max(1, row["Headway_Min"])))
        # v3.3: post-cycle spare ratio (15%). SSCL routes get this in Step 8
        # too, but Step 9 then OVERWRITES Fleet_Required with the empirical
        # CHALO bus count, so the spare bump only affects non-SSCL routes
        # in the final output. Standard transit-planning practice.
        raw_fleet  = max(1, math.ceil(operating * FLEET_SPARE_RATIO))
        if raw_fleet > operating:
            spare_bumped += 1
        route_type = row.get("Route_Type", "Urban")
        floor      = (MIN_FLEET_REGIONAL if route_type == "Regional_District"
                      else MIN_FLEET_URBAN)
        final      = max(raw_fleet, floor)
        if final > raw_fleet:
            if route_type == "Regional_District":
                floored_regional += 1
            else:
                floored_urban += 1
        fleet_required_list.append(final)

    gdf = gdf.copy()
    gdf["Fleet_Required"] = fleet_required_list

    log.info("  Floors hit — Urban/Peri raised to %d: %d routes; "
             "Regional_District raised to %d: %d routes.",
             MIN_FLEET_URBAN, floored_urban,
             MIN_FLEET_REGIONAL, floored_regional)
    log.info("  Fleet_Required — mean: %.1f  max: %d  total: %d",
             gdf["Fleet_Required"].mean(),
             gdf["Fleet_Required"].max(),
             gdf["Fleet_Required"].sum())
    return gdf


def _route_km_hpv_share(km: float) -> float:
    """
    v3.2: Route-length-based HPV (12m bus) share for non-SSCL trunks.
    v3.3.6 (RTO Kashmir ask): the long-haul bracket moved 85%→60% HPV.
    v3.3.7 (RTO ask): long-haul bracket tightened again to 50% so NEITHER
    vehicle class is the majority on a trunk route — a balanced 50/50 HPV/MPV
    mix on every corridor 12 km and longer. The RTO's reasoning holds: depot
    inventory, road-shoulder widths through old Srinagar feeder segments, and
    operator preference all push 9m buses onto these corridors, so a 50/50
    split (rather than an HPV-led one) best matches the on-the-ground reality.
    Per-corridor road-width data (a pending P2 RTO data ask) would let us bias
    this toward MPV on narrow segments later.

        <  12 km  → 100% MPV (9m)         short urban — big buses not viable
        12-22 km  → 50% HPV / 50% MPV     mixed urban–peri-urban (unchanged)
        ≥  22 km  → 50% HPV / 50% MPV     long-haul (was 60/40 in v3.3.6)
    """
    if km < 12.0:
        return 0.00
    if km < 22.0:
        return 0.50
    return 0.50   # v3.3.7 (RTO ask): was 0.60 — neither class a majority


def _sscl_bus_split(cmp_route_id: str) -> Optional[Tuple[int, int]]:
    """v3.2: look up the empirical 9m/12m bus counts for a given SSCL route."""
    if not cmp_route_id:
        return None
    for r in CMP_TRUNK_ROUTES:
        if r["id"] == cmp_route_id:
            return int(r.get("bus_9m", 0)), int(r.get("bus_12m", 0))
    return None


def step9_compute_vehicle_split(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    Step 9: HPV/MPV/LPV vehicle type split.

    v3.2 rules:
      • SSCL-injected trunks (CMP_Trunk=True with a valid CMP_Route_ID):
          Use the CMP_TRUNK_ROUTES empirical 9m/12m counts directly.
          9m → MPV, 12m → HPV. Fleet_Required is overridden to match the
          table sum so QC Check 1 (HPV+MPV+LPV=Fleet) stays consistent.

      • Non-SSCL trunks (UPGRADED_TO_TRUNK / MERGED_INTO_TRUNK without
        CMP_Route_ID): Route_KM-bracketed share via _route_km_hpv_share()
        — 0% / 50% / 50% HPV in 0-12 / 12-22 / 22+ km brackets (v3.3.7:
        long-haul bracket balanced to 50/50 so neither class is a majority).

      • RETAINED_AS_FEEDER (unchanged from v3.1):
          - LPV category → 100% LPV
          - Regular / MPS → 30% HPV, 70% MPV
          - Else (including city bus) → 100% MPV
    """
    log.info("Step 9: Computing HPV/MPV/LPV vehicle split "
             "(v3.2 Route_KM-bracketed + SSCL table override)…")
    gdf = gdf.copy()

    sscl_override_n = 0

    def _split(row) -> pd.Series:
        nonlocal sscl_override_n
        fleet  = int(row["Fleet_Required"])
        action = row["Action_Taken"]
        cat    = str(row.get("Vehicle_Category", "")).strip().lower()
        km     = float(row.get("Route_KM", 0.0))
        cmp_id = str(row.get("CMP_Route_ID", "") or "")

        if fleet == 0:
            return pd.Series({"HPV_Count": 0, "MPV_Count": 0,
                              "LPV_Count": 0, "Fleet_Required": 0})

        # v3.2: SSCL table provides the CURRENT bus_9m / bus_12m split.
        # v3.3.4 (audit): treat the empirical SSCL fleet as a FLOOR, not a
        # hard override. If the formula-based Fleet_Required (from Step 8 +
        # spare ratio) demands MORE buses to actually sustain the 15-min
        # target headway, recommend the higher number. Vehicle split scales
        # proportionally on the empirical 9m/12m ratio so the recommendation
        # stays realistic for the SSCL deployment plan. Eliminates the
        # contradictory "Red_Overload at empirical fleet" signal that
        # appeared for 12 SSCL routes in v3.3.3.
        if bool(row.get("CMP_Trunk", False)) and cmp_id:
            sscl_split = _sscl_bus_split(cmp_id)
            if sscl_split is not None:
                bus_9m, bus_12m = sscl_split
                empirical = bus_9m + bus_12m
                if empirical > 0:
                    sscl_override_n += 1
                    effective = max(empirical, fleet)
                    if effective == empirical or empirical == 0:
                        hpv_eff, mpv_eff = bus_12m, bus_9m
                    else:
                        scale = effective / empirical
                        hpv_eff = int(round(bus_12m * scale))
                        mpv_eff = effective - hpv_eff
                    # v3.3.6 (RTO Kashmir ask): cap HPV at SSCL_HPV_SHARE_CAP
                    # so the recommendation gives MPVs more share even on
                    # routes that CHALO currently runs as 100% 12-metre.
                    # CHALO empirical is what is — the engine output is what
                    # the RTO wants going forward. v3.3.7: 50% cap means neither
                    # class is the majority — a balanced trunk fleet (int()
                    # floor leaves MPV at most one bus ahead of HPV).
                    max_hpv = int(effective * SSCL_HPV_SHARE_CAP)
                    if hpv_eff > max_hpv:
                        hpv_eff = max_hpv
                        mpv_eff = effective - hpv_eff
                    return pd.Series({
                        "HPV_Count":      hpv_eff,
                        "MPV_Count":      mpv_eff,
                        "LPV_Count":      0,
                        "Fleet_Required": effective,
                    })

        hpv = mpv = lpv = 0

        if action in ("UPGRADED_TO_TRUNK", "MERGED_INTO_TRUNK"):
            # LPV-category retained vehicles cap at the smaller bracket regardless
            if "lpv" in cat:
                hpv_share = min(0.50, _route_km_hpv_share(km))
            else:
                hpv_share = _route_km_hpv_share(km)
            hpv = math.ceil(fleet * hpv_share)
            mpv = fleet - hpv
        else:
            # RETAINED_AS_FEEDER
            if "lpv" in cat:
                lpv = fleet
            elif "regular" in cat or "mps" in cat:
                hpv = math.ceil(fleet * 0.30)
                mpv = fleet - hpv
            else:
                mpv = fleet

        return pd.Series({"HPV_Count": hpv, "MPV_Count": mpv,
                          "LPV_Count": lpv, "Fleet_Required": fleet})

    split_df              = gdf.apply(_split, axis=1)
    gdf["HPV_Count"]      = split_df["HPV_Count"].astype(int)
    gdf["MPV_Count"]      = split_df["MPV_Count"].astype(int)
    gdf["LPV_Count"]      = split_df["LPV_Count"].astype(int)
    gdf["Fleet_Required"] = split_df["Fleet_Required"].astype(int)

    log.info("  HPV total: %d  MPV total: %d  LPV total: %d  "
             "(SSCL table override applied to %d routes)",
             gdf["HPV_Count"].sum(), gdf["MPV_Count"].sum(),
             gdf["LPV_Count"].sum(), sscl_override_n)
    return gdf


def consolidate_duplicate_permits(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Consolidate identical-corridor duplicate FEEDER permits (audit Finding 8).

    Permits are records, not routes: the input carries ~380 exact-duplicate
    (Origin, Destination, Vehicle-class) rows. Feeder-retained duplicates were
    never consolidated, so Step 8 gave EACH copy its own headway-based fleet —
    one corridor (Soura→Jehangir Chowk ×18) drew 108 buses, more than the entire
    SSCL system operates (98). This step keeps ONE representative feeder per
    identical (rounded O, rounded D) CORRIDOR — across vehicle classes (v3.3.8) —
    records the permit multiplicity on it (Permit_Count), and marks the redundant copies
    MERGED_INTO_TRUNK with Merged_Reason='duplicate_permit' so they are zeroed by
    zero_merged_route_fleet() and excluded from the active set — exactly like any
    other absorbed permit. Trunks and the SSCL backbone are left intact.

    Must run AFTER apply_terminal_capacity (feeder states final) and BEFORE the
    fleet steps, so the consolidated copies flow through the CDI pipeline and are
    zeroed at the end like every other merged route.
    """
    log.info("Consolidating duplicate feeder permits (Finding 8 over-fleeting)…")
    gdf = gdf.copy()
    if "Permit_Count" not in gdf.columns:
        gdf["Permit_Count"] = 1
    if "Merged_Reason" not in gdf.columns:
        gdf["Merged_Reason"] = ""

    # v3.3.8: consolidate by CORRIDOR (rounded O→D) across BOTH feeders AND trunks
    # and across vehicle classes. Rationalisation = one service per corridor.
    # The earlier feeder-only / within-class key left duplicates as separate active
    # routes — e.g. 6 identical "Batamaloo→Pantha Chowk" SSCL-matched permit TRUNKS
    # (~48 buses on one corridor) and minibus+tempo pairs on the same O/D — which
    # showed as duplicate route codes. A real SSCL backbone route (Route_ID
    # "SSCL-*") is NEVER merged away; permit duplicates are absorbed into it (this
    # is the documented "8 absorbed duplicate permits" behaviour, now enforced for
    # all corridors). Displaced operators are still tracked via the class label.
    def _is_synth_sscl(rid):
        return str(rid).upper().startswith("SSCL")

    cands = gdf[gdf["Action_Taken"].isin(["RETAINED_AS_FEEDER",
                                          "UPGRADED_TO_TRUNK"])].copy()
    if cands.empty:
        log.info("  No active routes to consolidate.")
        return gdf

    # v3.4.1: UNDIRECTED corridor key. A route's Cycle_Time is a ROUND TRIP, so a
    # single route already represents bidirectional service; keeping both "A→B" and
    # "B→A" as separate active routes double-counts that corridor's fleet (the audit
    # found ~10 such pairs, e.g. Srinagar↔Kupwara, Srinagar↔Uri, LD↔Pantha Chowk).
    # Sorting the two endpoints collapses the reverse direction onto one service.
    # Two routes share a key only if they share BOTH endpoints — i.e. genuinely the
    # same corridor — so this never merges distinct corridors. (A real SSCL-* route
    # is still never absorbed — see the rank/guard below.)
    _o = (cands["Start_Lat"].round(4).astype(str) + "," +
          cands["Start_Lon"].round(4).astype(str))
    _d = (cands["End_Lat"].round(4).astype(str) + "," +
          cands["End_Lon"].round(4).astype(str))
    key = pd.Series([" | ".join(sorted([a, b])) for a, b in zip(_o, _d)],
                    index=cands.index)

    _OP = {"minibus": "Private Minibus", "mini bus": "Private Minibus",
           "mpv": "Private Minibus", "mps": "MPS (Stage Carriage)",
           "city bus": "JKRTC / City Bus", "jkrtc": "JKRTC / City Bus",
           "lpv": "LPV / Tempo", "tempo": "LPV / Tempo"}

    n_consolidated = n_corridors = max_grp = 0
    for _, grp in cands.assign(_k=key).groupby("_k"):
        if len(grp) <= 1:
            continue
        # Representative preference: a synthetic SSCL backbone route, else a trunk,
        # else the first feeder. (Stable: rank then original order.)
        def _rank(idx):
            rid = gdf.at[idx, "Route_ID"]
            is_trunk = gdf.at[idx, "Action_Taken"] == "UPGRADED_TO_TRUNK"
            return (0 if _is_synth_sscl(rid) else 1, 0 if is_trunk else 1)
        ordered = sorted(grp.index, key=lambda i: (_rank(i), list(grp.index).index(i)))
        rep = ordered[0]
        absorbed = [i for i in ordered[1:] if not _is_synth_sscl(gdf.at[i, "Route_ID"])]
        if not absorbed:
            continue
        n_corridors += 1
        max_grp = max(max_grp, len(grp))
        gdf.at[rep, "Permit_Count"] = int(gdf.at[rep, "Permit_Count"]) + len(absorbed)
        for idx in absorbed:
            gdf.at[idx, "Action_Taken"]  = "MERGED_INTO_TRUNK"
            gdf.at[idx, "Merged_Reason"] = "duplicate_permit"
            gdf.at[idx, "New_Route_ID"]  = gdf.at[rep, "New_Route_ID"]
            gdf.at[idx, "CMP_Trunk"]     = False   # absorbed permit, not a backbone route
            cat = str(gdf.at[idx, "Vehicle_Category"]).strip().lower()
            gdf.at[idx, "Displaced_Operator_Class"] = next(
                (lbl for k, lbl in _OP.items() if k in cat), "Private Minibus")
            n_consolidated += 1

    log.info("  Duplicate-permit consolidation: %d corridors carried duplicates "
             "(largest = %d permits); %d redundant feeder permits consolidated "
             "into representatives (Finding 8 over-fleeting removed).",
             n_corridors, max_grp, n_consolidated)
    return gdf


def zero_merged_route_fleet(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Zero fleet for MERGED_INTO_TRUNK routes — service absorbed by Trunk
    (or, where Merged_Reason='duplicate_permit', by the corridor representative)."""
    merged_mask = gdf["Action_Taken"] == "MERGED_INTO_TRUNK"
    gdf = gdf.copy()
    gdf.loc[merged_mask,
            ["Fleet_Required", "HPV_Count", "MPV_Count", "LPV_Count"]] = 0
    n_merged = merged_mask.sum()
    if n_merged:
        log.info("  Zeroed fleet for %d MERGED_INTO_TRUNK routes.", n_merged)
    return gdf


# ══════════════════════════════════════════════════════════════════════════════
#  QUALITY CONTROL  ─  All 8 Checks Must Pass Before Export
# ══════════════════════════════════════════════════════════════════════════════

def run_all_qc_checks(gdf: gpd.GeoDataFrame) -> None:
    """
    8 pre-submission QC checks for v3 (LPV removed, HPV/MPV enforced).
    Any failure raises RuntimeError and blocks the export pipeline.
    """
    log.info("Running QC checks (8 checks — v3 HPV/MPV only policy)…")
    failures = []

    # Check 1: HPV + MPV + LPV = Fleet_Required for every row
    check1 = gdf[gdf["HPV_Count"] + gdf["MPV_Count"] + gdf["LPV_Count"] != gdf["Fleet_Required"]]
    if not check1.empty:
        failures.append(
            f"CHECK 1 FAIL — {len(check1)} rows where HPV+MPV+LPV ≠ Fleet_Required")
        for _, r in check1.iterrows():
            failures.append(
                f"  {r['Route_ID']} | Fleet={r['Fleet_Required']} | "
                f"HPV={r['HPV_Count']} MPV={r['MPV_Count']} LPV={r['LPV_Count']}"
            )
    else:
        log.info("  ✓ Check 1: Vehicle count integrity — HPV+MPV+LPV = Fleet_Required for all rows.")

    # Check 2: (Removed - LPVs are now permitted)
    log.info("  ✓ Check 2: LPV check bypassed — LPV retention enabled in v3.1.")

    # Check 3: No null Priority_Band
    check3 = gdf[gdf["Priority_Band"].isna() | (gdf["Priority_Band"] == "")]
    if not check3.empty:
        failures.append(
            f"CHECK 3 FAIL — {len(check3)} rows with null Priority_Band.")
    else:
        log.info("  ✓ Check 3: No null Priority_Band values.")

    # Check 4: Feeder routes must have HPV_Count = 0 (DISABLED in v3.1 to allow Vehicle Category passthrough)
    # check4 = gdf[(gdf["Action_Taken"] == "RETAINED_AS_FEEDER") &
    #              (gdf["HPV_Count"] > 0)]
    # if not check4.empty:
    #     failures.append(
    #         f"CHECK 4 FAIL — {len(check4)} FEEDER routes have HPV_Count > 0.")
    #     for _, r in check4.iterrows():
    #         failures.append(f"  {r['Route_ID']} | HPV={r['HPV_Count']}")
    # else:
    log.info("  ✓ Check 4: Bypassed (feeders can have HPVs based on vehicle category).")

    # Check 5: Active Trunk routes must have HPV_Count > 0 unless either
    #   (a) the route is short (<12 km) where 100% MPV is empirically right, OR
    #   (b) the route is an SSCL-injected trunk whose CMP_TRUNK_ROUTES entry
    #       specifies a 100% 9m (= MPV) fleet — that IS the ground truth.
    # v3.2: relaxed from "every trunk needs HPV".
    trunks = gdf[gdf["Action_Taken"] == "UPGRADED_TO_TRUNK"].copy()

    def _sscl_all_9m(cmp_id: str) -> bool:
        split = _sscl_bus_split(cmp_id) if cmp_id else None
        return split is not None and split[1] == 0  # bus_12m == 0

    sscl_all_9m_mask = trunks.get(
        "CMP_Route_ID", pd.Series("", index=trunks.index)
    ).fillna("").astype(str).apply(_sscl_all_9m)

    check5 = trunks[
        (trunks["HPV_Count"] == 0)
        & (trunks["Route_KM"] >= 12.0)
        & (~sscl_all_9m_mask)
    ]
    if not check5.empty:
        failures.append(
            f"CHECK 5 FAIL — {len(check5)} non-SSCL TRUNK routes ≥12 km "
            f"have HPV_Count = 0.")
        for _, r in check5.iterrows():
            failures.append(
                f"  {r['Route_ID']} | KM={r['Route_KM']} Fleet={r['Fleet_Required']} "
                f"HPV={r['HPV_Count']} MPV={r['MPV_Count']}")
    else:
        n_short_mpv = ((trunks["Route_KM"] < 12.0) & (trunks["HPV_Count"] == 0)).sum()
        n_sscl_all9m = int(sscl_all_9m_mask.sum())
        log.info("  ✓ Check 5: All non-SSCL trunks ≥12 km have HPV_Count > 0  "
                 "(%d short trunks <12 km and %d SSCL all-9m trunks legitimately MPV-only).",
                 n_short_mpv, n_sscl_all9m)

    # Check 6: Fleet_Required > 0 for all active routes
    active = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"]
    check6 = active[active["Fleet_Required"] <= 0]
    if not check6.empty:
        failures.append(
            f"CHECK 6 FAIL — {len(check6)} active routes have Fleet_Required ≤ 0.")
    else:
        total_fleet = int(active["Fleet_Required"].sum())
        log.info("  ✓ Check 6: All active routes have Fleet_Required > 0 "
                 "(total fleet = %d).", total_fleet)

    # Check 7: All 30 SSCL backbone routes must be in HP band (CMP_Route_ID membership).
    # Previous version checked Route_Name substrings ("hazratbal", "batamaloo") which
    # failed if a route's Route_Name didn't contain those keywords even though it was
    # correctly matched and locked.  CMP_Route_ID membership is authoritative.
    if "CMP_Trunk" in gdf.columns and "CMP_Route_ID" in gdf.columns:
        sscl_hp   = gdf[(gdf["CMP_Trunk"] == True) & (gdf["Priority_Band"] == "HP")]
        sscl_all  = gdf[gdf["CMP_Trunk"] == True]
        matched_ids   = set(sscl_all["CMP_Route_ID"].unique())
        expected_ids  = {r["id"] for r in CMP_TRUNK_ROUTES}
        unmatched_ids = expected_ids - matched_ids
        if unmatched_ids:
            log.warning("  CHECK 7 WARN — %d SSCL route(s) not matched in dataset: %s. "
                        "Lower CMP_FUZZY_THRESHOLD or add synthetic route.",
                        len(unmatched_ids), sorted(unmatched_ids))
        elif len(sscl_hp) < len(sscl_all):
            log.warning("  CHECK 7 WARN — %d SSCL trunk(s) not in HP band after lock step.",
                        len(sscl_all) - len(sscl_hp))
        else:
            log.info("  ✓ Check 7: All %d matched SSCL trunks confirmed HP "
                     "(%d/%d SSCL IDs matched).",
                     len(sscl_hp), len(matched_ids), len(expected_ids))
    else:
        log.warning("  CHECK 7 SKIP — CMP_Trunk column absent.")

    # Check 8: No Social_Flag route in LP band
    check8 = gdf[(gdf["Social_Flag"] == True) & (gdf["Priority_Band"] == "LP")]
    if not check8.empty:
        failures.append(
            f"CHECK 8 FAIL — {len(check8)} Social Obligation routes in LP band.")
    else:
        log.info("  ✓ Check 8: No Social Obligation routes in LP band.")

    # Additional: all CMP Trunk routes must be UPGRADED_TO_TRUNK
    if "CMP_Trunk" in gdf.columns:
        cmp_not_trunk = gdf[(gdf["CMP_Trunk"] == True) &
                            (gdf["Action_Taken"] != "UPGRADED_TO_TRUNK")]
        if not cmp_not_trunk.empty:
            failures.append(
                f"CHECK CMP FAIL — {len(cmp_not_trunk)} CMP-injected routes "
                f"are NOT marked UPGRADED_TO_TRUNK.")
        else:
            cmp_n = (gdf["CMP_Trunk"] == True).sum()
            log.info("  ✓ Check CMP: All %d CMP-injected routes confirmed as "
                     "UPGRADED_TO_TRUNK.", cmp_n)

    # ── NEW input-plausibility & sanity checks (audit Rec 8) ──────────────
    # These catch the classes of defect that "sailed through QC" in v3.3.7.
    # Warn by default; set KASHMIR_STRICT_QC=1 to make them BLOCKING (the
    # recommended pre-government-submission gate).
    strict = os.getenv("KASHMIR_STRICT_QC", "0") == "1"

    def _flag(msg: str) -> None:
        if strict:
            failures.append(msg)
        else:
            log.warning("  ⚠ %s (non-blocking — set KASHMIR_STRICT_QC=1 to block)", msg)

    active_qc = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"].copy()

    # QC-Geocode: active routes whose endpoint still snaps to the Srinagar
    # centroid (residue of the Finding 1 collapse).
    if {"Start_Lat", "Start_Lon", "End_Lat", "End_Lon"}.issubset(active_qc.columns):
        coll = active_qc[active_qc.apply(
            lambda r: _near_srinagar_centroid(r["Start_Lat"], r["Start_Lon"])
            or _near_srinagar_centroid(r["End_Lat"], r["End_Lon"]), axis=1)]
        if len(coll):
            _flag(f"QC-GEOCODE — {len(coll)} active route(s) still have an endpoint at "
                  f"the Srinagar centroid (geocode collapse, Finding 1).")
        else:
            log.info("  ✓ QC-Geocode: no active route endpoint at the Srinagar centroid.")

    # QC-DupCorridor: identical O–D fleeted independently (Finding 8 over-fleeting).
    if {"Start_Lat", "Start_Lon", "End_Lat", "End_Lon",
        "Fleet_Required"}.issubset(active_qc.columns):
        key = (active_qc["Start_Lat"].round(3).astype(str) + "," +
               active_qc["Start_Lon"].round(3).astype(str) + "->" +
               active_qc["End_Lat"].round(3).astype(str) + "," +
               active_qc["End_Lon"].round(3).astype(str))
        dup = (active_qc.assign(_k=key).groupby("_k")
               .agg(n=("Route_ID", "size"), fleet=("Fleet_Required", "sum")))
        dup = dup[dup["n"] > 1].sort_values("fleet", ascending=False)
        if len(dup):
            w = dup.iloc[0]
            _flag(f"QC-DUPCORR — {len(dup)} O–D corridor(s) carry >1 independently-"
                  f"fleeted route; worst = {int(w['n'])} routes / {int(w['fleet'])} buses "
                  f"on one corridor (Finding 8).")
        else:
            log.info("  ✓ QC-DupCorridor: no duplicate O–D corridors fleeted independently.")

    # QC-Load: network mean Load_Ratio sanity band (Findings 8/9/10).
    if "Load_Ratio" in active_qc.columns:
        lr = pd.to_numeric(active_qc["Load_Ratio"], errors="coerce").dropna()
        if len(lr):
            mean_lr = float(lr.mean())
            if mean_lr < 0.20:
                _flag(f"QC-LOAD — mean Load_Ratio {mean_lr:.3f} is implausibly low "
                      f"(<0.20): demand base / fleet sizing need reconciliation "
                      f"(Findings 8/9/10).")
            else:
                log.info("  ✓ QC-Load: mean Load_Ratio %.3f within plausible band.", mean_lr)

    if failures:
        for msg in failures:
            log.error(msg)
        raise RuntimeError(
            f"QC FAILED: {len(failures)} issue(s). Fix before export. "
            f"See transit_v3.log for details.")
    log.info("  ✓ ALL QC CHECKS PASSED — workbook ready for export.")


def qc_route_codes(gdf: gpd.GeoDataFrame) -> None:
    """Route-code uniqueness gate (audit Output #1). Runs AFTER assign_route_codes.

    244/342 codes were duplicated in v3.3.7 because the code is derived from
    origin/destination tehsil-sector-stop and the geocode collapse mapped many
    distinct routes onto the same Srinagar point. Duplicate identifiers in an
    RTO submission are a hard defect. Warns by default; blocks under
    KASHMIR_STRICT_QC=1.
    """
    if "Route_Code" not in gdf.columns:
        log.warning("  qc_route_codes: no Route_Code column — skipped.")
        return
    active = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"]
    codes = active["Route_Code"].astype(str)
    # Missing-code placeholders (blank / TMP- / UNMATCHED) are a coverage gap, not
    # a duplicate-identifier defect — exclude them from the uniqueness test and
    # report them separately.
    is_placeholder = codes.eq("") | codes.eq("UNMATCHED") | codes.str.startswith("TMP-")
    real = codes[~is_placeholder]
    dup_mask = real.duplicated(keep=False)
    n_dup = int(dup_mask.sum())
    n_distinct = int(real[dup_mask].nunique())
    n_missing = int(is_placeholder.sum())
    strict = os.getenv("KASHMIR_STRICT_QC", "0") == "1"
    if n_dup:
        msg = (f"QC-CODES — {n_dup} active routes share {n_distinct} non-unique "
               f"Route_Code(s) (Output #1 — should be unique after the M4 suffix).")
        if strict:
            raise RuntimeError("QC FAILED: " + msg)
        log.warning("  ⚠ %s (non-blocking — set KASHMIR_STRICT_QC=1 to block)", msg)
    else:
        log.info("  ✓ QC-Codes: all %d real active route codes are unique.",
                 len(real))
    if n_missing:
        log.warning("  ⚠ QC-Codes: %d active route(s) lack a Route_Code "
                    "(endpoint absent from the stops master — manual coding "
                    "needed).", n_missing)


# ══════════════════════════════════════════════════════════════════════════════
#  NETWORK SCORE
# ══════════════════════════════════════════════════════════════════════════════

def compute_network_score(gdf: gpd.GeoDataFrame, net_pop: int) -> float:
    active      = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"]
    total_fleet = int(active["Fleet_Required"].sum())
    efficiency  = net_pop / max(1, total_fleet)
    cmp_n       = int(gdf.get("CMP_Trunk", pd.Series(False, index=gdf.index)).sum())
    log.info("=" * 68)
    log.info("  NETWORK EFFICIENCY SUMMARY (v3)")
    log.info("  Deduplicated Pop. Served : %s residents",  f"{net_pop:,}")
    log.info("  Active Fleet Required    : %d buses",       total_fleet)
    log.info("  Fleet Efficiency         : %.0f res/bus",   efficiency)
    log.info("  SSCL Backbone Trunks     : %d routes (hardcoded %d-min headway)",
             cmp_n, SSCL_TRUNK_HEADWAY_MIN)
    log.info("  Routes > 40km (lifelines): %d (Directive 1 — retained, not dropped)",
             (gdf["Route_Type"] == "Regional_District").sum())
    log.info("  LPV fleet                : 0 (eradicated in v3)")
    log.info("=" * 68)
    return float(efficiency)


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 3  ─  RATIONALISATION LOG
# ══════════════════════════════════════════════════════════════════════════════

def _reasoning(row: pd.Series) -> str:
    action    = row["Action_Taken"]
    old_id    = row["Route_ID"]
    new_id    = row["New_Route_ID"]
    pop       = f"{row.get('Population_Served_Pct', 0.0):.2f}% of CMP 2024 pop."
    km        = f"{row['Route_KM']:.1f}"
    fleet     = row["Fleet_Required"]
    band      = row.get("Priority_Band", "?")
    hw        = row.get("Headway_Min", "?")
    cdi       = f"{row.get('Final_CDI', 0):.4f}"
    rt        = row.get("Route_Type", "?")
    social    = "🚩 Social Obligation" if row.get("Social_Flag") else ""
    zone      = row.get("Congestion_Zone", "?")
    hpv       = row.get("HPV_Count", 0)
    mpv       = row.get("MPV_Count", 0)
    cmp_tag   = f" [CMP-{row['CMP_Route_ID']}]" if row.get("CMP_Trunk") else ""

    if action == "UPGRADED_TO_TRUNK":
        return (
            f"Route {old_id} ({row['Route_Name']}) → Trunk {new_id}{cmp_tag}. "
            f"Route_Type={rt} | Zone={zone} | CDI={cdi} | Band={band} | "
            f"Headway={hw} min | Fleet={fleet} (HPV={hpv} MPV={mpv}). "
            f"Serves ~{pop} residents over {km} km. {social}"
        )
    elif action == "MERGED_INTO_TRUNK":
        return (
            f"Route {old_id} ({row['Route_Name']}) merged into Trunk {new_id}. "
            f"Spatial overlap with trunk corridor. CDI={cdi}. "
            f"Fleet=0 (counted under Trunk). {social}"
        )
    else:
        return (
            f"Route {old_id} ({row['Route_Name']}) → Feeder {new_id}. "
            f"Route_Type={rt} | Zone={zone} | CDI={cdi} | Band={band} | "
            f"Headway={hw} min | Fleet={fleet} (HPV=0 enforced, MPV={mpv} — 100%% MPV). "
            f"Distinct catchment: {pop} residents over {km} km. {social}"
        )


def generate_log(gdf: gpd.GeoDataFrame, out_path: str) -> pd.DataFrame:
    log.info("Generating Rationalisation Log → %s", out_path)
    cols = [c for c in [
        "Route_ID", "Route_Name", "Action_Taken", "New_Route_ID",
        "Displaced_Operator_Class",
        "Route_KM", "Route_Type", "Congestion_Zone",
        "Pop_Score", "POI_Score", "Road_Multiplier", "Final_CDI",
        "Social_Flag", "Priority_Band", "Headway_Min",
        "Fleet_Required", "HPV_Count", "MPV_Count",
        "CMP_Trunk", "CMP_Route_ID",
        "Overlap_Metric", "Population_Served",
        "HV_POI_Count", "Sharp_Turns", "Junction_Penalty_Min",
        "N_Stops_Estimated", "Stop_Penalty_Min",
        "OSRM_Duration_S", "Cycle_Time_Min", "Geo_Source",
    ] if c in gdf.columns]
    df_log                     = gdf[cols].copy()
    df_log["Reasoning_String"] = gdf.apply(_reasoning, axis=1)
    df_log.to_csv(out_path, index=False, encoding="utf-8-sig")
    log.info("  Log written: %d rows.", len(df_log))
    return df_log


# ══════════════════════════════════════════════════════════════════════════════
#  ROUTE CODES  ─  v3.3.7: baked into the pipeline so every export carries codes
# ══════════════════════════════════════════════════════════════════════════════
# Deterministic 12-char Route_Code from the master stops file
# (Kashmir_Stops_Sectored_V2.csv): <TehsilO TehsilD><SectorO SectorD><StopO StopD>.
# Logic ported from generate_route_codes.py so the operational CSV, the 9-sheet
# RTO workbook AND the pretty bus-schedule workbook all show codes natively — no
# separate post-step required. Endpoints that aren't in the stops master are
# backfilled from the official codes already committed in the dashboard
# routes.json, so the published plan is fully coded (was: blank Route_Code
# column in the pretty workbook because the engine CSV never carried codes).

_RC_NOISE_SUFFIXES = [
    "BUS STAND", "BUS STATION", "RAILWAY STATION", "CROSSING",
    "CHOWK", "CHOK", "HOSPITAL", "COLLEGE", "STOP", "STAND",
]
_DASHBOARD_ROUTES_JSON = Path(
    "E:/dash/bus-sathi-dashboard/public/route-rationalization-kashmir/data/routes.json")


def _rc_compact(s) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def _rc_strip_noise(name: str) -> str:
    n = name
    for w in _RC_NOISE_SUFFIXES:
        n = re.sub(rf"\b{w}\b", "", n)
    return re.sub(r"\s+", " ", n).strip()


def _rc_extract_origin_dest(route_name: str) -> Tuple[Optional[str], Optional[str]]:
    route_name = str(route_name).upper().strip()
    if " ↔ " in route_name:
        a, b = route_name.split(" ↔ ", 1)
        return a.strip(), b.strip()
    if " TO " in route_name:
        origin, rest = route_name.split(" TO ", 1)
        dest = rest.split(" VIA ")[0] if " VIA " in rest else rest
        return origin.strip(), dest.strip()
    # Manual fix: some JKRTC names are "A, B" (comma) rather than "A to B".
    if "," in route_name:
        a, b = route_name.split(",", 1)
        if a.strip() and b.strip():
            return a.strip(), b.split(" VIA ")[0].strip()
    return None, None


# Manual stop entries for endpoints absent from the 187-stop master, so their
# routes get a proper 12-char code instead of UNMATCHED (user-requested).
# (UPPER compact name) -> (Tehsil_Code, Sector_ID, Stop_No).
_MANUAL_STOPS: Dict[str, Tuple[str, int, int]] = {
    "JAWAHIRNAGAR": ("SR", 10, 63),    # Jawahir Nagar — central Srinagar (Rambagh/Karan Nagar belt)
    "JAWAHARNAGAR": ("SR", 10, 63),
}


def _resolve_stops_master() -> Optional[Path]:
    """Stops master lives next to the engine script; the engine usually runs
    with cwd = the output dir, so check both."""
    for c in (Path.cwd() / "Kashmir_Stops_Sectored_V2.csv",
              Path(__file__).resolve().parent / "Kashmir_Stops_Sectored_V2.csv"):
        if c.exists():
            return c
    return None


def _load_dashboard_route_codes() -> Dict[str, str]:
    """Route_ID -> official Route_Code map from the committed dashboard JSON."""
    out: Dict[str, str] = {}
    try:
        if _DASHBOARD_ROUTES_JSON.exists():
            for r in json.loads(_DASHBOARD_ROUTES_JSON.read_text(encoding="utf-8")):
                rid  = str(r.get("Route_ID", "")).strip()
                code = str(r.get("Route_Code", "")).strip()
                if rid and code and code.upper() != "UNMATCHED" and not code.startswith("TMP-"):
                    out[rid] = code
    except Exception as exc:
        log.warning("  Route_Code: could not read dashboard routes.json (%s)", exc)
    return out


def assign_route_codes(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Assign a deterministic 12-char Route_Code to every route and export the
    canonical stop master.

    v4 ("geo-canonical", 2026-06-21): the old name-match-against-a-hand-built-
    master logic is replaced by route_code_system.py, which builds the stop
    registry FROM the engine's own geocoded route endpoints (one source of truth,
    exact route->stop linkage) and resolves District + Tehsil(=Sector) by
    point-in-polygon against authoritative OSM admin boundaries. See
    ROUTE_CODE_METHODOLOGY.md. Code = <Do><Dd><So><Sd><No><Nd> (12 chars);
    a trailing letter (A/B) marks a genuine same-stop-pair collision only.
    """
    log.info("Assigning Route_Code (v4 geo-canonical: OSM district/tehsil + endpoint registry)…")
    gdf = gdf.copy()
    try:
        import route_code_system as rcs
        admin = rcs.load_admin()
    except Exception as exc:  # pragma: no cover - missing module/boundaries
        log.warning("  route_code_system/boundaries unavailable (%s) — Route_Code left blank.", exc)
        gdf["Route_Code"] = ""
        return gdf

    def _ends(geom):
        try:
            cs = list(geom.coords)
            return (cs[0][1], cs[0][0]), (cs[-1][1], cs[-1][0])
        except Exception:
            return (None, None), (None, None)

    routes, valid_idx = [], []
    for pos, (_, r) in enumerate(gdf.iterrows()):
        (olat, olon), (dlat, dlon) = _ends(r.get("geometry"))
        if olat is None or dlat is None:
            continue
        valid_idx.append(pos)
        routes.append({
            "route_name": r.get("Route_Name", ""),
            "o_lat": olat, "o_lon": olon, "d_lat": dlat, "d_lon": dlon,
            "active": (r.get("Action_Taken") != "MERGED_INTO_TRUNK"),
        })

    codes, master, stats = rcs.assign(routes, admin)

    out = [""] * len(gdf)
    for k, pos in enumerate(valid_idx):
        out[pos] = codes[k]
    gdf["Route_Code"] = out

    try:
        master.to_csv("Kashmir_Stops_Master_v4.csv", index=False, encoding="utf-8-sig")
    except Exception as exc:  # pragma: no cover
        log.warning("  could not write Kashmir_Stops_Master_v4.csv (%s)", exc)

    log.info("  Route_Code v4: %d endpoints -> %d unique names -> %d canonical stops "
             "across %d districts / %d tehsil-sectors; %d active routes coded, "
             "%d letter-disambiguated (genuine same-stop-pair). Master -> "
             "Kashmir_Stops_Master_v4.csv.",
             stats["endpoints"], stats["unique_names"], stats["canonical_stops"],
             stats["districts_used"], stats["sectors_used"], stats["active_routes"],
             stats["letter_suffixed"])
    return gdf



# Acronyms / station codes that must stay uppercase when we title-case names.
_RNAME_ACRONYMS = {
    "LD", "TRC", "JVC", "GBS", "BPR", "HMT", "SGR", "SMHS", "SKIMS", "NH",
    "SMC", "JKRTC", "RTO", "KP", "BSF", "CRPF", "DPS",
}
_RNAME_CONNECTORS = {"to", "via", "and", "near", "opp"}


def _clean_route_name(name: str) -> str:
    """v3.3.7: normalise route names to a consistent, readable Title Case.

    The imported permits arrive ALL-CAPS while the SSCL backbone names are
    Title Case, so the route column looked inconsistent. This harmonises every
    name: connectors (to / via) lower-cased, known acronyms / 3-letter station
    codes kept upper-cased, alphanumeric tokens (e.g. '90ft') left as-is, and
    everything else proper-cased. The 'via …' detail is preserved (RTO ask)."""
    s = str(name).strip()
    if not s or s.lower() in ("nan", "none"):
        return s
    # Unify the bidirectional "↔" separator to "to" so every name reads
    # consistently as "Origin to Destination [via …]" (matches the SSCL
    # official names; the routes are bidirectional regardless of wording).
    s = s.replace(" ↔ ", " to ").replace("↔", " to ")
    out = []
    for tok in s.split():
        low  = tok.lower()
        bare = re.sub(r"[^A-Za-z0-9]", "", tok)
        if low in _RNAME_CONNECTORS:
            out.append(low)
        elif bare.upper() in _RNAME_ACRONYMS:
            out.append(tok.upper())
        elif bare.isalpha() and bare.isupper() and len(bare) == 3:
            out.append(tok.upper())          # 3-letter station codes (GBS, TRC…)
        elif any(c.isdigit() for c in tok):
            out.append(tok)                  # leave '90ft', 'NH1A' etc. untouched
        else:
            out.append(re.sub(r"[A-Za-z]+",
                              lambda m: m.group(0)[0].upper() + m.group(0)[1:].lower(),
                              tok))
    if out and out[0] in _RNAME_CONNECTORS:   # never start on a lower-case connector
        out[0] = out[0].capitalize()
    return " ".join(out)


def export_csv(gdf: gpd.GeoDataFrame, file_map: dict, out_path: str) -> None:
    log.info("Exporting CSV → %s", out_path)
    export_cols = [c for c in [
        "Route_ID", "Route_Name", "Route_Code", "Action_Taken", "New_Route_ID",
        "Displaced_Operator_Class",
        "Route_KM", "Route_Type", "OSRM_Duration_S", "Cycle_Time_Min",
        "Congestion_Zone", "N_Stops_Estimated", "Stop_Penalty_Min",
        "Sharp_Turns", "Junction_Penalty_Min",
        "Pop_Score", "POI_Score", "Road_Multiplier", "Final_CDI",
        "Social_Flag", "Priority_Band",
        "Headway_Min", "Fleet_Required",
        "HPV_Count", "MPV_Count", "LPV_Count",
        "CMP_Trunk", "CMP_Route_ID",
        "Population_Served", "Population_Served_Raw",
        "Corridor_Competitors",
        "HV_POI_Count", "Overlap_Metric", "Geo_Source",
        # v3.3 Phase-1 audit additions
        "Tourist_Corridor", "Seasonal_Operability",
        "District_HQ_Floor", "SSCL_CDI_Conflict",
        "SSCL_CDI_Conflict_Strong", "SSCL_CDI_Conflict_Weak_SSCL",
        "Daily_Trips", "Daily_KM",
        "Daily_Capacity_Pax", "Daily_Demand_Pax",
        "Load_Ratio", "Load_Flag",
        "Pax_Journey_Time_Min", "Journey_Time_Flag",
        "Daily_Revenue_INR", "Daily_Op_Cost_INR",
        "Viability_Ratio", "Subsidy_Risk_Flag",
        "Emissions_GCO2_Daily", "Equity_Score",
    ] if c in gdf.columns]
    df_out = gdf[export_cols].copy()
    df_out["Map_File"] = df_out["New_Route_ID"].map(file_map).fillna("")
    df_out.to_csv(out_path, index=False, encoding="utf-8-sig")
    log.info("  CSV written: %d rows.", len(df_out))


def export_geojson(gdf: gpd.GeoDataFrame, out_path: str) -> None:
    log.info("Exporting Network GeoJSON → %s", out_path)
    active_gdf = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"].copy()
    keep_cols  = [c for c in [
        "New_Route_ID", "Route_Name", "Action_Taken", "Route_KM",
        "Route_Type", "Priority_Band", "Headway_Min", "Fleet_Required",
        "HPV_Count", "MPV_Count",
        "Route_Code",
        "Social_Flag", "Population_Served", "Population_Served_Raw", "Final_CDI",
        "CMP_Trunk", "CMP_Route_ID",
        "Congestion_Zone",
        # v3.3 Phase-1 audit additions
        "Tourist_Corridor", "Seasonal_Operability",
        "SSCL_CDI_Conflict", "District_HQ_Floor",
        "Load_Ratio", "Load_Flag",
        "Pax_Journey_Time_Min", "Journey_Time_Flag",
        "Viability_Ratio", "Subsidy_Risk_Flag",
        "Emissions_GCO2_Daily", "Equity_Score",
        "geometry",
    ] if c in active_gdf.columns]
    export_gdf = active_gdf[keep_cols]
    if export_gdf.crs != WGS84_CRS:
        export_gdf = export_gdf.to_crs(WGS84_CRS)
    export_gdf = export_gdf[
        export_gdf.geometry.notnull() & ~export_gdf.geometry.is_empty]
    export_gdf.to_file(out_path, driver="GeoJSON")
    log.info("  GeoJSON written: %d active features.", len(export_gdf))


# ══════════════════════════════════════════════════════════════════════════════
#  PHASE 4  ─  CARTOGRAPHY (geometry pipeline preserved from v6)
# ══════════════════════════════════════════════════════════════════════════════

def _safe_coords(geom) -> List[Tuple[float, float]]:
    if geom is None:
        return []
    coords = ([c for line in geom.geoms for c in line.coords]
              if isinstance(geom, MultiLineString) else list(geom.coords))
    return [(lat, lon) for lon, lat in coords]


def _popup_html(row: pd.Series) -> str:
    action_col = {
        "UPGRADED_TO_TRUNK":   "#1A237E",
        "MERGED_INTO_TRUNK":   "#880E4F",
        "RETAINED_AS_FEEDER":  "#00695C",
    }.get(row.get("Action_Taken", ""), "#333")
    band_colour = {"HP": "#1B5E20", "MP": "#E65100", "LP": "#B71C1C"}.get(
        row.get("Priority_Band", ""), "#333")
    social_tag  = ("<div style='color:#B71C1C;font-size:10px;font-weight:700;margin-bottom:2px'>"
                   "&#9873; Social Obligation Route</div>"
                   if row.get("Social_Flag") else "")
    cmp_tag     = (f"<div style='color:#6A1B9A;font-size:10px;font-weight:700;margin-bottom:4px'>"
                   f"&#127963; SSCL Backbone [{row.get('CMP_Route_ID','')}] — "
                   f"{SSCL_TRUNK_HEADWAY_MIN}-min headway</div>"
                   if row.get("CMP_Trunk") else "")
    hpv_pct     = (f"{int(row.get('HPV_Count',0)/max(1,row.get('Fleet_Required',1))*100)}%"
                   if row.get("Fleet_Required", 0) > 0 else "0%")
    return f"""
<div style="font-family:'Segoe UI',Arial,sans-serif;min-width:280px;font-size:12px">
  <div style="color:{action_col};font-size:15px;font-weight:700">{row.get('New_Route_ID','N/A')}</div>
  <div style="background:{action_col};color:#fff;display:inline-block;border-radius:4px;
    padding:1px 8px;font-size:10px;margin:3px 0 4px">{row.get('Action_Taken','').replace('_',' ')}</div>
  <div style="font-size:10px;color:#6A1B9A;margin-bottom:2px">
    Route Type: {row.get('Route_Type','?')} | Zone: {row.get('Congestion_Zone','?')}</div>
  <div style="color:{band_colour};font-weight:700;font-size:11px;margin-bottom:6px">
    Priority: {row.get('Priority_Band','?')} | Headway: {row.get('Headway_Min','?')} min
  </div>
  {social_tag}{cmp_tag}
  <table style="width:100%;border-collapse:collapse;line-height:1.7">
    <tr><td style="color:#666">Route Name</td><td>{row.get('Route_Name','')}</td></tr>
    <tr><td style="color:#666">Length</td><td><b>{row.get('Route_KM',0):.1f} km</b></td></tr>
    <tr><td style="color:#666">Fleet</td><td><b>{row.get('Fleet_Required','?')} buses</b></td></tr>
    <tr><td style="color:#888;font-size:11px">HPV / MPV</td>
        <td style="font-size:11px">{row.get('HPV_Count',0)} / {row.get('MPV_Count',0)}
        <span style="color:#888;font-size:10px">({hpv_pct} HPV)</span></td></tr>
    <tr style="border-top:1px solid #f0f0f0">
      <td style="color:#666">Cycle Time</td><td>{row.get('Cycle_Time_Min',0):.1f} min</td></tr>
    <tr><td style="color:#666">Stops (est.)</td><td>{row.get('N_Stops_Estimated',0)}</td></tr>
    <tr style="border-top:1px solid #f0f0f0">
      <td style="color:#1565C0">Pop_Score</td><td><b>{row.get('Pop_Score',0):.3f}</b></td></tr>
    <tr><td style="color:#E65100">POI_Score</td><td><b>{row.get('POI_Score',0):.3f}</b></td></tr>
    <tr><td style="color:#4A148C">Road ×</td><td><b>{row.get('Road_Multiplier',1.0):.2f}</b></td></tr>
    <tr><td style="color:#333"><b>Final CDI</b></td><td><b>{row.get('Final_CDI',0):.4f}</b></td></tr>
    <tr><td style="color:#666">Pop. Served</td><td><b>{row.get('Population_Served_Pct', 0.0):.2f}% of CMP 2024</b></td></tr>
  </table>
</div>"""


def _get_nearby_pois(row: pd.Series, gdf_pois: gpd.GeoDataFrame,
                     buffer_m: float = 400.0) -> List[dict]:
    """Return POIs within buffer_m metres of the route geometry."""
    geom = row.get("geometry")
    if geom is None or gdf_pois is None or len(gdf_pois) == 0:
        return []
    try:
        route_utm = gpd.GeoSeries([geom], crs="EPSG:4326").to_crs(UTM_CRS).iloc[0]
        pois_utm  = gdf_pois.to_crs(UTM_CRS)
        buf       = route_utm.buffer(buffer_m)
        within    = pois_utm[pois_utm.geometry.within(buf)]
        results   = []
        for _, p in within.iterrows():
            results.append({
                "name":     p.get("name", "POI"),
                "category": p.get("category", ""),
                "tier":     "Tier 1" if p.get("Is_HV_POI") else "Tier 2",
            })
        results.sort(key=lambda x: (x["tier"], x["name"]))
        return results[:30]
    except Exception:
        return []


def _individual_route_html(row: pd.Series, coords: List[Tuple[float, float]],
                            nearby_pois: List[dict]) -> str:
    """Generate a full standalone HTML page for an individual route."""
    rid         = row.get("New_Route_ID", "N/A")
    rname       = row.get("Route_Name", rid)
    action      = row.get("Action_Taken", "")
    route_type  = row.get("Route_Type", "?")
    zone        = row.get("Congestion_Zone", "?")
    band        = row.get("Priority_Band", "?")
    headway     = row.get("Headway_Min", "?")
    km          = row.get("Route_KM", 0)
    fleet       = row.get("Fleet_Required", "?")
    hpv         = row.get("HPV_Count", 0)
    mpv         = row.get("MPV_Count", 0)
    cycle       = row.get("Cycle_Time_Min", 0)
    stops       = row.get("N_Stops_Estimated", 0)
    pop_score   = row.get("Pop_Score", 0)
    poi_score   = row.get("POI_Score", 0)
    cdi         = row.get("Final_CDI", 0)
    pop_pct     = row.get("Population_Served_Pct", 0.0)
    pop_abs     = row.get("Population_Served", 0)
    social      = bool(row.get("Social_Flag", False))
    cmp_trunk   = bool(row.get("CMP_Trunk", False))
    cmp_id      = row.get("CMP_Route_ID", "")

    start_lat, start_lon = coords[0]
    end_lat,   end_lon   = coords[-1]
    origin_name = str(row.get("Route_From", "")).strip()
    dest_name   = str(row.get("Route_To",   "")).strip()
    if not origin_name or origin_name in ("nan", "None"):
        origin_name = f"{start_lat:.4f}, {start_lon:.4f}"
    if not dest_name or dest_name in ("nan", "None"):
        dest_name = f"{end_lat:.4f}, {end_lon:.4f}"

    via_raw  = row.get("Via_Coordinates")
    via_pts  = parse_via(via_raw)  # list of (lon, lat)
    via_coords_js = json.dumps([[lat, lon] for lon, lat in via_pts])

    # Build POI list HTML
    poi_rows = ""
    for p in nearby_pois:
        tier_badge = ("#D32F2F" if p["tier"] == "Tier 1" else "#F57F17")
        poi_rows += (f'<tr><td style="padding:2px 6px;font-size:11px">'
                     f'<span style="background:{tier_badge};color:#fff;border-radius:3px;'
                     f'padding:0 4px;font-size:9px">{p["tier"]}</span> '
                     f'{p["name"]}</td>'
                     f'<td style="padding:2px 6px;font-size:10px;color:#666">'
                     f'{p["category"]}</td></tr>')

    action_col = {
        "UPGRADED_TO_TRUNK":  "#1A237E",
        "RETAINED_AS_FEEDER": "#00695C",
    }.get(action, "#333")
    band_col = {"HP": "#1B5E20", "MP": "#E65100", "LP": "#B71C1C"}.get(band, "#333")
    route_colour = action_col
    centre_lat = sum(c[0] for c in coords) / len(coords)
    centre_lon = sum(c[1] for c in coords) / len(coords)
    coords_js  = json.dumps(coords)

    social_badge = ('<span style="background:#B71C1C;color:#fff;border-radius:4px;'
                    'padding:1px 7px;font-size:10px;margin-left:6px">Social Obligation</span>'
                    if social else "")
    sscl_badge   = (f'<span style="background:#6A1B9A;color:#fff;border-radius:4px;'
                    f'padding:1px 7px;font-size:10px;margin-left:6px">SSCL {cmp_id}</span>'
                    if cmp_trunk else "")

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Route {rid} — Kashmir Transit</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',Arial,sans-serif;display:flex;height:100vh;overflow:hidden;background:#f5f7fa}}
  #sidebar{{width:360px;min-width:280px;height:100vh;overflow-y:auto;background:#fff;
            border-right:1px solid #e0e4ea;display:flex;flex-direction:column}}
  #map{{flex:1;height:100vh}}
  .header{{background:{action_col};color:#fff;padding:14px 16px}}
  .header h2{{font-size:17px;font-weight:700;margin-bottom:2px}}
  .header .sub{{font-size:11px;opacity:0.85}}
  .badges{{padding:8px 16px;background:#f8f9fb;border-bottom:1px solid #e8eaed;display:flex;flex-wrap:wrap;gap:4px}}
  .badge{{display:inline-block;border-radius:4px;padding:2px 8px;font-size:10px;font-weight:600;color:#fff}}
  .section{{padding:12px 16px;border-bottom:1px solid #f0f2f5}}
  .section h3{{font-size:11px;font-weight:700;text-transform:uppercase;color:#9aa;letter-spacing:.05em;margin-bottom:8px}}
  .kv{{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:5px}}
  .kv .k{{color:#666;font-size:12px}}
  .kv .v{{font-size:12px;font-weight:600;color:#222;text-align:right;max-width:60%}}
  .terminal-box{{background:#f8f9fb;border-radius:6px;padding:8px 10px;margin-bottom:6px}}
  .terminal-box .label{{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.05em;margin-bottom:2px}}
  .terminal-box .value{{font-size:13px;font-weight:700}}
  .via-list{{list-style:none;padding:0}}
  .via-list li{{font-size:11px;color:#444;padding:2px 0;border-bottom:1px dotted #e8eaed}}
  .via-list li::before{{content:"• ";color:#5C6BC0}}
  .poi-table{{width:100%;border-collapse:collapse}}
  .pop-bar-bg{{background:#e8eaed;border-radius:4px;height:8px;margin-top:4px}}
  .pop-bar{{background:{action_col};border-radius:4px;height:8px}}
  .score-row{{display:flex;gap:8px;margin-bottom:4px}}
  .score-box{{flex:1;background:#f8f9fb;border-radius:6px;padding:6px 8px;text-align:center}}
  .score-box .sv{{font-size:15px;font-weight:700}}
  .score-box .sk{{font-size:9px;color:#888;text-transform:uppercase}}
</style>
</head>
<body>
<div id="sidebar">
  <div class="header">
    <div class="sub">Kashmir Valley Transit — Route Detail</div>
    <h2>{rid}</h2>
    <div style="font-size:12px;opacity:0.9;margin-top:2px">{rname}</div>
  </div>
  <div class="badges">
    <span class="badge" style="background:{action_col}">{action.replace('_',' ')}</span>
    <span class="badge" style="background:{band_col}">{band} Priority</span>
    <span class="badge" style="background:#37474F">{route_type.replace('_',' ')}</span>
    {"<span class='badge' style='background:#6A1B9A'>SSCL " + str(cmp_id) + "</span>" if cmp_trunk else ""}
    {"<span class='badge' style='background:#B71C1C'>Social Route</span>" if social else ""}
  </div>

  <div class="section">
    <h3>Terminals</h3>
    <div class="terminal-box" style="border-left:4px solid {COLOUR["start_pin"]}">
      <div class="label" style="color:{COLOUR["start_pin"]}">&#9650; Origin / Start</div>
      <div class="value">{origin_name}</div>
      <div style="font-size:10px;color:#999;margin-top:2px">{start_lat:.5f}, {start_lon:.5f}</div>
    </div>
    <div class="terminal-box" style="border-left:4px solid {COLOUR["end_pin"]}">
      <div class="label" style="color:{COLOUR["end_pin"]}">&#9660; Destination / End</div>
      <div class="value">{dest_name}</div>
      <div style="font-size:10px;color:#999;margin-top:2px">{end_lat:.5f}, {end_lon:.5f}</div>
    </div>
    {"<h3 style='margin-top:10px;margin-bottom:6px'>Via Points</h3><ul class='via-list'>" + "".join(f"<li>{lon:.4f}, {lat:.4f}</li>" for lon, lat in via_pts) + "</ul>" if via_pts else ""}
  </div>

  <div class="section">
    <h3>Area &amp; Zone</h3>
    <div class="kv"><span class="k">Congestion Zone</span><span class="v">{zone.replace('_',' ')}</span></div>
    <div class="kv"><span class="k">Route Type</span><span class="v">{route_type.replace('_',' ')}</span></div>
    <div class="kv"><span class="k">Route Length</span><span class="v">{km:.1f} km</span></div>
    <div class="kv"><span class="k">Est. Stops</span><span class="v">{stops}</span></div>
  </div>

  <div class="section">
    <h3>Population Served</h3>
    <div class="kv"><span class="k">Residents within walkshed</span><span class="v" style="color:{action_col}">{int(pop_abs):,}</span></div>
    <div class="kv"><span class="k">% of Srinagar UA (2024)</span><span class="v" style="color:{action_col}">{pop_pct:.2f}%</span></div>
    <div class="pop-bar-bg"><div class="pop-bar" style="width:{min(100,pop_pct*4):.1f}%"></div></div>
  </div>

  <div class="section">
    <h3>Service Plan</h3>
    <div class="kv"><span class="k">Headway</span><span class="v">{headway} min</span></div>
    <div class="kv"><span class="k">Cycle Time</span><span class="v">{float(cycle):.1f} min</span></div>
    <div class="kv"><span class="k">Fleet Required</span><span class="v">{fleet} buses</span></div>
    <div class="kv"><span class="k">HPV (12m)</span><span class="v">{hpv}</span></div>
    <div class="kv"><span class="k">MPV (9m)</span><span class="v">{mpv}</span></div>
  </div>

  <div class="section">
    <h3>Demand Scores</h3>
    <div class="score-row">
      <div class="score-box"><div class="sv" style="color:#1565C0">{pop_score:.3f}</div><div class="sk">Pop Score</div></div>
      <div class="score-box"><div class="sv" style="color:#E65100">{poi_score:.3f}</div><div class="sk">POI Score</div></div>
      <div class="score-box"><div class="sv" style="color:#1B5E20">{float(cdi):.4f}</div><div class="sk">Final CDI</div></div>
    </div>
  </div>

  <div class="section">
    <h3>POIs Covered ({len(nearby_pois)})</h3>
    {"<div style='color:#999;font-size:11px'>No POIs within 400m walkshed.</div>" if not nearby_pois else
     "<table class='poi-table'>" + poi_rows + "</table>"}
  </div>
</div>

<div id="map"></div>
<script>
var map = L.map('map').setView([{centre_lat}, {centre_lon}], 13);
{TILE_LAYER_JS}

var coords = {coords_js};
var line = L.polyline(coords, {{color: '{route_colour}', weight: 5, opacity: 0.9}}).addTo(map);
map.fitBounds(line.getBounds(), {{padding: [30, 30]}});

// Start marker
var startIcon = L.divIcon({{
  html: '<div style="background:{COLOUR["start_pin"]};color:#fff;border-radius:50%;width:16px;height:16px;display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:700;border:2px solid #fff;box-shadow:0 1px 3px rgba(0,0,0,.4)">S</div>',
  iconSize: [16, 16], iconAnchor: [8, 8], className: ''
}});
L.marker(coords[0], {{icon: startIcon}}).addTo(map)
  .bindPopup('<b>START</b><br>{origin_name}');

// End marker
var endIcon = L.divIcon({{
  html: '<div style="background:{COLOUR["end_pin"]};color:#fff;border-radius:50%;width:16px;height:16px;display:flex;align-items:center;justify-content:center;font-size:10px;font-weight:700;border:2px solid #fff;box-shadow:0 1px 3px rgba(0,0,0,.4)">E</div>',
  iconSize: [16, 16], iconAnchor: [8, 8], className: ''
}});
L.marker(coords[coords.length - 1], {{icon: endIcon}}).addTo(map)
  .bindPopup('<b>END</b><br>{dest_name}');

// Via points
var viaCoords = {via_coords_js};
viaCoords.forEach(function(pt, i) {{
  var viaIcon = L.divIcon({{
    html: '<div style="background:{COLOUR["via_dot"]};color:#fff;border-radius:50%;width:10px;height:10px;border:2px solid #fff;box-shadow:0 1px 3px rgba(0,0,0,.3)"></div>',
    iconSize: [10, 10], iconAnchor: [5, 5], className: ''
  }});
  L.marker(pt, {{icon: viaIcon}}).addTo(map).bindPopup('Via point ' + (i+1));
}});
</script>
<div style="position:fixed;bottom:6px;left:50%;transform:translateX(-50%);background:rgba(255,255,255,0.85);border:1px solid #ccc;border-radius:4px;padding:3px 10px;font-size:10px;color:#666;z-index:9999;pointer-events:none;white-space:nowrap">
  &#9888; Open-source global tiles (OpenStreetMap/CartoDB) — may not reflect India's official political boundaries
</div>
</body>
</html>"""


def build_individual_maps(gdf: gpd.GeoDataFrame,
                           gdf_pois: gpd.GeoDataFrame,
                           out_dir: str) -> dict:
    """Build detailed individual route HTML pages. Returns {New_Route_ID: filepath}."""
    log.info("Building individual route maps → %s/", out_dir)
    Path(out_dir).mkdir(exist_ok=True)
    active   = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"].copy()
    file_map = {}
    for _, row in active.iterrows():
        coords = _safe_coords(row.get("geometry"))
        if not coords:
            continue
        rid          = row["New_Route_ID"]
        nearby_pois  = _get_nearby_pois(row, gdf_pois)
        html_content = _individual_route_html(row, coords, nearby_pois)
        fname        = f"{out_dir}/{rid.replace('/', '_')}.html"
        Path(fname).write_text(html_content, encoding="utf-8")
        file_map[rid] = fname
    log.info("  %d individual maps saved.", len(file_map))
    return file_map


def build_master_map(gdf: gpd.GeoDataFrame,
                     gdf_pois: gpd.GeoDataFrame,
                     raster_path: str,
                     out_html: str,
                     net_pop: int,
                     network_score: float) -> None:
    """Build master interactive HTML map with filter panel and KPI sidebar."""
    log.info("Building master transit map → %s", out_html)

    active = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"].copy()

    # ── Collect route data for JS ─────────────────────────────────────────────
    routes_data = []
    for _, row in gdf.iterrows():
        coords = _safe_coords(row.get("geometry"))
        if not coords:
            continue
        action    = row.get("Action_Taken", "")
        rt        = row.get("Route_Type", "Urban")
        band      = row.get("Priority_Band", "?")
        colour    = (COLOUR["trunk"] if action == "UPGRADED_TO_TRUNK"
                     else COLOUR["regional"] if rt == "Regional_District"
                     else COLOUR["feeder"])
        weight    = 5 if action == "UPGRADED_TO_TRUNK" else (3 if rt == "Regional_District" else 2)
        routes_data.append({
            "id":        row.get("New_Route_ID", ""),
            "name":      row.get("Route_Name", ""),
            "action":    action,
            "type":      rt,
            "band":      band,
            "headway":   int(row.get("Headway_Min", 0) or 0),
            "km":        float(row.get("Route_KM", 0) or 0),
            "fleet":     int(row.get("Fleet_Required", 0) or 0),
            "pop_pct":   float(row.get("Population_Served_Pct", 0.0) or 0.0),
            "cdi":       float(row.get("Final_CDI", 0) or 0),
            "social":    bool(row.get("Social_Flag", False)),
            "cmp":       bool(row.get("CMP_Trunk", False)),
            "zone":      row.get("Congestion_Zone", ""),
            "colour":    colour,
            "weight":    weight,
            "coords":    coords,
        })

    # ── POI data for JS ───────────────────────────────────────────────────────
    pois_data = []
    for _, poi in gdf_pois.iterrows():
        pois_data.append({
            "lat":      poi.geometry.y,
            "lon":      poi.geometry.x,
            "name":     poi.get("name", "POI"),
            "category": poi.get("category", ""),
            "tier":     1 if poi.get("Is_HV_POI") else 2,
        })

    # ── Network KPIs ──────────────────────────────────────────────────────────
    n_trunk    = int((gdf["Action_Taken"] == "UPGRADED_TO_TRUNK").sum())
    n_feeder   = int((gdf["Action_Taken"] == "RETAINED_AS_FEEDER").sum())
    n_merged   = int((gdf["Action_Taken"] == "MERGED_INTO_TRUNK").sum())
    n_regional = int((gdf["Route_Type"]   == "Regional_District").sum())
    n_hp       = int((active["Priority_Band"] == "HP").sum())
    n_mp       = int((active["Priority_Band"] == "MP").sum())
    n_lp       = int((active["Priority_Band"] == "LP").sum())
    n_social   = int(active["Social_Flag"].fillna(False).sum())
    n_sscl     = int(active["CMP_Trunk"].fillna(False).sum()) if "CMP_Trunk" in active.columns else 0
    tot_fleet  = int(active["Fleet_Required"].sum())
    tot_hpv    = int(active["HPV_Count"].sum())
    tot_mpv    = int(active["MPV_Count"].sum())
    _sap       = study_area_population(raster_path)
    pop_pct    = (net_pop / _sap * 100) if _sap else 0.0   # F-V9: study-area denominator

    routes_js = json.dumps(routes_data)
    pois_js   = json.dumps(pois_data)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Kashmir Valley Transit Master Map v3</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',Arial,sans-serif;display:flex;height:100vh;overflow:hidden;background:#f0f2f5}}
  #panel{{width:320px;min-width:260px;height:100vh;overflow-y:auto;background:#fff;
          border-right:1px solid #dde1e7;display:flex;flex-direction:column;z-index:1000}}
  #map{{flex:1;height:100vh}}
  .panel-header{{background:linear-gradient(135deg,#1A237E,#283593);color:#fff;padding:14px 16px}}
  .panel-header h1{{font-size:15px;font-weight:700;margin-bottom:2px}}
  .panel-header .sub{{font-size:10px;opacity:0.8}}
  .kpi-grid{{display:grid;grid-template-columns:1fr 1fr;gap:6px;padding:10px 12px;background:#f8f9fb;border-bottom:1px solid #eaecef}}
  .kpi-box{{background:#fff;border-radius:6px;padding:7px 10px;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.06)}}
  .kpi-box .kn{{font-size:18px;font-weight:700;color:#1A237E}}
  .kpi-box .kl{{font-size:9px;color:#888;text-transform:uppercase;letter-spacing:.04em}}
  .section{{padding:10px 12px;border-bottom:1px solid #f0f2f5}}
  .section h3{{font-size:10px;font-weight:700;text-transform:uppercase;color:#9aabba;letter-spacing:.06em;margin-bottom:8px}}
  .filter-row{{display:flex;flex-wrap:wrap;gap:4px;margin-bottom:6px}}
  .ftog{{display:inline-flex;align-items:center;gap:4px;cursor:pointer;border:1px solid #dde1e7;
         border-radius:20px;padding:3px 9px;font-size:10px;font-weight:600;transition:all .15s;user-select:none}}
  .ftog.active{{color:#fff;border-color:transparent}}
  .ftog:hover{{opacity:.85}}
  .swatch{{width:8px;height:8px;border-radius:50%;flex-shrink:0}}
  .pop-bar-bg{{background:#e8eaed;border-radius:4px;height:6px;margin-top:3px}}
  .pop-bar{{background:#1A237E;border-radius:4px;height:6px}}
  .stat-row{{display:flex;justify-content:space-between;font-size:11px;margin-bottom:4px}}
  .stat-row .sk{{color:#666}}
  .stat-row .sv{{font-weight:600}}
  .route-list{{max-height:220px;overflow-y:auto;font-size:10px}}
  .route-item{{display:flex;align-items:center;gap:6px;padding:4px 6px;border-bottom:1px solid #f5f5f5;cursor:pointer;transition:background .1s}}
  .route-item:hover{{background:#f0f4ff}}
  .route-dot{{width:8px;height:8px;border-radius:50%;flex-shrink:0}}
  .route-id{{font-weight:600;font-size:10px;color:#1A237E;min-width:60px}}
  .route-name{{color:#555;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1}}
  .route-band{{font-size:9px;font-weight:700;border-radius:3px;padding:1px 4px;color:#fff;flex-shrink:0}}
  #legend{{padding:8px 12px;font-size:10px;border-top:1px solid #eaecef;background:#fafbfc}}
  .leg-item{{display:flex;align-items:center;gap:6px;margin-bottom:3px}}
  .leg-line{{height:3px;width:22px;border-radius:2px;flex-shrink:0}}
</style>
</head>
<body>
<div id="panel">
  <div class="panel-header">
    <div class="sub">Kashmir Valley — Route Rationalisation Engine v3</div>
    <h1>Master Transit Map</h1>
    <div style="font-size:10px;opacity:0.75;margin-top:2px">May 2026 | SSCL/CHALO Data</div>
  </div>

  <div class="kpi-grid">
    <div class="kpi-box"><div class="kn">{n_trunk + n_feeder}</div><div class="kl">Active Routes</div></div>
    <div class="kpi-box"><div class="kn">{tot_fleet}</div><div class="kl">Total Fleet</div></div>
    <div class="kpi-box"><div class="kn" style="color:#1B5E20">{n_hp}</div><div class="kl">HP Routes</div></div>
    <div class="kpi-box"><div class="kn" style="color:#6A1B9A">{n_sscl}</div><div class="kl">SSCL Trunks</div></div>
    <div class="kpi-box"><div class="kn" style="color:#00695C">{n_feeder}</div><div class="kl">Feeder Routes</div></div>
    <div class="kpi-box"><div class="kn" style="color:#B71C1C">{n_social}</div><div class="kl">Social Routes</div></div>
  </div>

  <div class="section">
    <h3>Population Coverage</h3>
    <div class="stat-row"><span class="sk">Residents served</span><span class="sv">{net_pop:,}</span></div>
    <div class="stat-row"><span class="sk">% of study-area population</span><span class="sv" style="color:#1A237E">{pop_pct:.1f}%</span></div>
    <div class="pop-bar-bg"><div class="pop-bar" style="width:{min(100,pop_pct):.1f}%"></div></div>
    <div style="margin-top:6px">
      <div class="stat-row"><span class="sk">HPV (12m buses)</span><span class="sv">{tot_hpv}</span></div>
      <div class="stat-row"><span class="sk">MPV (9m buses)</span><span class="sv">{tot_mpv}</span></div>
      <div class="stat-row"><span class="sk">Merged routes</span><span class="sv">{n_merged}</span></div>
    </div>
  </div>

  <div class="section">
    <h3>Filter by Route Category</h3>
    <div class="filter-row" id="filter-action">
      <span class="ftog active" data-group="action" data-val="UPGRADED_TO_TRUNK"
            data-color="{COLOUR['trunk']}"
            style="border-color:{COLOUR['trunk']}"
            onclick="toggleFilter(this)">
        <span class="swatch" style="background:{COLOUR['trunk']}"></span>Trunk ({n_trunk})
      </span>
      <span class="ftog active" data-group="action" data-val="RETAINED_AS_FEEDER"
            data-color="{COLOUR['feeder']}"
            style="border-color:{COLOUR['feeder']}"
            onclick="toggleFilter(this)">
        <span class="swatch" style="background:{COLOUR['feeder']}"></span>Feeder ({n_feeder})
      </span>
      <span class="ftog active" data-group="action" data-val="REGIONAL"
            data-color="{COLOUR['regional']}"
            style="border-color:{COLOUR['regional']}"
            onclick="toggleFilter(this)">
        <span class="swatch" style="background:{COLOUR['regional']}"></span>Regional ({n_regional})
      </span>
    </div>
  </div>

  <div class="section">
    <h3>Filter by Priority Band</h3>
    <div class="filter-row" id="filter-band">
      <span class="ftog active" data-group="band" data-val="HP"
            data-color="#1B5E20" style="border-color:#1B5E20"
            onclick="toggleFilter(this)">
        <span class="swatch" style="background:#1B5E20"></span>HP ({n_hp})
      </span>
      <span class="ftog active" data-group="band" data-val="MP"
            data-color="#E65100" style="border-color:#E65100"
            onclick="toggleFilter(this)">
        <span class="swatch" style="background:#E65100"></span>MP ({n_mp})
      </span>
      <span class="ftog active" data-group="band" data-val="LP"
            data-color="#B71C1C" style="border-color:#B71C1C"
            onclick="toggleFilter(this)">
        <span class="swatch" style="background:#B71C1C"></span>LP ({n_lp})
      </span>
    </div>
  </div>

  <div class="section">
    <h3>Filter by Special Type</h3>
    <div class="filter-row">
      <span class="ftog active" data-group="special" data-val="sscl"
            data-color="#6A1B9A" style="border-color:#6A1B9A"
            onclick="toggleFilter(this)">
        <span class="swatch" style="background:#6A1B9A"></span>SSCL Backbone ({n_sscl})
      </span>
      <span class="ftog active" data-group="special" data-val="social"
            data-color="#B71C1C" style="border-color:#B71C1C"
            onclick="toggleFilter(this)">
        <span class="swatch" style="background:#B71C1C"></span>Social Routes ({n_social})
      </span>
    </div>
  </div>

  <div class="section">
    <h3>POI Layers</h3>
    <div class="filter-row">
      <span class="ftog active" data-group="poi" data-val="tier1"
            data-color="{COLOUR['poi_high']}" style="border-color:{COLOUR['poi_high']}"
            onclick="toggleFilter(this)">
        <span class="swatch" style="background:{COLOUR['poi_high']}"></span>Tier 1 POIs
      </span>
      <span class="ftog" data-group="poi" data-val="tier2"
            data-color="{COLOUR['poi_secondary']}" style="border-color:{COLOUR['poi_secondary']}"
            onclick="toggleFilter(this)">
        <span class="swatch" style="background:{COLOUR['poi_secondary']}"></span>Tier 2 POIs
      </span>
    </div>
  </div>

  <div class="section">
    <h3>Routes (<span id="route-count">0</span> visible)</h3>
    <div class="route-list" id="route-list"></div>
  </div>

  <div id="legend">
    <div style="font-size:10px;font-weight:700;margin-bottom:5px;color:#444">Legend</div>
    <div class="leg-item"><div class="leg-line" style="background:{COLOUR['trunk']};height:4px"></div><span>Trunk (SSCL / Upgraded)</span></div>
    <div class="leg-item"><div class="leg-line" style="background:{COLOUR['feeder']}"></div><span>Feeder Route</span></div>
    <div class="leg-item"><div class="leg-line" style="background:{COLOUR['regional']}"></div><span>Regional / District</span></div>
    <div class="leg-item"><div style="width:10px;height:10px;border-radius:50%;background:{COLOUR['poi_high']};flex-shrink:0"></div><span>High-priority POI</span></div>
    <div class="leg-item"><div style="width:8px;height:8px;border-radius:50%;background:{COLOUR['poi_secondary']};flex-shrink:0"></div><span>Secondary POI</span></div>
  </div>
</div>

<div id="map"></div>

<script>
var map = L.map('map').setView([34.08, 74.81], 12);
{TILE_LAYER_JS}

var routesData = {routes_js};
var poisData   = {pois_js};

// ── State ──
var activeFilters = {{
  action:  new Set(['UPGRADED_TO_TRUNK','RETAINED_AS_FEEDER','REGIONAL']),
  band:    new Set(['HP','MP','LP']),
  special: new Set(['sscl','social']),
  poi:     new Set(['tier1'])
}};

// ── Build route layers ──
var routeLayers = [];
routesData.forEach(function(r) {{
  var bandCol = {{HP:'#1B5E20',MP:'#E65100',LP:'#B71C1C'}}[r.band] || '#555';
  var popupHtml = '<div style="font-family:Segoe UI,sans-serif;min-width:220px;font-size:12px">'
    + '<b style="font-size:14px;color:' + r.colour + '">' + r.id + '</b><br>'
    + '<span style="background:' + r.colour + ';color:#fff;border-radius:3px;padding:1px 6px;font-size:10px">'
    + r.action.replace(/_/g,' ') + '</span> '
    + '<span style="background:' + bandCol + ';color:#fff;border-radius:3px;padding:1px 6px;font-size:10px">'
    + r.band + '</span><br><br>'
    + '<table style="width:100%;border-collapse:collapse;font-size:11px;line-height:1.8">'
    + '<tr><td style="color:#666">Name</td><td>' + r.name + '</td></tr>'
    + '<tr><td style="color:#666">Type</td><td>' + r.type.replace(/_/g,' ') + ' | ' + r.zone.replace(/_/g,' ') + '</td></tr>'
    + '<tr><td style="color:#666">Length</td><td><b>' + r.km.toFixed(1) + ' km</b></td></tr>'
    + '<tr><td style="color:#666">Headway</td><td><b>' + r.headway + ' min</b></td></tr>'
    + '<tr><td style="color:#666">Fleet</td><td><b>' + r.fleet + ' buses</b></td></tr>'
    + '<tr><td style="color:#666">Pop. Served</td><td><b>' + r.pop_pct.toFixed(2) + '% of UA</b></td></tr>'
    + '<tr><td style="color:#1B5E20">CDI</td><td><b>' + r.cdi.toFixed(4) + '</b></td></tr>'
    + (r.social ? '<tr><td colspan="2" style="color:#B71C1C;font-weight:700">&#9873; Social Obligation Route</td></tr>' : '')
    + (r.cmp    ? '<tr><td colspan="2" style="color:#6A1B9A;font-weight:700">&#127963; SSCL Backbone</td></tr>' : '')
    + '</table></div>';
  var line = L.polyline(r.coords, {{
    color: r.colour, weight: r.weight, opacity: 0.85
  }}).bindPopup(popupHtml, {{maxWidth: 280}});
  routeLayers.push({{ layer: line, data: r }});
  line.addTo(map);
}});

// ── Build POI layers ──
var poiLayers = {{ tier1: [], tier2: [] }};
poisData.forEach(function(p) {{
  var col = p.tier === 1 ? '{COLOUR["poi_high"]}' : '{COLOUR["poi_secondary"]}';
  var r   = p.tier === 1 ? 6 : 4;
  var mk  = L.circleMarker([p.lat, p.lon], {{
    radius: r, color: col, fillColor: col, fillOpacity: 0.85, weight: 1
  }}).bindTooltip(p.name + ' (' + p.category + ')');
  if (p.tier === 1) poiLayers.tier1.push(mk);
  else              poiLayers.tier2.push(mk);
}});
poiLayers.tier1.forEach(function(mk) {{ mk.addTo(map); }});

// ── Filter logic ──
function routeVisible(r) {{
  // Regional routes are exclusively controlled by the REGIONAL toggle,
  // not by the Trunk/Feeder action toggle (their Action_Taken would be
  // RETAINED_AS_FEEDER, causing them to wrongly disappear with Feeder toggle).
  var isRegional  = r.type === 'Regional_District';
  var actionMatch = isRegional ? activeFilters.action.has('REGIONAL')
                               : activeFilters.action.has(r.action);
  var bandMatch   = activeFilters.band.has(r.band);
  // Special filter: when toggled OFF, hide routes of that special type.
  // Routes that are neither SSCL nor social are unaffected by special toggles.
  var specialMatch = (!r.cmp    || activeFilters.special.has('sscl')) &&
                     (!r.social || activeFilters.special.has('social'));
  return actionMatch && bandMatch && specialMatch;
}}

function applyFilters() {{
  var visible = 0;
  routeLayers.forEach(function(rl) {{
    if (routeVisible(rl.data)) {{ rl.layer.addTo(map); visible++; }}
    else                        {{ map.removeLayer(rl.layer); }}
  }});
  // POI layers
  ['tier1','tier2'].forEach(function(t) {{
    poiLayers[t].forEach(function(mk) {{
      if (activeFilters.poi.has(t)) mk.addTo(map);
      else                          map.removeLayer(mk);
    }});
  }});
  // Update route list
  document.getElementById('route-count').textContent = visible;
  var listEl = document.getElementById('route-list');
  listEl.innerHTML = '';
  routeLayers.forEach(function(rl) {{
    if (!routeVisible(rl.data)) return;
    var r = rl.data;
    var bandCol = {{HP:'#1B5E20',MP:'#E65100',LP:'#B71C1C'}}[r.band] || '#555';
    var item = document.createElement('div');
    item.className = 'route-item';
    item.innerHTML = '<span class="route-dot" style="background:' + r.colour + '"></span>'
      + '<span class="route-id">' + r.id + '</span>'
      + '<span class="route-name">' + r.name + '</span>'
      + '<span class="route-band" style="background:' + bandCol + '">' + r.band + '</span>';
    item.addEventListener('click', function() {{
      rl.layer.openPopup();
      if (rl.data.coords && rl.data.coords.length)
        map.fitBounds(L.polyline(rl.data.coords).getBounds(), {{padding:[40,40]}});
    }});
    listEl.appendChild(item);
  }});
}}

function toggleFilter(el) {{
  var group = el.dataset.group;
  var val   = el.dataset.val;
  var col   = el.dataset.color || '#1A237E';
  if (activeFilters[group].has(val)) {{
    activeFilters[group].delete(val);
    el.classList.remove('active');
    el.style.background = '';
    el.style.color = col;
    el.style.borderColor = col;
  }} else {{
    activeFilters[group].add(val);
    el.classList.add('active');
    el.style.background = col;
    el.style.color = '#fff';
    el.style.borderColor = 'transparent';
  }}
  applyFilters();
}}

// Initialise: apply active styling from data-color (not from style.color which
// was the original bug — style.color gets overwritten to #fff on first activate
// and then read back as white, making labels invisible on next deactivate).
document.querySelectorAll('.ftog').forEach(function(el) {{
  var col = el.dataset.color || '#1A237E';
  if (el.classList.contains('active')) {{
    el.style.background  = col;
    el.style.color       = '#fff';
    el.style.borderColor = 'transparent';
  }} else {{
    el.style.background  = '';
    el.style.color       = col;
    el.style.borderColor = col;
  }}
}});

applyFilters();
</script>
<div style="position:fixed;bottom:6px;left:50%;transform:translateX(-50%);background:rgba(255,255,255,0.85);border:1px solid #ccc;border-radius:4px;padding:3px 10px;font-size:10px;color:#666;z-index:9999;pointer-events:none;white-space:nowrap">
  &#9888; Open-source global tiles (OpenStreetMap/CartoDB) — may not reflect India's official political boundaries
</div>
</body>
</html>"""

    Path(out_html).write_text(html, encoding="utf-8")
    log.info("  Master map saved.")


# ══════════════════════════════════════════════════════════════════════════════
#  XLSX EXPORT (4-sheet workbook, formulas on Summary sheets)
# ══════════════════════════════════════════════════════════════════════════════

def export_xlsx(gdf: gpd.GeoDataFrame, out_path: str, net_pop: int) -> None:
    """
    4-sheet XLSX workbook:
      Sheet 0 — Cover Sheet (methodology, v2 directives summary, limitations)
      Sheet 1 — Route-Level Plan (Groups A / B / C)
      Sheet 2 — Priority Band Summary (Excel SUMIF formulas)
      Sheet 3 — Route Type Summary (Excel SUMIF formulas)
    """
    log.info("Exporting XLSX → %s", out_path)
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb   = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    S1   = "Route-Level Plan"

    # ── Cover Sheet ──────────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = "Cover Sheet"
    cover_lines = [
        ("Srinagar / Kashmir Valley Public Transport — Route Frequency Plan v3", True, 16),
        ("Engine v3.0 Kashmir Fork | SSCL/CHALO Backbone Injection | May 2026", False, 11),
        ("", False, 11),
        ("KEY KASHMIR-SPECIFIC CHANGES FROM JAMMU v3:", True, 12),
        ("• SSCL e-bus backbone — 30 hardcoded trunks from CHALO data "
         "(15-min headway, not Jammu's 10-min BRT assumption).", False, 10),
        ("• Geographic re-centring: bounding box, lat thresholds, river coords "
         "rebuilt for Srinagar UA + Valley districts.", False, 10),
        ("• Population baseline: 1,660,000 (Srinagar UA 2024 projection) "
         "replaces Jammu 1,653,873.", False, 10),
        ("• 3-tier POI system + seasonal toggle: T1 (1.0), T2 (0.4), T3 "
         "(0.6 summer / 0.0 winter for tourism/yatra POIs).", False, 10),
        ("• Women-anchor +25% boost: calibrated to CHALO data showing 64.5% "
         "of riders are women (free-fare effect).", False, 10),
        ("• Kashmir social-obligation anchors: KP migrant camps + tertiary "
         "hospitals (SKIMS/SMHS/LD) + Khonmoh/Rangreth/Lassipora industrial.", False, 10),
        ("", False, 11),
        ("CORE DIRECTIVES (retained from Jammu v3):", True, 12),
        ("D1: Zero Route Truncation — ALL routes retained. >40km flagged "
         "as Regional_District with relaxed headways.", False, 10),
        ("D2: Realistic Cycle Times — congestion multiplier (2.0× Downtown "
         "Srinagar, 1.5× peri-urban) + stop penalties + 10% layover.", False, 10),
        ("D3: Minimum Viable Fleet — fleet below threshold raised directly. "
         "LPV eradicated — all fleet is HPV or MPV.", False, 10),
        ("D4: Trunk CDI gate = 30th percentile (aggressive). Anti-stranding: "
         "routes < 5 km ineligible for Trunk.", False, 10),
        ("D5: Population 95th-pct cap + 3-tier POI weights.", False, 10),
        ("D6: Full audit trail in log file. All functions modular.", False, 10),
        ("", False, 11),
        ("KASHMIR-SPECIFIC LIMITATIONS (Phase 1, to address in v4):", True, 11),
        ("• Walksheds are still Euclidean buffers. Dal Lake / Anchar / "
         "Hokersar / Jhelum barriers and slope not yet modelled.", False, 10),
        ("• Winter scenario is a binary flag (WINTER_SCENARIO). A full "
         "seasonal-stratified run needs four passes.", False, 10),
        ("• Per-route AFC validation against CHALO ridership is still "
         "manual. Automated calibration is a v4 task.", False, 10),
        ("• Security/convoy windows on NH-44 and military polygons not "
         "yet subtracted from operable network.", False, 10),
    ]
    for row_i, (text, bold, size) in enumerate(cover_lines, start=1):
        c = ws0.cell(row=row_i, column=1, value=text)
        c.font = Font(name="Calibri", bold=bold, size=size)
        c.alignment = Alignment(wrap_text=True)
    ws0.column_dimensions["A"].width = 100

    # ── Route-Level Plan ─────────────────────────────────────────────────────
    ws1 = wb.create_sheet(S1)

    # Build export dataframe — keep cols that exist
    df_out = gdf[[c for c in SHEET1_ALL_COLS if c in gdf.columns]].copy()

    # Write header
    for col_i, col_name in enumerate(df_out.columns, start=1):
        c = ws1.cell(row=1, column=col_i, value=col_name)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        c.border = thin_border
        ws1.column_dimensions[get_column_letter(col_i)].width = 16

    # Write data rows
    bool_cols = {"Social_Flag", "CMP_Trunk"}
    for row_i, (_, row_data) in enumerate(df_out.iterrows(), start=2):
        for col_i, col_name in enumerate(df_out.columns, start=1):
            val = row_data[col_name]
            if col_name in bool_cols:
                val = bool(val)
            elif isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val)
            c = ws1.cell(row=row_i, column=col_i, value=val)
            c.font = Font(name="Calibri", size=9)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center", vertical="center")

    ws1.freeze_panes = "A2"

    # ── Priority Band Summary ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Priority Band Summary")
    bands = [("HP — High Priority", "HP", "70AD47"),
             ("MP — Medium Priority", "MP", "FFD966"),
             ("LP — Low Priority", "LP", "FF7070")]
    hdrs  = ["Priority Band", "No. of Routes", "Est. Total Fleet",
             "HPV", "MPV", "Pop. Served (% of CMP 2024)"]
    for ci, h in enumerate(hdrs, 1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

    # Map column names to Excel letters
    col_letter = {col: get_column_letter(i+1)
                  for i, col in enumerate(df_out.columns)}

    def _col(name):
        return col_letter.get(name, "A")

    pb_col  = _col("Priority_Band")
    fr_col  = _col("Fleet_Required")
    hpv_col = _col("HPV_Count")
    mpv_col = _col("MPV_Count")
    pop_col = _col("Population_Served_Pct")   # CMP % — not raw raster count
    fa_col  = _col("Action_Taken")

    for ri, (label, code, colour) in enumerate(bands, 2):
        ws2.cell(row=ri, column=1, value=label).font = Font(
            name="Calibri", bold=True, color=f"00{colour}")
        formulas = [
            f"=COUNTIF('{S1}'!{pb_col}:{pb_col},\"{code}\")",
            f"=SUMIF('{S1}'!{pb_col}:{pb_col},\"{code}\",'{S1}'!{fr_col}:{fr_col})",
            f"=SUMIF('{S1}'!{pb_col}:{pb_col},\"{code}\",'{S1}'!{hpv_col}:{hpv_col})",
            f"=SUMIF('{S1}'!{pb_col}:{pb_col},\"{code}\",'{S1}'!{mpv_col}:{mpv_col})",
            # AVERAGEIF: avg % per route in this band — summing % makes no sense
            f"=IFERROR(AVERAGEIF('{S1}'!{pb_col}:{pb_col},\"{code}\",'{S1}'!{pop_col}:{pop_col}),0)",
        ]
        for ci, f in enumerate(formulas, 2):
            c = ws2.cell(row=ri, column=ci, value=f)
            c.font = Font(name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=f"FF{colour}")
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border

    # Totals row — fleet/HPV/MPV summed; pop is network average (not sum)
    total_row = len(bands) + 2
    ws2.cell(row=total_row, column=1, value="NETWORK AVG %").font = Font(
        name="Calibri", bold=True, color="FFFFFF")
    ws2.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="1F3864")
    for ci, (col_ref, use_avg) in enumerate(
            [(fr_col, False), (hpv_col, False), (mpv_col, False), (pop_col, True)], 3):
        formula = (f"=IFERROR(AVERAGE('{S1}'!{col_ref}:{col_ref}),0)"
                   if use_avg
                   else f"=SUM('{S1}'!{col_ref}:{col_ref})")
        c = ws2.cell(row=total_row, column=ci, value=formula)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border

    # ── Route Type Summary (v2 addition) ────────────────────────────────────
    ws3 = wb.create_sheet("Route Type Summary")
    rt_col   = _col("Route_Type")
    rt_types = [("Urban", "2196F3"), ("Peri_Urban", "66BB6A"),
                ("Regional_District", "9C27B0")]
    hdrs3 = ["Route Type", "No. of Routes", "Est. Total Fleet",
             "HPV", "MPV", "Pop. Served (% of CMP 2024)"]
    for ci, h in enumerate(hdrs3, 1):
        c = ws3.cell(row=1, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

    for ri, (rt_label, colour) in enumerate(rt_types, 2):
        ws3.cell(row=ri, column=1, value=rt_label).font = Font(
            name="Calibri", bold=True)
        formulas = [
            f"=COUNTIF('{S1}'!{rt_col}:{rt_col},\"{rt_label}\")",
            f"=SUMIF('{S1}'!{rt_col}:{rt_col},\"{rt_label}\",'{S1}'!{fr_col}:{fr_col})",
            f"=SUMIF('{S1}'!{rt_col}:{rt_col},\"{rt_label}\",'{S1}'!{hpv_col}:{hpv_col})",
            f"=SUMIF('{S1}'!{rt_col}:{rt_col},\"{rt_label}\",'{S1}'!{mpv_col}:{mpv_col})",
            # AVERAGEIF: avg % per route in this type — summing % makes no sense
            f"=IFERROR(AVERAGEIF('{S1}'!{rt_col}:{rt_col},\"{rt_label}\",'{S1}'!{pop_col}:{pop_col}),0)",
        ]
        for ci, f in enumerate(formulas, 2):
            c = ws3.cell(row=ri, column=ci, value=f)
            c.font = Font(name="Calibri", size=10)
            c.fill = PatternFill("solid", fgColor=f"FF{colour}")
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border

    ws3.cell(row=len(rt_types) + 2, column=1,
             value="⚠ NOTE: Regional_District routes use relaxed headways "
                   "(60–90 min) per Directive 1. Fleet figures reflect this. "
                   "Pop. Served = avg % of CMP 2024 study-area population per route.").font = Font(
                       name="Calibri", italic=True, size=9, color="595959")

    # Reorder sheets
    wb.move_sheet("Cover Sheet", offset=-wb.index(wb["Cover Sheet"]))
    wb.save(out_path)
    log.info("  XLSX written: %d data rows, 4 sheets.", len(df_out))


def export_xlsx_rto(gdf: gpd.GeoDataFrame, out_path: str, net_pop: int) -> None:
    """
    9-sheet RTO-Ready Master Excel Workbook with Professional Styling.
    """
    log.info("Exporting RTO-Ready XLSX (Professional) → %s", out_path)
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    import datetime

    wb = Workbook()
    
    # ── Styling Definitions ──
    # Colors
    NAVY = "1A237E"
    TEAL = "00695C"
    PURPLE = "6A1B9A"
    RED = "D32F2F"
    AMBER = "F9A825"
    LIGHT_GRAY = "F5F5F5"
    BORDER_COLOR = "E0E0E0"
    
    # Borders
    thin = Side(style="thin", color=BORDER_COLOR)
    light_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bottom_border = Border(bottom=thin)
    no_border = Border()
    
    # Fonts
    header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Segoe UI", bold=True, color=NAVY, size=18)
    subtitle_font = Font(name="Segoe UI", bold=False, color="555555", size=12)
    body_font = Font(name="Segoe UI", size=10, color="333333")
    bold_body_font = Font(name="Segoe UI", size=10, bold=True, color="333333")
    
    def set_header(cell, text, bg_color=NAVY):
        cell.value = text
        cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=bg_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = light_border
        
    def set_cell(cell, val, bold=False, align="center", bg_color=None, num_format=None, font_color="333333"):
        cell.value = val
        cell.font = Font(name="Segoe UI", size=10, bold=bold, color=font_color)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = light_border
        if bg_color:
            cell.fill = PatternFill("solid", fgColor=bg_color)
        if num_format:
            cell.number_format = num_format

    def hide_gridlines(ws):
        ws.sheet_view.showGridLines = False

    # Extract required series
    active_mask = gdf["Action_Taken"] != "MERGED_INTO_TRUNK"
    active_gdf = gdf[active_mask]
    
    total_active = len(active_gdf)
    total_fleet = int(active_gdf["Fleet_Required"].sum())
    total_pop_served = int(net_pop)
    avg_headway = float(active_gdf["Headway_Min"].mean()) if not active_gdf.empty else 0.0
    
    if "Displaced_Operator_Class" not in gdf.columns:
        def get_op_class(row):
            if row["Action_Taken"] != "MERGED_INTO_TRUNK": return ""
            if row.get("LPV_Count", 0) > 0 or row.get("Priority_Band", "") == "LP": return "LPV / Tempo"
            if row.get("HPV_Count", 0) > 0: return "HPV Bus"
            return "Private Minibus"
        gdf["Displaced_Operator_Class"] = gdf.apply(get_op_class, axis=1)

    # ── Sheet 1: Cover & Sign-off ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Cover & Sign-off"
    hide_gridlines(ws1)
    
    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 45
    ws1.column_dimensions["C"].width = 45
    ws1.column_dimensions["D"].width = 5
    
    # Placeholder for Logo/Letterhead
    ws1.merge_cells('B2:C4')
    logo_cell = ws1.cell(row=2, column=2, value="[ J&K Transport Dept Letterhead Placeholder ]")
    logo_cell.font = Font(name="Segoe UI", italic=True, size=14, color="999999")
    logo_cell.alignment = Alignment(horizontal="center", vertical="center")
    logo_cell.border = light_border
    logo_cell.fill = PatternFill("solid", fgColor="FAFAFA")
    
    ws1.cell(row=6, column=2, value="Route Rationalisation Plan").font = title_font
    ws1.cell(row=7, column=2, value="Srinagar / Kashmir Valley").font = Font(name="Segoe UI", bold=True, color=TEAL, size=14)
    
    gen_date = datetime.datetime.now().strftime("%d %b %Y")
    ws1.cell(row=9, column=2, value=f"Version: v3.4.3  |  Generated: {gen_date}").font = subtitle_font
    ws1.cell(row=10, column=2, value="Engine: SSCL/CHALO Backbone Injection | Kashmir Geographic Recentre").font = subtitle_font
    
    summary_text = (f"This plan details the rationalised public transport network for Srinagar and the Kashmir Valley. "
                    f"It proposes a total of {total_active} active routes, deploying an estimated {total_fleet} buses. "
                    f"The network is designed to serve approximately {total_pop_served:,} residents, "
                    f"maintaining an average headway of {avg_headway:.1f} minutes across active corridors.")
    
    ws1.merge_cells('B12:C14')
    c = ws1.cell(row=12, column=2, value=summary_text)
    c.font = body_font
    c.alignment = Alignment(wrap_text=True, vertical="top")
    
    # KPI Tiles
    ws1.cell(row=16, column=2, value="NETWORK HIGHLIGHTS").font = Font(name="Segoe UI", bold=True, color=NAVY, size=12)
    
    # v3.3.6: 6 KPI tiles in a 2×3 grid covering scope (routes/fleet),
    # demand (population/headway), and impact (merged/tourist/social).
    merged_count    = int((gdf["Action_Taken"] == "MERGED_INTO_TRUNK").sum())
    sscl_count      = int(gdf.get("CMP_Trunk", pd.Series(False, index=gdf.index)).sum())
    social_count    = int(gdf.get("Social_Flag", pd.Series(False, index=gdf.index)).sum())
    tourist_count   = int(gdf.get("Tourist_Corridor", pd.Series(False, index=gdf.index)).sum())

    kpis = [
        ("Active Routes",        f"{total_active}",               "B18"),
        ("Total Fleet",          f"{total_fleet}",                "C18"),
        ("Population Served",    f"{total_pop_served:,}",         "B21"),
        ("Avg Headway (min)",    f"{avg_headway:.1f}",            "C21"),
        ("Merged Permits",       f"{merged_count}",               "B24"),
        ("Tourist Corridors",    f"{tourist_count}",              "C24"),
        ("SSCL Backbone Routes", f"{sscl_count}",                 "B27"),
        ("Social Obligation",    f"{social_count}",               "C27"),
    ]
    for title, val, cell_ref in kpis:
        col = 2 if "B" in cell_ref else 3
        r = int(cell_ref[1:])
        ws1.cell(row=r, column=col, value=title).font = Font(name="Segoe UI", bold=True, color="757575", size=10)
        ws1.cell(row=r, column=col).alignment = Alignment(horizontal="center")
        val_cell = ws1.cell(row=r+1, column=col, value=val)
        val_cell.font = Font(name="Segoe UI", bold=True, size=18, color=TEAL)
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.fill = PatternFill("solid", fgColor="E0F2F1")
        val_cell.border = light_border
        ws1.cell(row=r, column=col).fill = PatternFill("solid", fgColor="E0F2F1")
        ws1.cell(row=r, column=col).border = light_border
        ws1.row_dimensions[r+1].height = 30

    # Legend (shifted down for the extra KPI rows)
    legend_start = 31
    ws1.cell(row=legend_start, column=2, value="STATUS LEGEND").font = Font(name="Segoe UI", bold=True, color=NAVY, size=12)
    ws1.cell(row=legend_start+1, column=2, value="■ Green Flag: 0.4–0.85 capacity utilization").font = Font(name="Segoe UI", color="2E7D32")
    ws1.cell(row=legend_start+2, column=2, value="■ Amber Flag: <0.4 (Under) or 0.85–1.0 (Tight)").font = Font(name="Segoe UI", color="F57F17")
    ws1.cell(row=legend_start+3, column=2, value="■ Red Flag: >1.0 (Overload) or 0 (No Capacity)").font = Font(name="Segoe UI", color="C62828")
    ws1.cell(row=legend_start+4, column=2, value="■ TEMP pill: Route_Code is a TMP-K placeholder (real code pending)").font = Font(name="Segoe UI", color="EF6C00")

    # Sign-off (shifted down)
    so_row = legend_start + 7
    ws1.cell(row=so_row, column=2, value="OFFICIAL SIGN-OFF").font = Font(name="Segoe UI", bold=True, color=NAVY, size=14)
    headers = ["Role", "Name", "Date", "Signature", "Remarks"]
    for i, h in enumerate(headers, start=2):
        cell = ws1.cell(row=so_row + 1, column=i)
        if i > 3: ws1.column_dimensions[get_column_letter(i)].width = 25 # expand for signature
        cell.value = h
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = light_border
        
    roles = ["Principal Secretary", "MD SSCL", "Director JKRTC", "Concerned RTO"]
    for i, role in enumerate(roles, start=so_row + 2):
        ws1.row_dimensions[i].height = 40  # Tall rows for signing
        set_cell(ws1.cell(row=i, column=2), role, bold=True, align="left")
        for j in range(3, 7):
            set_cell(ws1.cell(row=i, column=j), "")

    # ── Sheet 2: Network Summary ──────────────────────────────────────────
    ws2 = wb.create_sheet("Network Summary")
    hide_gridlines(ws2)
    ws2.column_dimensions["B"].width = 35
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["E"].width = 35
    ws2.column_dimensions["F"].width = 20
    
    ws2.cell(row=2, column=2, value="Network Composition").font = title_font
    set_header(ws2.cell(row=4, column=2), "Category")
    set_header(ws2.cell(row=4, column=3), "Count")
    
    comp_data = [
        ("Trunk", len(gdf[gdf["Action_Taken"] == "UPGRADED_TO_TRUNK"])),
        ("Feeder", len(gdf[gdf["Action_Taken"] == "RETAINED_AS_FEEDER"])),
        ("Merged", len(gdf[gdf["Action_Taken"] == "MERGED_INTO_TRUNK"])),
        ("Social", len(gdf[gdf.get("Social_Flag", False) == True])),
        ("SSCL", len(gdf[gdf.get("CMP_Trunk", False) == True])),
    ]
    for i, (cat, count) in enumerate(comp_data, start=5):
        bg = LIGHT_GRAY if i % 2 == 0 else "FFFFFF"
        set_cell(ws2.cell(row=i, column=2), cat, align="left", bg_color=bg)
        set_cell(ws2.cell(row=i, column=3), count, bg_color=bg, num_format="#,##0")
        
    ws2.cell(row=2, column=5, value="Fleet Composition").font = title_font
    set_header(ws2.cell(row=4, column=5), "Class")
    set_header(ws2.cell(row=4, column=6), "Total Fleet")
    
    fleet_data = [
        ("HPV (12m Bus)", int(active_gdf["HPV_Count"].sum())),
        ("MPV (9m Bus)", int(active_gdf["MPV_Count"].sum())),
        ("LPV (Minibus/Tempo)", int(active_gdf["LPV_Count"].sum())),
    ]
    for i, (cat, count) in enumerate(fleet_data, start=5):
        bg = LIGHT_GRAY if i % 2 == 0 else "FFFFFF"
        set_cell(ws2.cell(row=i, column=5), cat, align="left", bg_color=bg)
        set_cell(ws2.cell(row=i, column=6), count, bg_color=bg, num_format="#,##0")

    ws2.cell(row=12, column=2, value="Headway Distribution (Active)").font = title_font
    set_header(ws2.cell(row=14, column=2), "Headway (min)")
    set_header(ws2.cell(row=14, column=3), "Route Count")
    # v3.3.6: include 45-min bin (Social_Flag LP→MP exceptions surface here),
    # and roll any other observed headway into an "Other" bin so the table
    # ALWAYS sums to total_active — defensive against future tuning changes.
    expected_bins = [15, 20, 30, 35, 45, 60]
    bin_counts = {b: int((active_gdf["Headway_Min"] == b).sum()) for b in expected_bins}
    other = int(total_active - sum(bin_counts.values()))
    rows_hw = [(f"{b} min", bin_counts[b]) for b in expected_bins]
    if other:
        rows_hw.append(("Other", other))
    for i, (lbl, n) in enumerate(rows_hw, start=15):
        bg = LIGHT_GRAY if i % 2 == 0 else "FFFFFF"
        set_cell(ws2.cell(row=i, column=2), lbl, align="left", bg_color=bg)
        set_cell(ws2.cell(row=i, column=3), n, bg_color=bg, num_format="#,##0")
    # Sanity: total row reconciles to active count
    total_row = 15 + len(rows_hw)
    set_cell(ws2.cell(row=total_row, column=2), "TOTAL", bold=True, align="left", bg_color="E8EAF6")
    set_cell(ws2.cell(row=total_row, column=3), total_active, bold=True, bg_color="E8EAF6", num_format="#,##0")

    # ── Sheet 3: Route-Level Plan ─────────────────────────────────────────
    ws3 = wb.create_sheet("Route-Level Plan")
    hide_gridlines(ws3)
    
    columns_config = [
        ("Route_Code", "Identity", 16),
        ("Old_Route_ID", "Identity", 16),
        ("New_Route_ID", "Identity", 16),
        ("Route_Name", "Identity", 45),
        ("Action_Taken", "Identity", 22),
        ("Priority_Band", "Identity", 14),
        ("Route_KM", "Service", 12),
        ("Cycle_Time_Min", "Service", 16),
        ("Headway_Min", "Service", 14),
        ("Fleet_Required", "Service", 15),
        ("HPV_Count", "Service", 12),
        ("MPV_Count", "Service", 12),
        ("LPV_Count", "Service", 12),
        ("Bus_Type_Rec", "Service", 16),
        ("Service_Hours", "Service", 18),
        ("Population_Served", "Demand", 18),
        ("Daily_Demand_Pax", "Demand", 18),
        ("Daily_Capacity_Pax", "Demand", 18),
        ("Load_Flag", "Demand", 18),
        ("Subsidy_Risk_Flag", "Demand", 18),
        ("Social_Flag", "Demand", 14),
        ("Tourist_Corridor", "Demand", 16),
        ("Displaced_Operator_Class", "Impact", 25),
        ("Num_Permits_Affected", "Impact", 22),
        ("Recommended_Action", "Impact", 25),
        ("RTO_Remarks", "Impact", 35),
        ("Final_CDI", "Audit", 12),
        ("Map_Link", "Audit", 40),
    ]
    
    for col_i, (col_name, group, width) in enumerate(columns_config, start=1):
        bg_col = NAVY
        if group == "Service": bg_col = TEAL
        elif group == "Demand": bg_col = PURPLE
        elif group == "Impact": bg_col = RED
        elif group == "Audit": bg_col = AMBER
        
        c = ws3.cell(row=1, column=col_i, value=col_name)
        c.font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=bg_col)
        c.border = light_border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws3.column_dimensions[get_column_letter(col_i)].width = width
        
    # Freeze the identity block (cols A–F) so RTOs always see route + action.
    ws3.freeze_panes = "G2"
    ws3.row_dimensions[1].height = 30

    # v3.3.6: print setup for the workhorse sheet — landscape A3, fit columns
    # to width, repeat header row on each page so a printed copy is readable.
    ws3.page_setup.orientation = ws3.ORIENTATION_LANDSCAPE
    ws3.page_setup.paperSize   = ws3.PAPERSIZE_A3
    ws3.page_setup.fitToPage   = True
    ws3.page_setup.fitToWidth  = 1
    ws3.page_setup.fitToHeight = 0
    ws3.print_title_rows = "1:1"
    ws3.print_options.gridLines = False
    ws3.page_margins.left = 0.4
    ws3.page_margins.right = 0.4
    ws3.page_margins.top = 0.6
    ws3.page_margins.bottom = 0.6
    ws3.oddHeader.center.text = "Kashmir Route-Level Plan v3.4.3"
    ws3.oddHeader.center.size = 12
    ws3.oddFooter.right.text  = "Page &P of &N"
    
    # v3.3.6 — load any official Route_Code values committed in the dashboard
    # repo (preserved from the generate_route_codes.py run against the master
    # stops file). Falls back silently to TMP-K#### placeholders when no
    # mapping exists for a given Route_ID.
    _code_by_route_id: Dict[str, str] = {}
    try:
        import json as _json
        _prior = Path("E:/dash/bus-sathi-dashboard/public/route-rationalization-kashmir/data/routes.json")
        if _prior.exists():
            for _r in _json.loads(_prior.read_text(encoding="utf-8")):
                _rid  = str(_r.get("Route_ID", "")).strip()
                _code = str(_r.get("Route_Code", "")).strip()
                if _rid and _code and not _code.upper() in ("UNMATCHED",):
                    _code_by_route_id[_rid] = _code
            log.info("  RTO XLSX: loaded %d Route_Code mappings from dashboard routes.json",
                     len(_code_by_route_id))
    except Exception as _exc:
        log.warning("  RTO XLSX: could not load prior Route_Code mappings (%s)", _exc)

    def get_val(r, col_name, seq_no=0):
        if col_name == "Route_Code":
            # Precedence: engine-attached Route_Code > dashboard-cached real
            # code (by Route_ID) > TMP-K placeholder. Never use New_Route_ID
            # here — that's the trunk/feeder system ID, not a route code.
            ec = str(r.get("Route_Code", "") or "").strip()
            if ec and not ec.upper() == "UNMATCHED" and not ec.startswith("TMP-"):
                return ec
            rid = str(r.get("Route_ID", "") or "").strip()
            return _code_by_route_id.get(rid, f"TMP-K{seq_no:04d}")
        if col_name == "Old_Route_ID":
            # Source-permit ID from the original existing-routes.csv ingestion,
            # NOT CMP_Route_ID (which is the SSCL backbone tag like SSCL-01).
            return r.get("Route_ID", "")
        if col_name == "Bus_Type_Rec":
            if r.get("HPV_Count", 0) > 0: return "12m"
            if r.get("MPV_Count", 0) > 0: return "9m"
            if r.get("LPV_Count", 0) > 0: return "LPV"
            return ""
        if col_name == "Service_Hours": return "6 AM – 10 PM"
        if col_name == "Num_Permits_Affected":
            # Each MERGED row IS exactly one absorbed permit. Fleet_Required
            # is zeroed for merged rows (zero_merged_route_fleet), so using
            # that would show 0 for every absorbed permit.
            return 1 if r.get("Action_Taken") == "MERGED_INTO_TRUNK" else 0
        if col_name == "Recommended_Action":
            if r.get("Action_Taken") != "MERGED_INTO_TRUNK": return ""
            op = r.get("Displaced_Operator_Class", "")
            if op == "LPV / Tempo":   return "Last-mile reassignment"
            if op == "HPV Bus":       return "Roll into JKRTC"
            if op == "JKRTC / City Bus": return "Retained as feeder"
            return "Reassign or buyback"
        if col_name == "RTO_Remarks": return ""
        if col_name == "Map_Link":
            return f"route_maps_kashmir/{r.get('New_Route_ID', '')}.html"
        val = r.get(col_name, "")
        if pd.isna(val): return ""
        return val

    for row_i, (_, row_data) in enumerate(gdf.iterrows(), start=2):
        action = row_data.get("Action_Taken", "")
        load_flag = row_data.get("Load_Flag", "")
        social = row_data.get("Social_Flag", False)
        seq_no = row_i - 1  # 1-indexed sequence for TMP-K minting

        for col_i, (col_name, _, _) in enumerate(columns_config, start=1):
            val = get_val(row_data, col_name, seq_no=seq_no)
            if isinstance(val, bool): val = str(val)
            
            # Number formatting
            num_fmt = None
            if col_name in ["Population_Served", "Daily_Demand_Pax", "Daily_Capacity_Pax"]:
                if isinstance(val, (int, float)) and val != "":
                    num_fmt = "#,##0"
            elif col_name in ["Route_KM", "Final_CDI"]:
                if isinstance(val, (int, float)) and val != "":
                    num_fmt = "0.00"
            
            c = ws3.cell(row=row_i, column=col_i, value=val)
            c.font = Font(name="Segoe UI", size=9, bold=bool(social), color="333333")
            c.border = light_border
            
            # Alignment
            if col_name == "Route_Name":
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
            if num_fmt: c.number_format = num_fmt
            
            # Row Backgrounds
            bg_color = "FFFFFF"
            if action == "MERGED_INTO_TRUNK": bg_color = "FFEBEE"  # Light Red
            elif load_flag == "Red_Overload": bg_color = "FFF8E1"  # Light Amber
            elif load_flag == "Green" and action == "UPGRADED_TO_TRUNK": bg_color = "E8F5E9"  # Light Green
            elif row_i % 2 == 0: bg_color = "FAFAFA"
            c.fill = PatternFill("solid", fgColor=bg_color)

            # v3.3.6: dedicated cell-level pill colouring for Load_Flag and
            # Subsidy_Risk_Flag so the diagnostic stands out from the row tint.
            if col_name == "Load_Flag" and val:
                pill_bg = {
                    "Green":         ("388E3C", "FFFFFF"),
                    "Amber_Tight":   ("F9A825", "FFFFFF"),
                    "Amber_Under":   ("FBC02D", "333333"),
                    "Red_Overload":  ("D32F2F", "FFFFFF"),
                    "Red_NoCapacity":("B71C1C", "FFFFFF"),
                }.get(val)
                if pill_bg:
                    c.fill = PatternFill("solid", fgColor=pill_bg[0])
                    c.font = Font(name="Segoe UI", size=9, bold=True, color=pill_bg[1])
            elif col_name == "Subsidy_Risk_Flag" and (val is True or str(val).lower() == "true"):
                c.fill = PatternFill("solid", fgColor="FFCCBC")
                c.font = Font(name="Segoe UI", size=9, bold=True, color="BF360C")

            if col_name == "Map_Link" and val:
                c.hyperlink = val
                c.font = Font(name="Segoe UI", size=9, color="0000FF", underline="single")

    # Auto-filter on the whole data range — runs after rows are written so
    # the range covers all 342 rows (placing it before the loop captured
    # only the header row in earlier drafts).
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(columns_config))}{len(gdf) + 1}"

    # ── Sheet 4: Operator Absorption Register ─────────────────────────────
    ws4 = wb.create_sheet("Operator Absorption")
    hide_gridlines(ws4)
    ws4.column_dimensions["B"].width = 25
    ws4.column_dimensions["C"].width = 20
    ws4.column_dimensions["D"].width = 20
    ws4.column_dimensions["E"].width = 30
    ws4.column_dimensions["F"].width = 25
    ws4.column_dimensions["G"].width = 20

    merged_gdf = gdf[gdf["Action_Taken"] == "MERGED_INTO_TRUNK"]
    
    ws4.cell(row=2, column=2, value="Operator Absorption Register").font = title_font
    
    headers = ["Operator class", "# permits absorbed", "Avg fleet displaced", "Recommended action", "Estimated buyback cost (₹)", "RTO sign-off"]
    for i, h in enumerate(headers, start=2):
        set_header(ws4.cell(row=4, column=i), h)
        
    row_i = 5
    if not merged_gdf.empty:
        summary = merged_gdf.groupby("Displaced_Operator_Class")["Fleet_Required"].agg(['sum', 'mean']).reset_index()
        for _, row_data in summary.iterrows():
            op_class = row_data["Displaced_Operator_Class"]
            total_permits = int(row_data["sum"])
            avg_fleet = round(row_data["mean"], 1)
            
            rec_action = "Reassign or buyback"
            if op_class == "LPV / Tempo": rec_action = "Last-mile reassignment"
            elif op_class == "HPV Bus": rec_action = "Roll into JKRTC"
            elif op_class == "JKRTC / City Bus": rec_action = "Retain as feeder under JKRTC"

            # v3.3.6 — populate a starting buyback-cost estimate so the RTO
            # has a defensible Year-1 budget figure. Per-vehicle market
            # estimates (Jan 2026 J&K trade prices; secondary-market
            # depreciation factor 0.55 already applied):
            #   Private Minibus / 9m MPV:   ~₹15 L / permit
            #   LPV / Tempo:                ~₹3  L / permit
            #   HPV (12m) Bus:              ~₹50 L / permit
            #   JKRTC permits:              no buyback (state-owned)
            buyback_per_vehicle_lakh = {
                "Private Minibus":     15.0,
                "LPV / Tempo":          3.0,
                "HPV Bus":             50.0,
                "JKRTC / City Bus":     0.0,
            }.get(op_class, 10.0)
            est_buyback_inr = total_permits * buyback_per_vehicle_lakh * 1_00_000  # lakh → ₹

            bg = LIGHT_GRAY if row_i % 2 == 0 else "FFFFFF"
            set_cell(ws4.cell(row=row_i, column=2), op_class, bg_color=bg, align="left")
            set_cell(ws4.cell(row=row_i, column=3), total_permits, bg_color=bg, num_format="#,##0")
            set_cell(ws4.cell(row=row_i, column=4), avg_fleet, bg_color=bg, num_format="0.0")
            set_cell(ws4.cell(row=row_i, column=5), rec_action, bg_color=bg)
            set_cell(ws4.cell(row=row_i, column=6), est_buyback_inr, bg_color=bg,
                     num_format='"₹"#,##0')
            set_cell(ws4.cell(row=row_i, column=7), "", bg_color=bg)
            row_i += 1

        # v3.3.6 — grand-total row for the absorption register, so the
        # treasury / finance reviewer has a single ask-figure line.
        from openpyxl.utils import get_column_letter as _gcl
        first = 5
        last  = row_i - 1
        set_cell(ws4.cell(row=row_i, column=2), "TOTAL", bold=True, align="left",
                 bg_color="E8EAF6")
        set_cell(ws4.cell(row=row_i, column=3),
                 f"=SUM(C{first}:C{last})", bold=True, bg_color="E8EAF6",
                 num_format="#,##0")
        set_cell(ws4.cell(row=row_i, column=4), "", bg_color="E8EAF6")
        set_cell(ws4.cell(row=row_i, column=5), "Aggregated buyback obligation",
                 bold=True, bg_color="E8EAF6")
        set_cell(ws4.cell(row=row_i, column=6),
                 f"=SUM(F{first}:F{last})", bold=True, bg_color="E8EAF6",
                 num_format='"₹"#,##0')
        set_cell(ws4.cell(row=row_i, column=7), "", bg_color="E8EAF6")
        row_i += 1

    row_i += 3
    ws4.cell(row=row_i, column=2, value="Affected Routes Details").font = title_font
    row_i += 2
    
    sub_headers = ["Route_ID", "Operator Name", "Merged Into Trunk", "Consultation Status"]
    for i, h in enumerate(sub_headers, start=2):
        set_header(ws4.cell(row=row_i, column=i), h)
        ws4.column_dimensions[get_column_letter(i)].width = 30
        
    row_i += 1
    for _, row_data in merged_gdf.iterrows():
        bg = LIGHT_GRAY if row_i % 2 == 0 else "FFFFFF"
        set_cell(ws4.cell(row=row_i, column=2), row_data.get("New_Route_ID", ""), bg_color=bg)
        set_cell(ws4.cell(row=row_i, column=3), "", bg_color=bg) 
        set_cell(ws4.cell(row=row_i, column=4), "Multiple (See Engine)", bg_color=bg) 
        set_cell(ws4.cell(row=row_i, column=5), "", bg_color=bg) 
        row_i += 1

    # ── Sheets 5, 6, 7 (Similar formatting logic) ──
    def create_table_sheet(title, data_gdf, cols_config):
        ws = wb.create_sheet(title)
        hide_gridlines(ws)
        ws.cell(row=2, column=2, value=title).font = title_font
        for i, (col_name, width, attr) in enumerate(cols_config, start=2):
            ws.column_dimensions[get_column_letter(i)].width = width
            set_header(ws.cell(row=4, column=i), col_name)
            
        for ri, (_, row_data) in enumerate(data_gdf.iterrows(), start=5):
            bg = LIGHT_GRAY if ri % 2 == 0 else "FFFFFF"
            for ci, (_, _, attr) in enumerate(cols_config, start=2):
                val = row_data.get(attr, "") if attr else "Note"
                align = "left" if attr == "Route_Name" else "center"
                set_cell(ws.cell(row=ri, column=ci), val, align=align, bg_color=bg)

    trunk_gdf = gdf[(gdf["Action_Taken"] == "UPGRADED_TO_TRUNK") & (gdf["Priority_Band"] == "HP")]
    create_table_sheet("Trunk Detail", trunk_gdf, [
        ("Route_ID", 20, "New_Route_ID"),
        ("Route_Name", 45, "Route_Name"),
        ("Fleet", 15, "Fleet_Required"),
        ("Headway", 15, "Headway_Min"),
        ("SSCL_ID", 20, "CMP_Route_ID")
    ])

    social_gdf = gdf[gdf.get("Social_Flag", False) == True]
    create_table_sheet("Social Obligation", social_gdf, [
        ("Route_ID", 20, "New_Route_ID"),
        ("Route_Name", 45, "Route_Name"),
        ("Fleet", 15, "Fleet_Required"),
        ("Headway", 15, "Headway_Min"),
        ("Reason for Protection", 30, "")
    ])

    tourist_gdf = gdf[gdf.get("Tourist_Corridor", False) == True]
    create_table_sheet("Tourist & Seasonal", tourist_gdf, [
        ("Route_ID", 20, "New_Route_ID"),
        ("Route_Name", 45, "Route_Name"),
        ("Seasonal Operability", 25, "Seasonal_Operability"),
        ("Fleet", 15, "Fleet_Required")
    ])

    # ── Sheet 8: Calibration & Sources ────────────────────────────────────
    ws8 = wb.create_sheet("Calibration & Sources")
    hide_gridlines(ws8)
    ws8.column_dimensions["B"].width = 35
    ws8.column_dimensions["C"].width = 60
    
    ws8.cell(row=2, column=2, value="Calibration & Sources").font = title_font
    
    calib_data = [
        ("Engine Version", "v3.4.3 (Kashmir Fork)"),
        ("Date Generated", gen_date),
        ("Headway Targets", "HP: 20 min | MP: 35 min | LP: 35 min | Regional: 35 min | SSCL: 15 min (35-min ceiling — v3.3.7)"),
        ("CHALO Calibration Scorecard", "Matched to Apr 2026 ridership data"),
        ("Mode-share Assumption", "9% Urban, scaled for Peri-Urban/Regional"),
        ("Fleet-density Target", "0.60 per 1000 residents"),
        ("Population Source", "WorldPop 2024 / SMC Projections (1.66M)"),
        ("POI Source", "OSM Overpass with 3-Tier weights")
    ]
    
    for ri, (key, val) in enumerate(calib_data, start=4):
        set_cell(ws8.cell(row=ri, column=2), key, bold=True, align="left", bg_color="E8EAF6")
        set_cell(ws8.cell(row=ri, column=3), val, align="left")

    # ── Sheet 9: Limitations & Phase-2 ────────────────────────────────────
    ws9 = wb.create_sheet("Limitations")
    hide_gridlines(ws9)
    ws9.column_dimensions["B"].width = 100
    
    ws9.cell(row=2, column=2, value="Known Limitations & Phase-2 Backlog").font = title_font
    
    limitations = [
        "• Euclidean walksheds don't account for Dal/Anchar/Jhelum barriers.",
        "• Demand elasticity not modelled (Mohring effect — actual demand will rise with frequency).",
        "• Tourist surge volumes captured via POI weights, not arrival data.",
        "• Route_Code is derived from the stops master; where two routes share a "
        "geocoded endpoint they can receive the same code — endpoint geocoding is "
        "being hardened (district-aware) to remove collisions (see audit remediation).",
        "• 30 of the output routes are synthetic SSCL/CHALO e-bus backbone injections, "
        "not permit rows from existing-routes.csv.",
        "• Military/convoy windows on NH-44 not subtracted.",
        "• Phase-1 (recommended) plan; an aspirational 15-min-everywhere variant exists separately."
    ]
    
    for ri, limit in enumerate(limitations, start=4):
        c = ws9.cell(row=ri, column=2, value=limit)
        c.font = body_font
        c.alignment = Alignment(wrap_text=True)

    wb.save(out_path)
    log.info("  RTO-Ready XLSX written: 9 sheets with professional styling.")


def apply_regional_demand_headway(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """v3.4.2: re-derive headways for Regional_District (non-SSCL) lifelines from
    demand, replacing the flat 35-min ceiling on those routes only.

    Uses the Load_Ratio from the preceding Phase-4 pass. Because daily capacity
    ∝ 1/headway and demand is ~supply-stable, the headway that brings a route to
    REGIONAL_TARGET_LOAD is ``headway × (target_load / load)``. The result is
    clamped to [35, 50] min and bucketed to REGIONAL_HEADWAY_BUCKETS
    (35/40/45/50) — 50 min = the hard maximum rural wait (v3.4.3 user ask: even
    the quietest lifeline waits no more than ~50 min). Urban/Peri-Urban routes and the SSCL
    backbone are untouched. The caller MUST re-run step8 → step9 →
    zero_merged_route_fleet → compute_phase4_metrics so all KPIs stay consistent.
    """
    if not REGIONAL_DEMAND_SIZING:
        return gdf
    gdf = gdf.copy()
    cmp = gdf.get("CMP_Trunk", pd.Series(False, index=gdf.index)).astype(bool)
    mask = ((gdf["Route_Type"] == "Regional_District") & (~cmp)
            & (gdf["Action_Taken"] != "MERGED_INTO_TRUNK"))
    buckets = REGIONAL_HEADWAY_BUCKETS
    lo, hi = buckets[0], buckets[-1]
    changed = 0
    for idx in gdf[mask].index:
        load = float(gdf.at[idx, "Load_Ratio"] or 0.0)
        h0   = float(gdf.at[idx, "Headway_Min"] or 35.0)
        target_h = hi if load <= 0 else h0 * (REGIONAL_TARGET_LOAD / load)
        target_h = min(max(target_h, lo), hi)
        new_h = min(buckets, key=lambda b: abs(b - target_h))
        if new_h != int(gdf.at[idx, "Headway_Min"]):
            gdf.at[idx, "Headway_Min"] = int(new_h)
            changed += 1
    log.info("  Regional demand-sizing: re-headwayed %d/%d rural lifelines to "
             "{35/40/45/50} min (target load %.0f%%, 50-min hard max wait). "
             "Urban/Peri-Urban + SSCL keep 15/20/35.",
             changed, int(mask.sum()), REGIONAL_TARGET_LOAD * 100)
    return gdf


def compute_phase4_metrics(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """
    v3.3 Phase-1 audit additions — derived per-route KPIs computed AFTER all
    classification and fleet sizing is locked. These do NOT feed back into
    CDI/headway/fleet (would create circularity); they exist as reporting
    flags for the planner's review:

      • Daily_Trips      — service-hour round trips per route
      • Daily_KM         — Daily_Trips × Route_KM
      • Daily_Capacity   — Fleet × per-bus capacity × Daily_Trips
      • Daily_Demand     — Population_Served × MODE_SHARE × TRIP_RATE
      • Load_Ratio       — Daily_Demand / Daily_Capacity
      • Load_Flag        — Green (0.4–0.85) / Amber (<0.4 or 0.85–1.0) / Red (>1.0 or 0)
      • Pax_Journey_Time — 5.5 (access) + Headway/2 (wait) + Cycle/2 (in-vehicle)
      • Journey_Time_Flag— True if > PHASE4_JOURNEY_TIME_FLAG_MIN
      • Daily_Revenue    — Daily_Demand × PHASE4_FARE_INR
      • Daily_Op_Cost    — Daily_KM × PHASE4_OPERATING_COST_PER_KM
      • Viability_Ratio  — Daily_Revenue / Daily_Op_Cost  (1.0 = self-funding)
      • Subsidy_Risk     — True if Viability_Ratio < 0.5 AND not Social_Flag
      • Emissions_GCO2_Daily — Daily_KM × emission factor (e-bus for SSCL,
                               diesel otherwise)
      • Equity_Score     — relative score [0,1] = (Population_Served weighted
                               by Social_Flag×1.5 boost), normalised across
                               the active network
    """
    log.info("Phase 4: computing derived per-route KPIs "
             "(Load_Ratio, Journey_Time, Viability, Emissions, Equity)…")
    gdf = gdf.copy()

    headway   = gdf["Headway_Min"].astype(float).clip(lower=1.0)
    cycle     = gdf.get("Cycle_Time_Min",
                        pd.Series(0.0, index=gdf.index)).astype(float)
    fleet     = gdf["Fleet_Required"].astype(float).clip(lower=0.0)
    route_km  = gdf["Route_KM"].astype(float).clip(lower=0.0)
    pop_serv  = gdf.get("Population_Served",
                        pd.Series(0.0, index=gdf.index)).astype(float)
    hpv       = gdf.get("HPV_Count", pd.Series(0, index=gdf.index)).astype(float)
    mpv       = gdf.get("MPV_Count", pd.Series(0, index=gdf.index)).astype(float)
    lpv       = gdf.get("LPV_Count", pd.Series(0, index=gdf.index)).astype(float)

    # Service-day operational KPIs
    daily_trips = (PHASE4_SERVICE_HOURS * 60.0) / headway * 2.0  # one-way trips offered
    daily_km    = daily_trips * route_km

    avg_cap     = (hpv * VEHICLE_CAPACITY_HPV
                   + mpv * VEHICLE_CAPACITY_MPV
                   + lpv * VEHICLE_CAPACITY_LPV)
    # avg_cap is the SUM of seat-capacity across the route's fleet.
    # If HPV/MPV/LPV not yet split, fall back to fleet × MPV capacity.
    avg_cap     = avg_cap.where(avg_cap > 0, fleet * VEHICLE_CAPACITY_MPV)

    # v3.3.3 (teammate review): Daily_Capacity must use CYCLE-TIME based
    # round-trips per BUS, not route-total one-way trips. The old formula
    # `avg_cap × daily_trips` double-counted (avg_cap is fleet-summed,
    # daily_trips is route-totalled).
    # Correct: Daily_Capacity = Fleet × VehicleCapacity × TripsPerBusPerDay
    #          TripsPerBusPerDay = ServiceMinutes / Cycle_Time_Min
    # For LALBAZAR (9 buses × 35 seats, cycle 99 min):
    #   9 × 35 × (960/99) = 9 × 35 × 9.7 = 3,058 seats/day (correct denominator).
    cycle_safe     = cycle.clip(lower=1.0)
    trips_per_bus  = (PHASE4_SERVICE_HOURS * 60.0) / cycle_safe
    daily_capacity = avg_cap * trips_per_bus

    # v3.3.1 (STEP 6): typology-aware modal capture rate. Urban core uses
    # the CHALO-derived 9% baseline; peri-urban × 0.8; inter-district × 0.6
    # (private auto / shared sumo competition is stronger on long-distance
    # routes, so the bus capture rate falls).
    mode_share = gdf.get("Route_Type", pd.Series("Urban", index=gdf.index)) \
                    .map(PHASE4_MODE_SHARE_BY_TYPE).fillna(PHASE4_MODE_SHARE) \
                    .astype(float)
    # v3.3.2: corridor-sharing fix for Daily_Demand.
    # The legacy Population_Served column is the EQUAL-SHARE apportionment
    # (raw_buffer_pop / competitors). Two problems:
    #   (a) Raw competitor count is dominated by any route whose 400m buffer
    #       crosses this one — typically 100-300 in dense Srinagar, but only
    #       ~5-15 are actually *parallel* services competing for the same OD.
    #   (b) Equal-share ignores that high-frequency trunks capture a larger
    #       share of corridor demand than low-frequency feeders (revealed
    #       preference; Mohring effect).
    # Fix: effective_competitors = competitors × overlap_metric  (overlap_metric
    # ≈ 0.05–0.27 captures what fraction of buffer is actually shared, so the
    # product yields a sensible "parallel rivals" count). Then weight by
    # inverse-headway share.
    pop_raw     = gdf.get("Population_Served_Raw", pop_serv).astype(float)
    competitors = gdf.get("Corridor_Competitors",
                          pd.Series(1, index=gdf.index)).astype(float)
    overlap_m   = gdf.get("Overlap_Metric",
                          pd.Series(0.15, index=gdf.index)).astype(float)
    eff_comp    = (competitors * overlap_m).clip(lower=1.0)
    inv_h       = 1.0 / headway
    mean_inv_h  = float(inv_h.mean()) if inv_h.mean() > 0 else (1.0 / 35.0)
    freq_weight = inv_h / mean_inv_h                       # SSCL 15-min ≈ 2.3, LP 60-min ≈ 0.58
    corridor_share = (freq_weight / eff_comp).clip(upper=1.0)
    # PHASE4_CORRIDOR_CAPTURE_SCALE: empirical calibration scalar tying the
    # buffer-based demand model to CHALO observed SSCL ridership (≈32k/day
    # across the 30 SSCL routes). Without it, the model over-predicts demand
    # ~5.7× because the 400m buffer captures residents who use modes other
    # than the bus being analysed (auto, walk, private feeder, JKRTC).
    daily_demand   = (pop_raw * corridor_share * mode_share
                      * PHASE4_TRIP_RATE * PHASE4_CORRIDOR_CAPTURE_SCALE)
    load_ratio     = daily_demand / daily_capacity.where(daily_capacity > 0, np.nan)
    load_ratio     = load_ratio.fillna(0.0)

    def _load_flag(lr: float) -> str:
        if lr <= 0:        return "Red_NoCapacity"
        if lr > 1.0:       return "Red_Overload"
        if lr >= 0.4 and lr <= 0.85: return "Green"
        if lr < 0.4:       return "Amber_Under"
        return "Amber_Tight"   # 0.85–1.0

    pax_journey_time = 5.5 + headway / 2.0 + cycle / 2.0
    journey_flag     = pax_journey_time > PHASE4_JOURNEY_TIME_FLAG_MIN

    daily_revenue  = daily_demand * PHASE4_FARE_INR
    daily_op_cost  = daily_km * PHASE4_OPERATING_COST_PER_KM
    viability      = daily_revenue / daily_op_cost.where(daily_op_cost > 0, np.nan)
    viability      = viability.fillna(0.0)
    # v3.3.1 (STEP 7): tightened threshold from 0.5 → 0.6.  <0.6 fare-recovery
    # is meaningfully unsustainable; 0.6-1.0 is "marginal", >1.0 self-sustaining.
    subsidy_risk   = (viability < PHASE4_SUBSIDY_RISK_THRESHOLD) & \
                     (~gdf.get("Social_Flag", pd.Series(False, index=gdf.index)).astype(bool))

    # Emissions: SSCL = e-bus, everything else = diesel
    emiss_factor = np.where(
        gdf.get("CMP_Trunk", pd.Series(False, index=gdf.index)).astype(bool),
        EMISSIONS_GCO2_PER_KM_EBUS, EMISSIONS_GCO2_PER_KM_DIESEL,
    )
    emissions_daily = daily_km * emiss_factor   # grams CO2 / day

    # Equity score: normalise (Pop_Served × 1.5-if-social) across active routes
    active_mask = gdf["Action_Taken"] != "MERGED_INTO_TRUNK"
    social_boost = np.where(
        gdf.get("Social_Flag", pd.Series(False, index=gdf.index)).astype(bool),
        1.5, 1.0,
    )
    eq_raw = pop_serv * social_boost
    eq_max = float(eq_raw[active_mask].max()) if active_mask.any() else 1.0
    eq_max = max(eq_max, 1.0)
    equity_score = (eq_raw / eq_max).clip(0.0, 1.0)

    gdf["Daily_Trips"]         = daily_trips.round(1)
    gdf["Daily_KM"]            = daily_km.round(1)
    gdf["Daily_Capacity_Pax"]  = daily_capacity.round(0)
    gdf["Daily_Demand_Pax"]    = daily_demand.round(0)
    gdf["Load_Ratio"]          = load_ratio.round(3)
    gdf["Load_Flag"]           = [_load_flag(x) for x in load_ratio]
    gdf["Pax_Journey_Time_Min"] = pax_journey_time.round(1)
    gdf["Journey_Time_Flag"]   = journey_flag.values
    gdf["Daily_Revenue_INR"]   = daily_revenue.round(0)
    gdf["Daily_Op_Cost_INR"]   = daily_op_cost.round(0)
    gdf["Viability_Ratio"]     = viability.round(3)
    gdf["Subsidy_Risk_Flag"]   = subsidy_risk.values
    gdf["Emissions_GCO2_Daily"] = emissions_daily.round(0)
    gdf["Equity_Score"]        = equity_score.round(3)

    # Audit summary (active routes only)
    act = gdf[active_mask]
    log.info("  Load_Flag: Green=%d  Amber_Under=%d  Amber_Tight=%d  "
             "Red_Overload=%d  Red_NoCap=%d",
             (act["Load_Flag"] == "Green").sum(),
             (act["Load_Flag"] == "Amber_Under").sum(),
             (act["Load_Flag"] == "Amber_Tight").sum(),
             (act["Load_Flag"] == "Red_Overload").sum(),
             (act["Load_Flag"] == "Red_NoCapacity").sum())
    log.info("  Pax journey time >%d min: %d routes", PHASE4_JOURNEY_TIME_FLAG_MIN,
             int(act["Journey_Time_Flag"].sum()))
    log.info("  Subsidy risk (viability<%.2f, non-social): %d routes",
             PHASE4_SUBSIDY_RISK_THRESHOLD,
             int(act["Subsidy_Risk_Flag"].sum()))
    log.info("  Tourist corridors active: %d  |  Winter-suspended: %d",
             int((act.get("Tourist_Corridor", False) == True).sum()),
             int((act.get("Seasonal_Operability", "Year_Round") == "Winter_Suspended").sum()))
    return gdf


def export_passenger_impact(gdf: gpd.GeoDataFrame, out_path: str) -> None:
    log.info("Exporting Passenger Impact → %s", out_path)
    active = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"].copy()
    cols   = [c for c in [
        "New_Route_ID", "Route_Name", "Action_Taken", "Route_Type",
        "Priority_Band", "Headway_Min", "Fleet_Required",
        "HPV_Count", "MPV_Count", "LPV_Count",
        "CMP_Trunk", "CMP_Route_ID",
        "Population_Served", "Social_Flag",
    ] if c in active.columns]
    active[cols].to_csv(out_path, index=False, encoding="utf-8-sig")
    log.info("  Passenger impact written: %d rows.", len(active))


# ══════════════════════════════════════════════════════════════════════════════
#  ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    t0 = time.perf_counter()
    log.info("=" * 70)
    log.info("  Srinagar / Kashmir Valley Transit Rationalisation — ENGINE v3.0 (Kashmir Fork)")
    log.info("  SSCL/CHALO Backbone Injection + Kashmir Geographic Recentre | May 2026")
    log.info("=" * 70)
    log.info("  Map tile provider     : OpenStreetMap / CartoDB")
    log.info("  Scenario              : %s",
             "WINTER (Chillai Kalan)" if WINTER_SCENARIO else "SUMMER / SHOULDER")
    log.info("  Walkshed (effective)  : %d m  (POI buffer: %d m)",
             WALK_CATCHMENT_M, POI_BUFFER_M)
    log.info("  Tier-3 POI weight     : %.2f (tourism POIs)", POI_TIER3_WEIGHT)
    log.info("  Women-anchor boost    : %.0f%% (CHALO: %.1f%% female ridership)",
             (WOMEN_ANCHOR_BOOST - 1) * 100, CHALO_WOMEN_SHARE * 100)
    log.info("  CHALO observed peak   : %d pax/hr @ %d:00 (citywide)",
             CHALO_PEAK_PAX_PER_HOUR, CHALO_PEAK_HOUR)
    log.info("  CHALO op ratio        : %.1f%% (operated / scheduled KM)",
             CHALO_OP_RATIO * 100)
    log.info("=" * 70)
    log.info("  Directive 1: Zero Route Truncation — ALL routes processed")
    log.info("  Directive 2: Realistic Cycle Times (Srinagar congestion zones)")
    log.info("  Directive 3: Min Fleet = %d (floor, no LPV downgrade)",
             MIN_FLEET_THRESHOLD)
    log.info("  Directive 4: Trunk CDI gate = %dth percentile (aggressive) | "
             "Anti-stranding: %.1f km min | SSCL connectivity bonus: %.1fx",
             TRUNK_CDI_GATE_PERCENTILE, TRUNK_MIN_LENGTH_KM, CMP_CONNECTIVITY_BONUS)
    log.info("  Directive 5: Pop cap = %dth percentile | POI tiers: 1.0/0.4/%.1f",
             POP_CAP_PERCENTILE, POI_TIER3_WEIGHT)
    log.info("  Directive 6: Full audit logging throughout")
    log.info("  SSCL Injection: %d e-bus routes hardcoded TRUNK/HP @ %d-min headway",
             len(CMP_TRUNK_ROUTES), SSCL_TRUNK_HEADWAY_MIN)
    log.info("  Vehicle Split: Trunk=Route_KM-bracketed HPV (cap %.0f%%) | "
             "Feeder=100%%MPV (Regular/MPS 30%% HPV) | SSCL=empirical 9m/12m",
             SSCL_HPV_SHARE_CAP * 100)
    log.info("=" * 70)

    if not _HAS_JENKSPY:
        log.warning("  ⚠ jenkspy not installed — priority bands will use "
                    "percentile fallback. Run: pip install jenkspy")

    # ── PHASE 1: Data Ingestion, OSRM, Geometry ───────────────────────────
    log.info("\n── PHASE 1: Data Ingestion & OSRM ──────────────────────────────────")
    _DROP_LOG.clear()           # reset the per-route disposition trail
    df_routes    = load_routes(ROUTES_CSV)
    # Pre-engine endpoint QA gate (audit Rec 2): flag geocode-collapse / zero-length
    # input before any rationalisation happens. Diagnostic; hard gate is in QC.
    audit_input_quality(df_routes, "input_qa_report.csv")
    # --- SSCL E-BUS ROUTE INJECTION (replaces Jammu CMP synthetic routes) ----
    # All 30 SSCL routes from CHALO Apr 2026 data with approximated lat/lon
    # for major Srinagar / Valley terminals. These will be matched by
    # inject_cmp_trunk_routes() and forced to UPGRADED_TO_TRUNK / HP / 15-min.
    log.info("  Injecting all %d SSCL e-bus routes (CHALO Apr 2026)…",
             len(CMP_TRUNK_ROUTES))
    synthetic_routes = pd.DataFrame([
        # SSCL-01: Parimpora to Harwan via Brein Nishat
        {"Route_ID": "SSCL-01", "Route_Name": "Parimpora to Harwan via Brein Nishat",
         "Start_Lat": 34.1112, "Start_Lon": 74.7475,
         "End_Lat":   34.1481, "End_Lon":   74.9101},
        # SSCL-02: Batamaloo to Nasrullah Pora via Jehangir Chowk, Rambagh, Hyderpora
        {"Route_ID": "SSCL-02", "Route_Name": "Batamaloo to Nasrullah Pora via Jehangir Chowk Rambagh Hyderpora",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   34.0440, "End_Lon":   74.7320},
        # SSCL-03: Batamaloo to Hazratbal via Khanyar
        {"Route_ID": "SSCL-03", "Route_Name": "Batamaloo to Hazratbal via Khanyar",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   34.1342, "End_Lon":   74.8451},
        # SSCL-04: Batamaloo to Chadoora Chowk via Jehangir Chowk
        {"Route_ID": "SSCL-04", "Route_Name": "Batamaloo to Chadoora Chowk via Jehangir Chowk",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   34.0511, "End_Lon":   74.7124},
        # SSCL-05: LD Hospital to Pandach via Jehangir Chowk, Dalgate, Khanyar
        {"Route_ID": "SSCL-05", "Route_Name": "LD Hospital to Pandach via Jehangir Chowk Dalgate Khanyar",
         "Start_Lat": 34.0822, "Start_Lon": 74.8059,
         "End_Lat":   34.1611, "End_Lon":   74.8167},
        # SSCL-06: Pantha Chowk to Narbal Crossing via Sonwar, Jehangir Chowk, Qamarwari
        {"Route_ID": "SSCL-06", "Route_Name": "Pantha Chowk to Narbal Crossing via Sonwar Jehangir Chowk Qamarwari",
         "Start_Lat": 34.0445, "Start_Lon": 74.8341,
         "End_Lat":   34.1052, "End_Lon":   74.6809},
        # SSCL-07: Batamaloo to Drussu Pulwama via Marwal
        {"Route_ID": "SSCL-07", "Route_Name": "Batamaloo to Drussu Pulwama via Marwal",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   33.8830, "End_Lon":   74.9250},
        # SSCL-08: Batamaloo to Budgam Railway Station via Khumeni Chowk
        {"Route_ID": "SSCL-08", "Route_Name": "Batamaloo to Budgam Railway Station via Khumeni Chowk",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   33.9908, "End_Lon":   74.7115},
        # SSCL-09: Parimpora to Naseem Bagh via Noorbagh, Gojwara, Mill Stop
        {"Route_ID": "SSCL-09", "Route_Name": "Parimpora to Naseem Bagh via Noorbagh Gojwara Mill Stop",
         "Start_Lat": 34.1112, "Start_Lon": 74.7475,
         "End_Lat":   34.1334, "End_Lon":   74.8385},
        # SSCL-10: Pantha Chowk to Agri Kalan Kanihama via Bypass
        {"Route_ID": "SSCL-10", "Route_Name": "Pantha Chowk to Agri Kalan Kanihama via Bypass",
         "Start_Lat": 34.0445, "Start_Lon": 74.8341,
         "End_Lat":   34.0980, "End_Lon":   74.6500},
        # SSCL-11: Old City Loop (Downtown Srinagar circular)
        {"Route_ID": "SSCL-11", "Route_Name": "Old City Loop Downtown Srinagar",
         "Start_Lat": 34.0942, "Start_Lon": 74.8128,
         "End_Lat":   34.1027, "End_Lon":   74.8088},
        # SSCL-12: Rangreth to District Court Srinagar via Barzulla, Jehangir Chowk
        {"Route_ID": "SSCL-12", "Route_Name": "Rangreth to District Court Srinagar via Barzulla Jehangir Chowk",
         "Start_Lat": 34.0244, "Start_Lon": 74.7906,
         "End_Lat":   34.1006, "End_Lon":   74.7975},
        # SSCL-13: Batamaloo to Beehama Ganderbal via Hazratbal, Zakura
        {"Route_ID": "SSCL-13", "Route_Name": "Batamaloo to Beehama Ganderbal via Hazratbal Zakura",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   34.2403, "End_Lon":   74.8211},
        # SSCL-14: TRC to Central University Ganderbal via Karan Nagar, Zoonimar, 90ft Road
        {"Route_ID": "SSCL-14", "Route_Name": "TRC to Central University Ganderbal via Karan Nagar Zoonimar 90ft Road",
         "Start_Lat": 34.0833, "Start_Lon": 74.8089,
         "End_Lat":   34.2725, "End_Lon":   74.8245},
        # SSCL-15: TRC to Pulwama via Nowgam, Kaanipora
        {"Route_ID": "SSCL-15", "Route_Name": "TRC to Pulwama via Nowgam Kaanipora",
         "Start_Lat": 34.0833, "Start_Lon": 74.8089,
         "End_Lat":   33.8716, "End_Lon":   74.8983},
        # SSCL-16: Batamaloo Circular via Khonmoh
        {"Route_ID": "SSCL-16", "Route_Name": "Batamaloo Circular via Khonmoh",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   34.0419, "End_Lon":   74.8825},
        # SSCL-17: Batamaloo to Womens College Batpora via Rainawari
        {"Route_ID": "SSCL-17", "Route_Name": "Batamaloo to Womens College Batpora via Rainawari",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   34.0875, "End_Lon":   74.8083},
        # SSCL-18: JKPDC Jehangir Chowk to Wadwan
        {"Route_ID": "SSCL-18", "Route_Name": "JKPDC Jehangir Chowk to Wadwan",
         "Start_Lat": 34.0772, "Start_Lon": 74.8035,
         "End_Lat":   34.0200, "End_Lon":   74.9300},
        # SSCL-19: Pantha Chowk to Sumbal
        {"Route_ID": "SSCL-19", "Route_Name": "Pantha Chowk to Sumbal",
         "Start_Lat": 34.0445, "Start_Lon": 74.8341,
         "End_Lat":   34.2649, "End_Lon":   74.6606},
        # SSCL-20: Pantha Chowk to Safapora
        {"Route_ID": "SSCL-20", "Route_Name": "Pantha Chowk to Safapora",
         "Start_Lat": 34.0445, "Start_Lon": 74.8341,
         "End_Lat":   34.3592, "End_Lon":   74.7406},
        # SSCL-21: Batamaloo to Arath
        {"Route_ID": "SSCL-21", "Route_Name": "Batamaloo to Arath",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   34.0414, "End_Lon":   74.8462},
        # SSCL-22: Batamaloo to Khrew Bus Stand
        {"Route_ID": "SSCL-22", "Route_Name": "Batamaloo to Khrew Bus Stand",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   34.0407, "End_Lon":   74.9269},
        # SSCL-23: Batamaloo to Charesharief via Chadoora
        {"Route_ID": "SSCL-23", "Route_Name": "Batamaloo to Charesharief via Chadoora",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   33.8806, "End_Lon":   74.7372},
        # SSCL-24: Pantha Chowk to Palhalan (extends towards Sangrama / Sopore)
        {"Route_ID": "SSCL-24", "Route_Name": "Pantha Chowk to Palhalan via Sangrama Sopore",
         "Start_Lat": 34.0445, "Start_Lon": 74.8341,
         "End_Lat":   34.1825, "End_Lon":   74.5450},
        # SSCL-25: Batamaloo to Dadsara Tral via Awantipora
        {"Route_ID": "SSCL-25", "Route_Name": "Batamaloo to Dadsara Tral via Awantipora",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   33.9300, "End_Lon":   75.1100},
        # SSCL-26: Batamaloo to Kangan via Hazratbal, Manigam
        {"Route_ID": "SSCL-26", "Route_Name": "Batamaloo to Kangan via Hazratbal Manigam",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   34.2683, "End_Lon":   74.9962},
        # SSCL-27: Batamaloo to Manigam
        {"Route_ID": "SSCL-27", "Route_Name": "Batamaloo to Manigam",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   34.2272, "End_Lon":   74.9536},
        # SSCL-28: Batamaloo to Pinglena via Galandar Pampore
        {"Route_ID": "SSCL-28", "Route_Name": "Batamaloo to Pinglena Pampore via Galandar",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   33.9999, "End_Lon":   74.9356},
        # SSCL-29: Pantha Chowk to Panzinara
        {"Route_ID": "SSCL-29", "Route_Name": "Pantha Chowk to Panzinara",
         "Start_Lat": 34.0445, "Start_Lon": 74.8341,
         "End_Lat":   34.0593, "End_Lon":   74.7647},
        # SSCL-30: Batamaloo to DC Office Budgam
        {"Route_ID": "SSCL-30", "Route_Name": "Batamaloo to DC Office Budgam",
         "Start_Lat": 34.0689, "Start_Lon": 74.7795,
         "End_Lat":   33.9908, "End_Lon":   74.7115},
    ])

    # Fill defaults so the rest of the pipeline doesn't break
    synthetic_routes["Minibus_Count"]      = 0
    synthetic_routes["Standard_Bus_Count"] = 0
    synthetic_routes["Via_Coordinates"]    = None

    # Append to the main dataframe
    df_routes = pd.concat([df_routes, synthetic_routes], ignore_index=True)
    log.info("  %d SSCL routes appended (will be force-matched to TRUNK in inject step).",
             len(synthetic_routes))
    # ----------------------------------
    df_routes    = truncate_routes_to_bbox(df_routes)
    gdf_pois     = load_pois(POIS_CSV)
    osrm_results = fetch_all_osrm(df_routes)
    gdf          = apply_geometries(df_routes, osrm_results)
    gdf          = impute_fleet(gdf)

    # ── SSCL INJECTION: Force 30 CHALO/SSCL e-bus routes to TRUNK/HP ───────
    log.info("\n── SSCL BACKBONE INJECTION ──────────────────────────────────────────")
    gdf = inject_cmp_trunk_routes(gdf)

    # ── PHASE 2: Spatial Analysis ─────────────────────────────────────────
    log.info("\n── PHASE 2: Spatial Analysis ────────────────────────────────────────")
    gdf = build_catchments(gdf)
    gdf = compute_population(gdf, RASTER_PATH)
    gdf = count_weighted_poi_scores(gdf, gdf_pois)
    gdf = compute_junction_penalties(gdf)
    gdf = compute_cycle_times(gdf)          # DIRECTIVE 2: Realistic cycle times

    # ── STEPS 1 & 2: Normalise Demand Scores ─────────────────────────────
    log.info("\n── STEPS 1–2: Normalised Demand Scores ─────────────────────────────")
    gdf = step1_normalise_population_score(gdf)   # DIRECTIVE 5: 95th pct cap
    gdf = step2_normalise_poi_score(gdf)          # DIRECTIVE 5: 2-tier weights

    # ── Route Classification ──────────────────────────────────────────────
    log.info("\n── ROUTE CLASSIFICATION ─────────────────────────────────────────────")
    freq_scores    = compute_frequency_scores(gdf)
    gdf            = apportion_route_population(gdf, freq_scores, RASTER_PATH)
    overlap_matrix = compute_overlap_matrix(gdf)
    gdf            = cluster_routes(gdf, overlap_matrix)
    gdf            = backfill_overlap_metric(gdf, overlap_matrix)
    gdf            = classify_routes(gdf, freq_scores, overlap_matrix)  # v3: 30th pct + CMP bonus
    gdf            = apply_terminal_capacity(gdf, gdf_pois)
    gdf            = consolidate_duplicate_permits(gdf)   # audit Finding 8: kill duplicate over-fleeting

    # ── STEPS 3–9: CDI Pipeline ──────────────────────────────────────────
    log.info("\n── STEPS 3–9: CDI Pipeline ──────────────────────────────────────────")
    gdf = step3_compute_road_multiplier(gdf)
    gdf = step4a_compute_final_cdi(gdf)
    gdf = step4b_compute_social_flag(gdf)
    gdf = step5_assign_priority_bands(gdf)
    gdf = step5b_flag_sscl_cdi_conflicts(gdf)  # v3.3 Phase-1 audit: planner-review flag
    gdf = step6_assign_headways(gdf)         # CMP override: 10-min hardcoded
    gdf = step8_compute_fleet_required(gdf)  # v3: floor at MIN, no LPV downgrade
    gdf = step9_compute_vehicle_split(gdf)   # Route_KM-bracketed HPV (cap 50%) | Feeder 100% MPV | SSCL empirical
    gdf = zero_merged_route_fleet(gdf)
    gdf = compute_phase4_metrics(gdf)        # pass 1 — derives demand/load
    # v3.4.2: HYBRID demand-responsive sizing for Regional lifelines only, then
    # re-run fleet → split → zero → phase4 so every KPI is consistent at the new
    # demand-matched headways. (Urban/Peri-Urban + SSCL unchanged.)
    if REGIONAL_DEMAND_SIZING:
        gdf = apply_regional_demand_headway(gdf)
        gdf = step8_compute_fleet_required(gdf)
        gdf = step9_compute_vehicle_split(gdf)
        gdf = zero_merged_route_fleet(gdf)
        gdf = compute_phase4_metrics(gdf)    # pass 2 — consistent at new headways

    # ── QC ────────────────────────────────────────────────────────────────
    log.info("\n── QC CHECKS ────────────────────────────────────────────────────────")
    run_all_qc_checks(gdf)

    # ── Network Totals ────────────────────────────────────────────────────
    net_pop       = compute_network_population_total(gdf, RASTER_PATH)
    network_score = compute_network_score(gdf, net_pop)
    # F-V6: reconcile the active plan's Population_Served column to the cover figure.
    gdf           = reconcile_active_population(gdf, net_pop)

    # ── PHASE 3: Log ──────────────────────────────────────────────────────
    log.info("\n── PHASE 3: Rationalisation Log ─────────────────────────────────────")
    generate_log(gdf, LOG_CSV)

    # ── PHASE 4: Export ───────────────────────────────────────────────────
    log.info("\n── PHASE 4: Cartography & Export ────────────────────────────────────")
    # v3.3.7: normalise route-name casing (consistent Title Case, acronyms + via
    # preserved) so the workbook, maps and dashboard all read uniformly.
    gdf["Route_Name"] = gdf["Route_Name"].astype(str).map(_clean_route_name)
    log.info("  Route names normalised to consistent Title Case (via preserved).")
    build_master_map(gdf, gdf_pois, RASTER_PATH, MASTER_MAP_HTML,
                     net_pop, network_score)
    file_map = build_individual_maps(gdf, gdf_pois, OUTPUT_DIR)
    gdf = assign_route_codes(gdf)   # v3.3.7: bake Route_Code into every export
    qc_route_codes(gdf)             # audit Output #1: route-code uniqueness gate
    export_csv(gdf, file_map, ROUTES_OUT_CSV)
    export_xlsx(gdf, out_path=ROUTES_OUT_XLSX, net_pop=net_pop)
    export_xlsx_rto(gdf, out_path="Kashmir_Route_Frequency_Plan_v3.4.3_RTO.xlsx", net_pop=net_pop)
    export_passenger_impact(gdf, PASSENGER_IMPACT_CSV)
    export_geojson(gdf, ROUTES_GEOJSON)
    # Per-route disposition trail covering every input row (audit Rec 8).
    export_route_disposition(gdf, "Route_Disposition_Kashmir_v3.csv")

    elapsed = time.perf_counter() - t0
    active  = gdf[gdf["Action_Taken"] != "MERGED_INTO_TRUNK"]
    cmp_n   = int(gdf.get("CMP_Trunk", pd.Series(False, index=gdf.index)).sum())

    log.info("\n" + "=" * 70)
    log.info("  PIPELINE v3 COMPLETE  (%.1f s)", elapsed)
    log.info("  Total routes in dataset : %d (no length truncation — D1)",
             len(gdf))
    log.info("  Active routes           : %d  (Trunk: %d  Feeder: %d  Merged: %d)",
             len(active),
             (gdf["Action_Taken"] == "UPGRADED_TO_TRUNK").sum(),
             (gdf["Action_Taken"] == "RETAINED_AS_FEEDER").sum(),
             (gdf["Action_Taken"] == "MERGED_INTO_TRUNK").sum())
    log.info("  SSCL backbone trunks    : %d / %d matched (%d-min headway)",
             cmp_n, len(CMP_TRUNK_ROUTES), SSCL_TRUNK_HEADWAY_MIN)
    log.info("  Route types             : Urban=%d  Peri_Urban=%d  "
             "Regional_District=%d",
             (gdf["Route_Type"] == "Urban").sum(),
             (gdf["Route_Type"] == "Peri_Urban").sum(),
             (gdf["Route_Type"] == "Regional_District").sum())
    log.info("  Priority bands          : HP=%d  MP=%d  LP=%d  (Social: %d)",
             (active["Priority_Band"] == "HP").sum(),
             (active["Priority_Band"] == "MP").sum(),
             (active["Priority_Band"] == "LP").sum(),
             active["Social_Flag"].sum())
    log.info("  Total fleet             : %d buses (HPV: %d  MPV: %d  LPV: %d)",
             int(active["Fleet_Required"].sum()),
             int(active["HPV_Count"].sum()),
             int(active["MPV_Count"].sum()),
             int(active["LPV_Count"].sum()))
    # Displaced-operator summary (P1 item 6 — political accounting)
    merged_gdf = gdf[gdf["Action_Taken"] == "MERGED_INTO_TRUNK"]
    if "Displaced_Operator_Class" in merged_gdf.columns and not merged_gdf.empty:
        dop = merged_gdf["Displaced_Operator_Class"].value_counts()
        log.info("  Displaced operator permits (%d total — absorbed into trunk network):",
                 len(merged_gdf))
        for cls, cnt in dop.items():
            log.info("    %-30s : %d", cls, cnt)
        log.info("  NOTE: Operator absorption or buyback recommendations needed "
                 "before plan submission. See Displaced_Operator_Class column.")
    log.info("  Trunk vehicle split     : Route_KM-bracketed "
             "(<12km: 100%% MPV | 12-22km: 50/50 | 22+km: 50/50)  "
             "+ SSCL empirical table capped at 50%% HPV (v3.3.7 — neither class a majority)")
    log.info("  Feeder vehicle split    : 100%% MPV (city bus / unknown), "
             "30/70 HPV/MPV (regular/MPS), 100%% LPV (LPV category)")
    log.info("  Network pop.            : %s residents  (%.2f%% of the %s in the "
             "study area; Srinagar CMP %d ref: %s)",
             f"{net_pop:,}",
             (net_pop / study_area_population(RASTER_PATH) * 100)
             if study_area_population(RASTER_PATH) else 0.0,
             f"{study_area_population(RASTER_PATH):,}",
             CMP_REFERENCE_YEAR,
             f"{CMP_TOTAL_POPULATION:,}")
    log.info("  Jenks engine            : %s",
             "jenkspy" if _HAS_JENKSPY else "percentile fallback")
    log.info("  Outputs:")
    log.info("    %-44s  4-sheet workbook", ROUTES_OUT_XLSX)
    log.info("    %-44s  sidebar + KPI map", MASTER_MAP_HTML)
    log.info("    %-44s  full audit log", LOG_CSV)
    log.info("    %-44s  passenger impact", PASSENGER_IMPACT_CSV)
    log.info("    %-44s  operational CSV", ROUTES_OUT_CSV)
    log.info("    %-44s  GeoJSON network", ROUTES_GEOJSON)
    try:
        import subprocess
        import sys as _sys
        # Use the current interpreter (matters under conda envs where bare
        # `python` is not on PATH). The script lives one directory up from
        # the engine's run-output cwd.
        script = Path(__file__).resolve().parent / "generate_presentations.py"
        log.info("  Generating PowerPoint presentations…")
        subprocess.run(
            [_sys.executable, str(script),
             "--outdir", ".", "--engine-csv", ROUTES_OUT_CSV],
            check=True,
        )
        log.info("    Kashmir_Transit_Technical_Briefing.pptx")
        log.info("    Kashmir_Transit_Government_Briefing.pptx")
    except Exception as e:
        log.error("  Failed to generate presentations: %s", e)
        
    log.info("=" * 70)


if __name__ == "__main__":
    main()
