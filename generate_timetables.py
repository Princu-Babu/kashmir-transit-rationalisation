#!/usr/bin/env python
"""
Kashmir_Timetables_v1.xlsx — expand the plan's headways into ACTUAL departure
boards, the artifact a depot manager operates from.

The plan speaks in headways ("a bus every 20 min"); an RTO runs timetables.
For every active route this expands headway -> departure times from EACH
terminus across the service day. The service window is anchored on MEASURED
reality (bus-sathi-trace GPS): in-service time ramps from ~08:00 and collapses
after ~19:00, so the standard day is 08:00-19:00 (SSCL e-bus backbone keeps
its design 07:00-20:00). Symmetric service: the same board applies from both
termini (a bus departs each end every headway).

Sheets: Read Me + one per origin district (route-code prefix), ordered by
sector/route code. Each route = one row: code, name, headway, cycle, fleet,
first/last departure + the full departure board.

Output: outputs_v3.4.5/Kashmir_Timetables_v1.xlsx (+ dashboard copy)
Run:  & "D:\\plotting\\ana\\python.exe" generate_timetables.py
"""
import sys, os, shutil
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "outputs_v3.4.5"
CSV = f"{OUT_DIR}/Rationalised_Routes_Kashmir_v3.csv"
OUT = f"{OUT_DIR}/Kashmir_Timetables_v1.xlsx"
DASH = r"E:\dash\bus-sathi-dashboard\public\route-rationalization-kashmir"

DAY_START, DAY_END = 8 * 60, 19 * 60          # measured: ramps 08:00, collapses ~19:00
SSCL_START, SSCL_END = 7 * 60, 20 * 60        # e-bus backbone design day

DISTRICT = {"SR": "Srinagar", "BG": "Budgam", "GB": "Ganderbal", "BR": "Baramulla",
            "BP": "Bandipora", "PW": "Pulwama", "SP": "Shopian", "AN": "Anantnag",
            "KG": "Kulgam", "KW": "Kupwara"}

TEAL = "0F6E56"; MIST = "E9F2EE"; INK = "102A2E"; WHITE = "FFFFFF"
HDR_FONT = Font(bold=True, color=WHITE, size=10.5, name="Calibri")
HDR_FILL = PatternFill("solid", fgColor=TEAL)
ALT = PatternFill("solid", fgColor=MIST)
THIN = Border(*[Side(style="thin", color="D9E2DE")] * 4)


def hhmm(m):
    return f"{m // 60:02d}:{m % 60:02d}"


def board(headway, start, end):
    return [hhmm(t) for t in range(start, end + 1, int(headway))]


def main():
    df = pd.read_csv(CSV)
    act = df[df.Action_Taken != "MERGED_INTO_TRUNK"].copy()
    act = act.drop_duplicates(subset="New_Route_ID")
    print(f"Active routes: {len(act)}")

    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet("Read Me")
    notes = [
        ("Kashmir Route Timetables v1 — departure boards from the v3.4.5 frequency plan", Font(bold=True, size=14, name="Cambria", color=INK)),
        ("", None),
        ("HOW THESE WERE BUILT", Font(bold=True, size=11, color=TEAL)),
        ("Each route's planned headway is expanded into departure times from EACH terminus. "
         "Service runs symmetrically: the same board applies at both ends of the route.", None),
        ("Standard service day 08:00–19:00 — anchored on MEASURED operations from the Bus Sathi "
         "GPS layer (in-service time ramps from ~08:00 and collapses after ~19:00). "
         "SSCL e-bus backbone uses its design day 07:00–20:00.", None),
        ("", None),
        ("HOW TO READ", Font(bold=True, size=11, color=TEAL)),
        ("One row per route: code · name · a bus every N min · round-trip cycle · buses assigned · "
         "first/last departure · the full board. Sheets are grouped by ORIGIN district (route-code prefix).", None),
        ("These are PLAN timetables (design frequencies), not a transcription of current informal "
         "operations. Extending service past 19:00 is a policy decision the measured data flags as a gap.", None),
    ]
    for i, (t, f) in enumerate(notes, 1):
        c = ws.cell(i, 1, t)
        if f: c.font = f
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 105

    act["d2"] = act.Route_Code.astype(str).str[:2]
    cols = ["Route code", "Route", "Every (min)", "Cycle (min)", "Buses", "First dep", "Last dep", "Departures (from each terminus)"]
    n_rows = 0
    for d2 in sorted(act.d2.unique(), key=lambda x: (x != "SR", x)):
        dname = DISTRICT.get(d2, d2)
        sub = act[act.d2 == d2].sort_values("Route_Code")
        ws = wb.create_sheet(dname[:28])
        ws.cell(1, 1, f"{dname} — {len(sub)} routes (by origin district)").font = Font(bold=True, size=13, name="Cambria", color=INK)
        for j, cname in enumerate(cols, 1):
            c = ws.cell(3, j, cname); c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        r = 4
        for _, row in sub.iterrows():
            sscl = str(row.New_Route_ID).startswith("SSCL-")
            s, e = (SSCL_START, SSCL_END) if sscl else (DAY_START, DAY_END)
            hw = int(row.Headway_Min)
            b = board(hw, s, e)
            vals = [row.Route_Code, row.Route_Name, hw, round(float(row.Cycle_Time_Min), 0),
                    int(row.Fleet_Required), b[0], b[-1], "  ".join(b)]
            for j, v in enumerate(vals, 1):
                c = ws.cell(r, j, v); c.border = THIN
                c.font = Font(size=9.5, name="Calibri")
                c.alignment = Alignment(wrap_text=(j == 8), vertical="top")
                if (r % 2) == 0: c.fill = ALT
            ws.row_dimensions[r].height = max(14, 12 * (len("  ".join(b)) // 95 + 1))
            r += 1; n_rows += 1
        widths = [14, 30, 9, 9, 7, 9, 9, 95]
        for j, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A4"

    wb.save(OUT)
    shutil.copy2(OUT, os.path.join(DASH, "Kashmir_Timetables_v1.xlsx"))
    print(f"Wrote {OUT} ({n_rows} route boards) + dashboard copy")


if __name__ == "__main__":
    main()
