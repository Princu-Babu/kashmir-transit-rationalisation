"""
_sync_dashboard.py — copies the latest engine outputs into the
bus-sathi-dashboard public folder and regenerates the JSON data files.

Preserves prior official Route_Codes via Route_ID lookup and mints
TMP-K#### placeholders for any genuinely-new routes.
"""

from __future__ import annotations

import csv
import json
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Dict, List

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

ENGINE_OUT = Path("E:/kash/outputs_v3.4.4")
ENGINE_LOG = Path("E:/kash/engine_run_v3.4.4.log")

DASH_PUBLIC = Path("E:/dash/bus-sathi-dashboard/public/route-rationalization-kashmir")
DASH_DATA   = DASH_PUBLIC / "data"

# v3.3.7: also push the RTO master workbook, the pretty bus-schedule workbook,
# and the route-code workbook so the dashboard always serves the LIVE engine
# artefacts (previously the pretty workbook was hand-copied → went stale).
COPY_FILES = [
    "Rationalised_Routes_Kashmir_v3.csv",
    "Rationalised_Routes_Kashmir_v3.geojson",
    "Passenger_Impact_Kashmir_v3.csv",
    "Rationalisation_Log_Kashmir_v3.csv",
    "Kashmir_Route_Frequency_Plan_v3.xlsx",
    # NOTE: the 9-sheet "..._RTO.xlsx" master workbook is intentionally NOT pushed
    # to the dashboard (user ask — internal detail pack, not for public download).
    # It is also listed in STALE_FILES below so any previously-copied build is purged.
    "Kashmir_Route_Frequency_Plan_v3.4.4_RTO_Pretty.xlsx",
    "Kashmir_Route_Verification_Appendix_v3.4.4_RTO.xlsx",
    "Routes_with_Codes.xlsx",
    "Master_Transit_Map_Kashmir_v3.html",
]
COPY_DIRS = ["route_maps_kashmir"]

# v3.3.7: stale per-version downloads to purge from the dashboard public folder
# so the RTO can never grab an out-of-date workbook. The cleanup keeps exactly
# one RTO master + one pretty bus-schedule file (the v3.3.7 pair above).
STALE_FILES = [
    "Kashmir_Route_Frequency_Plan_v3.3.5_RTO.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.3.6_RTO.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.3.7_RTO.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.3.7_RTO_Pretty.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.3.8_RTO.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.3.8_RTO_Pretty.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.3.9_RTO.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.3.9_RTO_Pretty.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.4.0_RTO.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.4.0_RTO_Pretty.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.4.1_RTO.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.4.1_RTO_Pretty.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.4.2_RTO.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.4.2_RTO_Pretty.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.4.3_RTO.xlsx",
    "Kashmir_Route_Frequency_Plan_v3.4.3_RTO_Pretty.xlsx",
    # 9-sheet master workbook — hidden from the dashboard (user ask); purge if present
    "Kashmir_Route_Frequency_Plan_v3.4.4_RTO.xlsx",
    "Formatted_Kashmir_Routes_Pretty.xlsx",
    "Formatted_Kashmir_Routes.xlsx",
    "generate_route_codes (1).py",
    "kashmir_routes_geocoded.csv",
]


def _temp_route_code(seq: int) -> str:
    return f"TMP-K{seq:04d}"


def _copy_assets():
    for fname in COPY_FILES:
        src = ENGINE_OUT / fname
        if not src.exists():
            print(f"  ! missing {src}", file=sys.stderr); continue
        shutil.copy2(src, DASH_PUBLIC / fname)
        print(f"  copied {fname}")
    if ENGINE_LOG.exists():
        shutil.copy2(ENGINE_LOG, DASH_PUBLIC / "transit_v3.log.txt")
        print("  copied engine log -> transit_v3.log.txt")
    for dname in COPY_DIRS:
        src_dir = ENGINE_OUT / dname
        dst_dir = DASH_PUBLIC / dname
        if not src_dir.is_dir():
            print(f"  ! missing dir {src_dir}", file=sys.stderr); continue
        if dst_dir.exists():
            shutil.rmtree(dst_dir)
        shutil.copytree(src_dir, dst_dir)
        n = sum(1 for _ in dst_dir.rglob("*.html"))
        print(f"  copied {dname}/ ({n} html files)")


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _write_json(path: Path, rows: List[Dict[str, str]]):
    with path.open("w", encoding="utf-8") as fh:
        json.dump(rows, fh, ensure_ascii=False, separators=(",", ":"))
    print(f"  wrote {path.name} ({len(rows)} rows)")


def _load_prior_codes() -> Dict[str, str]:
    out: Dict[str, str] = {}
    try:
        res = subprocess.run(
            ["git", "show", "HEAD:public/route-rationalization-kashmir/data/routes.json"],
            cwd=str(DASH_PUBLIC.parents[1]),
            capture_output=True, text=True, encoding="utf-8", check=True,
        )
        data = json.loads(res.stdout)
        for r in data if isinstance(data, list) else []:
            rid  = str(r.get("Route_ID") or "").strip()
            code = str(r.get("Route_Code") or "").strip()
            # Skip blank, TMP- placeholders, and UNMATCHED entries from the
            # generate_route_codes.py output — those should re-mint as TMP- this run.
            if (rid and code
                    and not code.startswith("TMP-")
                    and code.upper() != "UNMATCHED"):
                out[rid] = code
    except Exception as exc:
        print(f"  ! could not load prior route codes ({exc})", file=sys.stderr)
    return out


def _load_fresh_codes() -> Dict[str, str]:
    """
    Read Routes_with_Codes.xlsx (output of generate_route_codes.py).
    Returns a Route_ID -> Route_Code map by joining the freshly-generated
    codes against the engine CSV on row position (both files are 342 rows
    in identical order — same source DataFrame).
    """
    out: Dict[str, str] = {}
    codes_xlsx = ENGINE_OUT / "Routes_with_Codes.xlsx"
    engine_csv = ENGINE_OUT / "Rationalised_Routes_Kashmir_v3.csv"
    if not codes_xlsx.exists():
        print(f"  ! no fresh Routes_with_Codes.xlsx — skipping fresh-code lookup",
                 file=sys.stderr)
        return out
    if not engine_csv.exists():
        return out
    try:
        import openpyxl
        wb = openpyxl.load_workbook(codes_xlsx, data_only=True)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        code_col = headers.index("Route_Code") + 1 if "Route_Code" in headers else None
        if code_col is None:
            return out
        # Codes in order:
        fresh_codes = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=code_col).value
            fresh_codes.append(str(v) if v is not None else "")
        # Engine CSV row order:
        with engine_csv.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            rids = [str(row.get("Route_ID", "")).strip() for row in reader]
        # Pair by index. UNMATCHED stays UNMATCHED so the fallback layer
        # promotes it to a TMP-K placeholder.
        for rid, code in zip(rids, fresh_codes):
            if (rid and code
                    and code.upper() != "UNMATCHED"
                    and not code.startswith("TMP-")):
                out[rid] = code
        print(f"  Fresh codes loaded: {len(out)} from {codes_xlsx.name} "
                f"({sum(1 for c in fresh_codes if c.upper()=='UNMATCHED')} UNMATCHED)")
    except Exception as exc:
        print(f"  ! could not load fresh Route_Codes ({exc})", file=sys.stderr)
    return out


def _add_codes(rows: List[Dict[str, str]],
               fresh: Dict[str, str],
               prior: Dict[str, str]) -> List[Dict[str, str]]:
    """
    Precedence per route:
      1. Embedded Route_Code in the CSV row (if present and not UNMATCHED).
      2. Freshly-generated code from generate_route_codes.py (this run).
      3. Carried-forward official code from the last commit on the dashboard.
      4. TMP-K#### placeholder.
    """
    out, embedded, fresh_n, prior_n, minted = [], 0, 0, 0, 0
    seq = 0
    for r in rows:
        seq += 1
        existing = (r.get("Route_Code") or "").strip()
        rid      = (r.get("Route_ID") or "").strip()
        is_unmatched = existing.upper() == "UNMATCHED"
        if existing and not existing.startswith("TMP-") and not is_unmatched:
            code = existing; embedded += 1
        elif rid in fresh:
            code = fresh[rid]; fresh_n += 1
        elif rid in prior:
            code = prior[rid]; prior_n += 1
        else:
            code = _temp_route_code(seq); minted += 1
        new = {"Route_Code": code}
        for k, v in r.items():
            if k == "Route_Code":
                continue
            new[k] = v
        out.append(new)
    print(f"  Route_Code:  embedded {embedded}  fresh {fresh_n}  "
          f"prior {prior_n}  TMP minted {minted}")
    return out


def _build_jsons():
    fresh = _load_fresh_codes()
    prior = _load_prior_codes()
    routes_rows = _add_codes(
        _read_csv_rows(ENGINE_OUT / "Rationalised_Routes_Kashmir_v3.csv"),
        fresh, prior,
    )
    code_by_id = {r["Route_ID"]: r["Route_Code"] for r in routes_rows if r.get("Route_ID")}

    def _backfill(rows):
        out = []
        seq = 0
        for r in rows:
            seq += 1
            rid = (r.get("Route_ID") or "").strip()
            existing = (r.get("Route_Code") or "").strip()
            if existing and not existing.startswith("TMP-") and existing.upper() != "UNMATCHED":
                code = existing
            elif rid in code_by_id:
                code = code_by_id[rid]
            else:
                code = _temp_route_code(seq)
            new = {"Route_Code": code}
            for k, v in r.items():
                if k == "Route_Code":
                    continue
                new[k] = v
            out.append(new)
        return out

    impact_rows = _backfill(_read_csv_rows(ENGINE_OUT / "Passenger_Impact_Kashmir_v3.csv"))
    log_rows    = _backfill(_read_csv_rows(ENGINE_OUT / "Rationalisation_Log_Kashmir_v3.csv"))

    _write_json(DASH_DATA / "routes.json", routes_rows)
    _write_json(DASH_DATA / "impact.json", impact_rows)
    _write_json(DASH_DATA / "log.json",    log_rows)

    csv_path = DASH_DATA / "Rationalised_Routes_Kashmir_v3.csv"
    if routes_rows:
        with csv_path.open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=list(routes_rows[0].keys()))
            writer.writeheader()
            writer.writerows(routes_rows)
        print(f"  wrote data/Rationalised_Routes_Kashmir_v3.csv ({len(routes_rows)} rows)")


def _purge_stale():
    for fname in STALE_FILES:
        p = DASH_PUBLIC / fname
        if p.exists():
            p.unlink()
            print(f"  purged stale {fname}")


def main():
    print(f"Sync {ENGINE_OUT.name} engine outputs -> dashboard...")
    DASH_DATA.mkdir(parents=True, exist_ok=True)
    _copy_assets()
    _purge_stale()
    _build_jsons()
    print("Done.")


if __name__ == "__main__":
    main()
