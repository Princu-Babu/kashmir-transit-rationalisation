#!/usr/bin/env python
"""Build the RTO verification appendix workbook for v3.4.4 from the audit CSVs."""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HDR=PatternFill('solid',fgColor='1F4E78'); HF=Font(color='FFFFFF',bold=True,size=11)
TITLE=Font(bold=True,size=14,color='1F4E78'); SUB=Font(italic=True,size=9,color='666666')
thin=Side(style='thin',color='D9D9D9'); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical='top'); CTR=Alignment(horizontal='center',vertical='center')
GROW=PatternFill('solid',fgColor='FCE4D6'); SHR=PatternFill('solid',fgColor='E2EFDA')
wb=Workbook()
def style_header(ws,row=1):
    for c in ws[row]:
        if c.value is not None: c.fill=HDR;c.font=HF;c.alignment=CTR;c.border=BORD

ws=wb.active; ws.title='Summary'
ws['A1']='Appendix R — Route Verification & Corrections (v3.4.4)'; ws['A1'].font=TITLE
ws['A2']='Independent real-world verification of all 186 active routes; 49 distance corrections applied. Prepared 2026-06-25.'; ws['A2'].font=SUB
rows=[('','',''),
 ('Verification outcome','Routes','Note'),
 ('PASS — carried forward unchanged',93,'Distance within +/-15% of real road km'),
 ('REVIEW — checked, one number off',88,'Real corridor; distance/pin reviewed'),
 ('FAIL — geocoding error',5,'4 corrected, 1 deferred to RTO stop register'),
 ('','',''),
 ('Corrections in v3.4.4','Count','Note'),
 ('Routes corrected (km set to verified real)',49,'inc. 4 re-geocoded+re-routed; 43 shrank, 6 grew'),
 ('   incl. re-geocoded + re-routed','4','Budgam, GBS, Hazratbal, Manigam (geometry redrawn)'),
 ('Routes deferred',44,'11 SSCL via-loop, 19 name-unverifiable, 14 within tolerance'),
 ('PASS routes unchanged',93,'byte-identical to v3.4.3'),
 ('','',''),
 ('Fleet impact','v3.4.3','v3.4.4'),
 ('Total buses',1044,1004),
 ('HPV / MPV / LPV','185 / 776 / 83','187 / 748 / 69'),
 ('Active routes / SSCL trunks / districts','186 / 30 / 10','186 / 30 / 10'),
]
r=4
for a,b,c in rows:
    ws.cell(r,1,a); ws.cell(r,2,b); ws.cell(r,3,c)
    if a in ('Verification outcome','Corrections in v3.4.4','Fleet impact'):
        for col in (1,2,3): ws.cell(r,col).fill=HDR; ws.cell(r,col).font=HF
    r+=1
ws.column_dimensions['A'].width=44; ws.column_dimensions['B'].width=16; ws.column_dimensions['C'].width=52
for row in ws.iter_rows(min_row=4,max_row=r-1,max_col=3):
    for c in row: c.alignment=WRAP

ws=wb.create_sheet('Corrections Applied')
ap=list(csv.DictReader(open('corrections_applied_v344.csv',encoding='utf-8-sig')))
cols=['Route_Code','Route_Name','Eng_Type','Old_KM','New_KM','Old_Fleet','New_Fleet','Confidence','Headway','Reason','Sources']
ws.append(['Route Code','Route Name','Type','Old km','Real km','Old fleet','New fleet','Conf.','Headway','Reason (verified)','Sources']); style_header(ws)
for x in sorted(ap,key=lambda r:(r['Confidence']!='HIGH',r['Route_Code'])):
    ws.append([x.get(c,'') for c in cols]); rr=ws.max_row
    try:
        f=SHR if float(x['New_KM'])<float(x['Old_KM']) else GROW
        ws.cell(rr,4).fill=f; ws.cell(rr,5).fill=f
    except: pass
for i,w in enumerate([14,30,15,8,8,9,9,7,9,60,40],1): ws.column_dimensions[get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=11):
    for c in row: c.alignment=WRAP; c.border=BORD
ws.freeze_panes='A2'

ws=wb.create_sheet('Deferred Worklist')
df=list(csv.DictReader(open('corrections_deferred_v344.csv',encoding='utf-8-sig')))
ws.append(['Route Code','Route Name','Why deferred / action']); style_header(ws)
for x in df: ws.append([x['Route_Code'],x['Route_Name'],x['Why_deferred']])
for i,w in enumerate([14,32,70],1): ws.column_dimensions[get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=3):
    for c in row: c.alignment=WRAP; c.border=BORD
ws.freeze_panes='A2'

ws=wb.create_sheet('Full Ledger (186)')
led=list(csv.DictReader(open('ROUTE_DEEPDIVE_LEDGER.csv',encoding='utf-8-sig')))
lc=['Route_Code','Route_Name','Eng_KM','Real_RoadKM','Verdict','Finding','Sources']
ws.append(['Route Code','Route Name','Model km','Real road km','Verdict','Finding','Sources']); style_header(ws)
vf={'PASS':PatternFill('solid',fgColor='E2EFDA'),'REVIEW':PatternFill('solid',fgColor='FFF2CC'),'FAIL':PatternFill('solid',fgColor='F8CBAD')}
for x in led:
    ws.append([x.get(c,'') for c in lc]); rr=ws.max_row
    ws.cell(rr,5).fill=vf.get(x['Verdict'].strip(),PatternFill())
for i,w in enumerate([14,30,9,13,9,60,34],1): ws.column_dimensions[get_column_letter(i)].width=w
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,max_col=7):
    for c in row: c.alignment=WRAP; c.border=BORD
ws.freeze_panes='A2'

out='outputs_v3.4.4/Kashmir_Route_Verification_Appendix_v3.4.4_RTO.xlsx'
wb.save(out); print('wrote',out)
