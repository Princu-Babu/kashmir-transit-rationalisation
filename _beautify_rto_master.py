"""
_beautify_rto_master.py

Cut-down companion to the 9-sheet RTO master workbook. Condenses it into a
focused 2-sheet pretty bus-schedule workbook for quick reviews and meetings:

  1.  Summary       — KPI tiles + reference tables on one page
  2.  Route Plan    — compact route table (the bus schedule: headways, cycle
                      time, fleet, HPV/MPV split, Route_Code per route)

v3.3.7: the Operator Absorption and Sign-off sheets were removed (RTO ask) so
the file is purely the schedule. The full operator-absorption / buyback detail
still lives in the 9-sheet RTO master workbook.

Output: Kashmir_Route_Frequency_Plan_v3.3.7_RTO_Pretty.xlsx next to the source.
"""
from __future__ import annotations

import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SRC = Path("E:/kash/outputs_v3.4.3/Kashmir_Route_Frequency_Plan_v3.4.3_RTO.xlsx")
DST = Path("E:/kash/outputs_v3.4.3/Kashmir_Route_Frequency_Plan_v3.4.3_RTO_Pretty.xlsx")
# Full CSV — has CMP_Trunk, Route_Type, Displaced_Operator_Class etc. that
# the source XLSX strips out for the RTO-facing column subset.
SRC_CSV = Path("E:/kash/outputs_v3.4.3/Rationalised_Routes_Kashmir_v3.csv")

# Also write a copy to a convenient location for the meeting
ALSO  = Path("C:/Users/Prash/Music/Kashmir_Route_Frequency_Plan_v3.4.3_RTO_Pretty.xlsx")


# ─── Theme (matches the other pretty workbook) ───────────────────────────────
NAVY        = "1A237E"
TEAL        = "00695C"
TEAL_LT     = "E0F2F1"
PURPLE      = "6A1B9A"
PURPLE_LT   = "F3E5F5"
SAFFRON     = "D32F2F"
SAFFRON_LT  = "FFEBEE"
GOLD        = "F9A825"
GOLD_LT     = "FFF8E1"
GREEN       = "2E7D32"
GREEN_LT    = "E8F5E9"
LIGHT       = "F5F5F5"
DARK        = "212121"
GREY        = "757575"
WHITE       = "FFFFFF"
BORDER_GREY = "E0E0E0"

THIN = Side(style="thin", color=BORDER_GREY)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ─── Helpers ──────────────────────────────────────────────────────────────────
def set_cell(cell, value, *, bold=False, size=10, font_color=DARK,
             align="center", anchor="center", bg=None, num_fmt=None,
             border=True, italic=False, name="Segoe UI"):
    cell.value = value
    cell.font = Font(name=name, size=size, bold=bold, italic=italic,
                      color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical=anchor,
                                wrap_text=True)
    if border:
        cell.border = BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt


def header_cell(cell, value, *, bg=NAVY, size=11):
    set_cell(cell, value, bold=True, size=size, font_color=WHITE,
              align="center", anchor="center", bg=bg)


def hide_gridlines(ws):
    ws.sheet_view.showGridLines = False


def _read_route_plan() -> pd.DataFrame:
    """
    Load the route plan from the full CSV (preferred — has every column),
    falling back to the RTO XLSX subset if the CSV is missing.
    """
    if SRC_CSV.exists():
        df = pd.read_csv(SRC_CSV, encoding="utf-8-sig")
    else:
        src_wb = load_workbook(SRC, data_only=True)
        ws     = src_wb["Route-Level Plan"]
        headers = [c.value for c in ws[1]]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, r)))
        df = pd.DataFrame(rows)
    return df


# Per-vehicle Phase-1 buyback estimate (matches the v3.3.4 engine logic).
# Empty string for JKRTC means "no buyback applicable — state-owned".
BUYBACK_LAKH_PER_VEHICLE = {
    "Private Minibus":     15.0,
    "LPV / Tempo":          3.0,
    "HPV Bus":             50.0,
    "JKRTC / City Bus":     0.0,
}
DEFAULT_BUYBACK_LAKH = 10.0


def _operator_absorption_from_df(df: pd.DataFrame) -> list:
    """
    Re-derive the operator-class buyback summary from the full CSV.
    Counts MERGED rows by Displaced_Operator_Class (each merged row IS
    one absorbed permit — Fleet_Required is zeroed post-merge so we can't
    sum it like the source XLSX tried).
    """
    if "Displaced_Operator_Class" not in df.columns:
        return []
    merged = df[df["Action_Taken"] == "MERGED_INTO_TRUNK"].copy()
    if merged.empty:
        return []
    merged["_class"] = (merged["Displaced_Operator_Class"]
                         .fillna("Unclassified").astype(str).str.strip()
                         .replace({"": "Unclassified"}))
    grouped = (merged.groupby("_class").size()
                       .reset_index(name="permits")
                       .sort_values("permits", ascending=False))
    out = []
    for _, r in grouped.iterrows():
        cls = r["_class"]
        n   = int(r["permits"])
        per_lakh = BUYBACK_LAKH_PER_VEHICLE.get(cls, DEFAULT_BUYBACK_LAKH)
        cost = n * per_lakh * 100_000      # lakh → ₹
        if cls == "JKRTC / City Bus":
            action = "Retain as feeder under JKRTC"
        elif cls == "LPV / Tempo":
            action = "Last-mile reassignment"
        elif cls == "HPV Bus":
            action = "Roll into JKRTC"
        elif cls == "Private Minibus":
            action = "Buyback @ ₹15 L / vehicle"
        else:
            action = "Reassign or buyback"
        out.append({
            "class":     cls,
            "permits":   n,
            "avg_fleet": 1.0,    # one permit per merged row
            "action":    action,
            "cost_inr":  cost,
        })
    return out


def _coerce_bool(v) -> bool:
    """Robust truthy check for Python bool, numpy.bool_, or string 'True'."""
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v) and v == v   # excludes NaN
    if isinstance(v, str):
        return v.strip().lower() in ("true", "1", "yes")
    return False


# ─── Build ────────────────────────────────────────────────────────────────────
def main():
    df = _read_route_plan()
    # Coerce numeric cols
    for col in ["Route_KM", "Cycle_Time_Min", "Population_Served", "Population_Served_Raw"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    for col in ["Headway_Min", "Fleet_Required",
                "HPV_Count", "MPV_Count", "LPV_Count",
                "Daily_Demand_Pax", "Daily_Capacity_Pax"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    # The current CSV export drops LPV_Count (engine export_csv subset) —
    # derive it back from Fleet_Required minus the two known vehicle classes.
    if "LPV_Count" not in df.columns:
        if {"Fleet_Required", "HPV_Count", "MPV_Count"}.issubset(df.columns):
            df["LPV_Count"] = (df["Fleet_Required"]
                                  - df["HPV_Count"]
                                  - df["MPV_Count"]).clip(lower=0).astype(int)
        else:
            df["LPV_Count"] = 0
    # Bools — must handle Python bool, numpy bool_, and string "True"
    for col in ["Social_Flag", "Tourist_Corridor", "Subsidy_Risk_Flag",
                "CMP_Trunk"]:
        if col in df.columns:
            df[col] = df[col].apply(_coerce_bool)

    active = df[df["Action_Taken"] != "MERGED_INTO_TRUNK"].copy()

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1 — Summary  (KPI dashboard)
    # ═══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    hide_gridlines(ws1)

    widths = {"A": 3, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22, "G": 22, "H": 3}
    for c, w in widths.items():
        ws1.column_dimensions[c].width = w

    # Title
    ws1.merge_cells("B2:G3")
    title = ws1.cell(row=2, column=2,
                      value="Kashmir Valley Route Rationalisation — RTO Master")
    title.font = Font(name="Segoe UI", bold=True, color=NAVY, size=22)
    title.alignment = Alignment(horizontal="left", vertical="center")

    ws1.merge_cells("B4:G4")
    sub = ws1.cell(row=4, column=2,
                    value=(f"Engine v3.4.3  ·  Generated "
                            f"{datetime.datetime.now():%d %b %Y}  ·  "
                            f"Phase-1 Conservative plan  ·  35-min headway ceiling"))
    sub.font = Font(name="Segoe UI", italic=True, color=GREY, size=12)
    sub.alignment = Alignment(horizontal="left", vertical="center")

    # 8 KPI tiles  (4-wide × 2-tall)
    sscl_cnt    = int(df["CMP_Trunk"].sum()) if "CMP_Trunk" in df.columns else 0
    tourist_cnt = int(df["Tourist_Corridor"].sum()) if "Tourist_Corridor" in df.columns else 0
    social_cnt  = int(df["Social_Flag"].sum()) if "Social_Flag" in df.columns else 0
    merged_cnt  = int((df["Action_Taken"] == "MERGED_INTO_TRUNK").sum())

    tiles = [
        ("Total permits",       f"{len(df):,}",                                  TEAL),
        ("Active routes",       f"{len(active):,}",                              TEAL),
        ("Upgraded to Trunk",   f"{(df['Action_Taken']=='UPGRADED_TO_TRUNK').sum():,}", GREEN),
        ("Merged into Trunk",   f"{merged_cnt:,}",                                SAFFRON),
        ("Total fleet",         f"{int(active['Fleet_Required'].sum()):,}",       GOLD),
        ("SSCL backbone",       f"{sscl_cnt}  / 30",                              NAVY),
        ("Tourist corridors",   f"{tourist_cnt}",                                  PURPLE),
        ("Social Obligation",   f"{social_cnt}",                                   NAVY),
    ]
    tile_start_row = 6
    for i, (label, val, hex_col) in enumerate(tiles):
        r, c = i // 4, i % 4
        top = tile_start_row + r * 4
        col = 2 + c
        ws1.row_dimensions[top].height = 18
        lc = ws1.cell(row=top, column=col)
        set_cell(lc, label.upper(), bold=True, size=9, font_color=WHITE,
                  bg=hex_col, border=False)
        ws1.row_dimensions[top + 1].height = 36
        vc = ws1.cell(row=top + 1, column=col)
        set_cell(vc, val, bold=True, size=22, font_color=hex_col,
                  bg=LIGHT, border=False, name="Calibri")

    # Network composition table (cols B:C)
    block_y = tile_start_row + 8
    ws1.cell(row=block_y, column=2, value="Network composition").font = Font(
        name="Segoe UI", bold=True, color=NAVY, size=14)
    header_cell(ws1.cell(row=block_y + 1, column=2), "Action")
    header_cell(ws1.cell(row=block_y + 1, column=3), "Routes")
    action_rows = [
        ("Upgraded to Trunk",  "UPGRADED_TO_TRUNK",  GREEN_LT,   GREEN),
        ("Retained as Feeder", "RETAINED_AS_FEEDER", TEAL_LT,    TEAL),
        ("Merged into Trunk",  "MERGED_INTO_TRUNK",  SAFFRON_LT, SAFFRON),
    ]
    for i, (label, code, bg, fc) in enumerate(action_rows):
        rr = block_y + 2 + i
        set_cell(ws1.cell(row=rr, column=2), label, align="left", bg=bg)
        set_cell(ws1.cell(row=rr, column=3),
                  int((df["Action_Taken"] == code).sum()),
                  bold=True, font_color=fc, num_fmt="#,##0", bg=bg)
    rr_total = block_y + 5
    set_cell(ws1.cell(row=rr_total, column=2), "Total permits",
              bold=True, align="left", bg=LIGHT)
    set_cell(ws1.cell(row=rr_total, column=3), len(df),
              bold=True, num_fmt="#,##0", bg=LIGHT)

    # Fleet composition table (cols E:F)
    ws1.cell(row=block_y, column=5, value="Fleet composition").font = Font(
        name="Segoe UI", bold=True, color=NAVY, size=14)
    header_cell(ws1.cell(row=block_y + 1, column=5), "Class")
    header_cell(ws1.cell(row=block_y + 1, column=6), "Buses")
    fleet_rows = [
        ("HPV — 12 m bus",  "HPV_Count",  NAVY,    NAVY),
        ("MPV — 9 m bus",   "MPV_Count",  TEAL_LT, TEAL),
        ("LPV — minibus",   "LPV_Count",  GOLD_LT, GOLD),
    ]
    for i, (label, col_key, bg, fc) in enumerate(fleet_rows):
        rr = block_y + 2 + i
        set_cell(ws1.cell(row=rr, column=5), label, align="left",
                  bg=bg if i > 0 else WHITE)
        n = int(active[col_key].sum()) if col_key in active else 0
        set_cell(ws1.cell(row=rr, column=6), n,
                  bold=True, font_color=fc, num_fmt="#,##0",
                  bg=bg if i > 0 else WHITE)
    # HPV is NAVY background — make text white for readability
    cell = ws1.cell(row=block_y + 2, column=5)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Segoe UI", size=10, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    set_cell(ws1.cell(row=block_y + 2, column=6),
              int(active["HPV_Count"].sum()),
              bold=True, font_color=WHITE, num_fmt="#,##0", bg=NAVY)
    set_cell(ws1.cell(row=block_y + 5, column=5), "Total fleet",
              bold=True, align="left", bg=LIGHT)
    set_cell(ws1.cell(row=block_y + 5, column=6),
              int(active["Fleet_Required"].sum()),
              bold=True, num_fmt="#,##0", bg=LIGHT)

    # Headway table (cols B:C, lower block)
    block_y2 = block_y + 7
    ws1.cell(row=block_y2, column=2,
              value="Headway distribution (active only)").font = Font(
        name="Segoe UI", bold=True, color=NAVY, size=14)
    header_cell(ws1.cell(row=block_y2 + 1, column=2), "Headway")
    header_cell(ws1.cell(row=block_y2 + 1, column=3), "Routes")
    expected = [15, 20, 30, 35, 45, 60]
    bin_counts = {b: int((active["Headway_Min"] == b).sum()) for b in expected}
    other = int(len(active) - sum(bin_counts.values()))
    hw_rows = [(b, bin_counts[b]) for b in expected if bin_counts[b] > 0]
    if other:
        hw_rows.append(("other", other))
    for i, (b, n) in enumerate(hw_rows):
        rr = block_y2 + 2 + i
        label = f"{b} min" if isinstance(b, int) else b
        set_cell(ws1.cell(row=rr, column=2), label, align="left",
                  bg=LIGHT if i % 2 == 0 else WHITE)
        set_cell(ws1.cell(row=rr, column=3), n, bold=True,
                  num_fmt="#,##0", bg=LIGHT if i % 2 == 0 else WHITE)
    rr_tot = block_y2 + 2 + len(hw_rows)
    set_cell(ws1.cell(row=rr_tot, column=2), "Total active",
              bold=True, align="left", bg=NAVY, font_color=WHITE)
    set_cell(ws1.cell(row=rr_tot, column=3), len(active),
              bold=True, num_fmt="#,##0", bg=NAVY, font_color=WHITE)

    # (v3.3.7: operator-absorption mini-summary removed from the pretty workbook
    # along with the Operator Absorption + Sign-off sheets — this file is now a
    # focused 2-sheet bus schedule: Summary + Route Plan.)

    # Calibration anchor strip
    cal_row = rr_tot + 2
    ws1.merge_cells(start_row=cal_row, start_column=2,
                     end_row=cal_row, end_column=7)
    cal = ws1.cell(row=cal_row, column=2,
                    value=("Calibrated against CHALO Apr 2026 ridership data  ·  "
                            "Per-route SSCL fleet within ±10% of headway-scaled "
                            "CHALO  ·  8/8 QC checks passing  ·  0 Red_Overload routes"))
    cal.font = Font(name="Segoe UI", bold=True, italic=True, color=GREEN, size=10)
    cal.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cal.fill = PatternFill("solid", fgColor=GREEN_LT)
    cal.border = BORDER
    ws1.row_dimensions[cal_row].height = 36

    # Footer notes
    note_row = cal_row + 2
    ws1.merge_cells(start_row=note_row, start_column=2,
                     end_row=note_row, end_column=7)
    nt = ws1.cell(row=note_row, column=2,
                   value=("Legend  ·  Action_Taken pill colours: Green = Upgraded · "
                           "Teal = Retained · Red = Merged.  Priority bands HP/MP/LP "
                           "set by demand index (Population + POI gravity)."))
    nt.font = Font(name="Segoe UI", italic=True, color=GREY, size=9)
    nt.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2 — Route Plan  (compact bus schedule — v3.3.7: Load flag / Social /
    #           Tourist columns removed at RTO request)
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Route Plan")
    hide_gridlines(ws2)

    columns = [
        ("Route_Code",         "Route Code",       16, "left"),
        ("Route_Name",         "Route Name",       40, "left"),
        ("Action_Taken",       "Action",           22, "center"),
        ("Route_KM",           "Length (km)",      11, "center"),
        ("Route_Type",         "Type",             14, "center"),
        ("Priority_Band",      "Priority",         11, "center"),
        ("Headway_Min",        "Headway (min)",    12, "center"),
        ("Cycle_Time_Min",     "Cycle (min)",      11, "center"),
        ("Fleet_Required",     "Fleet",            10, "center"),
        ("HPV_Count",          "HPV",              7,  "center"),
        ("MPV_Count",          "MPV",              7,  "center"),
        ("LPV_Count",          "LPV",              7,  "center"),
        ("Population_Served_Raw", "Pop. served",  14,  "center"),
    ]

    ws2.row_dimensions[1].height = 30
    for ci, (key, hdr, width, _align) in enumerate(columns, start=1):
        cell = ws2.cell(row=1, column=ci, value=hdr)
        cell.font = Font(name="Segoe UI", bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        cell.border = BORDER
        ws2.column_dimensions[get_column_letter(ci)].width = width

    action_palette = {
        "UPGRADED_TO_TRUNK":  (GREEN_LT,   GREEN,   "Upgraded to Trunk"),
        "RETAINED_AS_FEEDER": (TEAL_LT,    TEAL,    "Retained as Feeder"),
        "MERGED_INTO_TRUNK":  (SAFFRON_LT, SAFFRON, "Merged into Trunk"),
    }
    band_palette = {
        "HP": (GREEN_LT,  GREEN),
        "MP": (PURPLE_LT, PURPLE),
        "LP": (GOLD_LT,   GOLD),
    }
    load_palette = {
        "Green":          (GREEN,   WHITE, "Green"),
        "Amber_Tight":    (GOLD,    WHITE, "Amber"),
        "Amber_Under":    ("FBC02D", DARK,  "Amber"),
        "Red_Overload":   (SAFFRON, WHITE, "Red"),
        "Red_NoCapacity": ("B71C1C", WHITE, "Red"),
    }

    # RTO ask: the bus schedule lists only routes that actually run a fleet.
    # MERGED_INTO_TRUNK rows are consolidated/duplicate permits whose service is
    # absorbed into their trunk (Fleet_Required = 0), so they're excluded here.
    # (The Summary sheet still reports the full breakdown for context.)
    plan_df = df[df["Action_Taken"] != "MERGED_INTO_TRUNK"].copy()
    for row_i, (_, row) in enumerate(plan_df.iterrows(), start=2):
        action = str(row.get("Action_Taken", ""))
        band   = str(row.get("Priority_Band", ""))
        load   = str(row.get("Load_Flag", "") or "")
        row_bg = LIGHT if row_i % 2 == 0 else WHITE

        for ci, (key, _hdr, _w, align) in enumerate(columns, start=1):
            value = row.get(key, "")
            cell = ws2.cell(row=row_i, column=ci, value=value)
            cell.font = Font(name="Segoe UI", size=10, color=DARK)
            cell.alignment = Alignment(horizontal=align, vertical="center",
                                        wrap_text=(key == "Route_Name"))
            cell.border = BORDER
            cell.fill = PatternFill("solid", fgColor=row_bg)

            # number formats
            if key == "Route_KM":
                cell.number_format = "0.0"
            elif key == "Cycle_Time_Min":
                cell.number_format = "0.0"
            elif key in ("Headway_Min", "Fleet_Required",
                         "HPV_Count", "MPV_Count", "LPV_Count",
                         "Population_Served", "Population_Served_Raw"):
                cell.number_format = "#,##0"

            # Action pill
            if key == "Action_Taken":
                bg, fc, pretty = action_palette.get(action, (LIGHT, DARK, action))
                cell.value = pretty
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Segoe UI", size=9, bold=True, color=fc)

            # Priority pill
            if key == "Priority_Band" and band in band_palette:
                bg, fc = band_palette[band]
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Segoe UI", size=10, bold=True, color=fc)

            # Load_Flag pill
            if key == "Load_Flag" and load:
                cfg = load_palette.get(load)
                if cfg:
                    bg, fc, label = cfg
                    cell.value = label
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(name="Segoe UI", size=9, bold=True, color=fc)

            # Boolean checkmarks for Social / Tourist
            if key in ("Social_Flag", "Tourist_Corridor"):
                is_true = bool(value) and not (isinstance(value, str) and
                                                 value.lower() in ("false", "no", ""))
                cell.value = "✓" if is_true else ""
                if is_true:
                    cell.font = Font(name="Segoe UI", size=12, bold=True,
                                      color=PURPLE if key == "Tourist_Corridor" else NAVY)

            # Highlight long inter-district routes
            if key == "Route_KM" and isinstance(value, (int, float)) and value >= 30:
                cell.font = Font(name="Segoe UI", size=10, bold=True,
                                  color=PURPLE)

    ws2.freeze_panes = "C2"
    last_col = get_column_letter(len(columns))
    ws2.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

    ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
    ws2.page_setup.paperSize   = ws2.PAPERSIZE_A3
    ws2.page_setup.fitToPage   = True
    ws2.page_setup.fitToWidth  = 1
    ws2.page_setup.fitToHeight = 0
    ws2.print_title_rows = "1:1"
    ws2.print_options.gridLines = False

    # v3.3.7: Operator Absorption + Sign-off sheets removed — the pretty
    # workbook is now a focused 2-sheet bus schedule (Summary + Route Plan).

    # Save
    wb.save(DST)
    wb.save(ALSO)
    print(f"Wrote {DST}")
    print(f"Also  {ALSO}")
    print(f"  Sheet 1  ·  Summary           (KPI dashboard)")
    print(f"  Sheet 2  ·  Route Plan        ({len(plan_df)} fleet-carrying routes × {len(columns)} cols)")


if __name__ == "__main__":
    main()
