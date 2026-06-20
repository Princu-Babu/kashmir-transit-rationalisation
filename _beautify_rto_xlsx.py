"""
_beautify_rto_xlsx.py

Quick visual polish for Formatted_Kashmir_Routes.xlsx:
  - Sheet 1  Summary  (8 KPI tiles + 4 quick tables on a single dashboard)
  - Sheet 2  Route-Level Plan  (the original 12 cols, restyled, frozen panes,
             auto-filter, colour-coded Action_Taken and Priority_Band pills)
  - Sheet 3  Quick filters  (top-fleet routes, longest, LPV-only — for an RTO
             reviewer skimming for hot spots)

Output: Formatted_Kashmir_Routes_Pretty.xlsx next to the source.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SRC = Path(r"C:/Users/Prash/Music/Formatted_Kashmir_Routes.xlsx")
DST = Path(r"C:/Users/Prash/Music/Formatted_Kashmir_Routes_Pretty.xlsx")


# ─── Theme ────────────────────────────────────────────────────────────────────
NAVY        = "1A237E"
TEAL        = "00695C"
TEAL_LT     = "E0F2F1"
PURPLE      = "6A1B9A"
PURPLE_LT   = "F3E5F5"
SAFFRON     = "D32F2F"
SAFFRON_LT  = "FFEBEE"
GOLD        = "F9A825"
GOLD_LT     = "FFF8E1"
GREEN       = "2E7D32"
GREEN_LT    = "E8F5E9"
LIGHT       = "F5F5F5"
DARK        = "212121"
GREY        = "757575"
WHITE       = "FFFFFF"
BORDER_GREY = "E0E0E0"

THIN = Side(style="thin", color=BORDER_GREY)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ─── Helpers ──────────────────────────────────────────────────────────────────
def set_cell(cell, value, *, bold=False, size=10, font_color=DARK,
             align="center", anchor="center", bg=None, num_fmt=None,
             border=True, italic=False, name="Segoe UI"):
    cell.value = value
    cell.font = Font(name=name, size=size, bold=bold, italic=italic,
                     color=font_color)
    cell.alignment = Alignment(horizontal=align, vertical=anchor,
                                wrap_text=True)
    if border:
        cell.border = BORDER
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt


def header_cell(cell, value, *, bg=NAVY, size=11):
    set_cell(cell, value, bold=True, size=size, font_color=WHITE,
             align="center", anchor="center", bg=bg)


def hide_gridlines(ws):
    ws.sheet_view.showGridLines = False


# ─── Build ────────────────────────────────────────────────────────────────────
def main():
    df = pd.read_excel(SRC, sheet_name=0)
    # Defensive coercions
    for col in ["Route_KM", "Cycle_Time_Min"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    for col in ["Headway_Min", "Fleet_Required",
                "HPV_Count", "MPV_Count", "LPV_Count"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    active = df[df["Action_Taken"] != "MERGED_INTO_TRUNK"].copy()

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1 — Summary
    # ═══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"
    hide_gridlines(ws1)

    # column widths
    widths = {"A": 3, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22, "G": 22, "H": 3}
    for c, w in widths.items():
        ws1.column_dimensions[c].width = w

    # Title block
    ws1.merge_cells("B2:G3")
    title = ws1.cell(row=2, column=2, value="Kashmir Valley Route Rationalisation")
    title.font = Font(name="Segoe UI", bold=True, color=NAVY, size=22)
    title.alignment = Alignment(horizontal="left", vertical="center")

    ws1.merge_cells("B4:G4")
    sub = ws1.cell(row=4, column=2,
                    value="At-a-glance summary  ·  342 in-scope permits  ·  Kashmir Valley study area")
    sub.font = Font(name="Segoe UI", italic=True, color=GREY, size=12)
    sub.alignment = Alignment(horizontal="left", vertical="center")

    # KPI tiles — 8 tiles in a 4×2 grid
    tiles = [
        ("Total permits",      f"{len(df):,}",                                  TEAL),
        ("Active routes",      f"{len(active):,}",                              TEAL),
        ("Upgraded to Trunk",  f"{(df['Action_Taken']=='UPGRADED_TO_TRUNK').sum():,}", GREEN),
        ("Merged into Trunk",  f"{(df['Action_Taken']=='MERGED_INTO_TRUNK').sum():,}", SAFFRON),
        ("Total fleet",        f"{int(active['Fleet_Required'].sum()):,}",      GOLD),
        ("HPV (12 m)",         f"{int(active['HPV_Count'].sum()):,}",           NAVY),
        ("MPV (9 m)",          f"{int(active['MPV_Count'].sum()):,}",           NAVY),
        ("LPV (small)",        f"{int(active['LPV_Count'].sum()):,}",           NAVY),
    ]
    tile_start_row = 6
    for i, (label, val, col_hex) in enumerate(tiles):
        r, c = i // 4, i % 4
        top = tile_start_row + r * 4
        col = 2 + c       # cols B (2) → E (5)
        # label row
        ws1.row_dimensions[top].height = 18
        lc = ws1.cell(row=top, column=col)
        set_cell(lc, label.upper(), bold=True, size=9, font_color=WHITE,
                 bg=col_hex, border=False)
        # value row
        ws1.row_dimensions[top + 1].height = 36
        vc = ws1.cell(row=top + 1, column=col)
        set_cell(vc, val, bold=True, size=22, font_color=col_hex,
                 bg=LIGHT, border=False, name="Calibri")

    # Network composition table (Sheet1 cols B:C, rows 14+)
    network_y = tile_start_row + 8
    ws1.cell(row=network_y, column=2,
              value="Network composition").font = Font(
                  name="Segoe UI", bold=True, color=NAVY, size=14)
    nh_row = network_y + 1
    header_cell(ws1.cell(row=nh_row, column=2), "Action")
    header_cell(ws1.cell(row=nh_row, column=3), "Routes")
    action_rows = [
        ("Upgraded to Trunk",  "UPGRADED_TO_TRUNK",  GREEN_LT,  GREEN),
        ("Retained as Feeder", "RETAINED_AS_FEEDER", TEAL_LT,   TEAL),
        ("Merged into Trunk",  "MERGED_INTO_TRUNK",  SAFFRON_LT, SAFFRON),
    ]
    for i, (label, code, bg, fc) in enumerate(action_rows):
        n = int((df["Action_Taken"] == code).sum())
        rr = nh_row + 1 + i
        set_cell(ws1.cell(row=rr, column=2), label, align="left", bg=bg)
        set_cell(ws1.cell(row=rr, column=3), n, bold=True, font_color=fc,
                 num_fmt="#,##0", bg=bg)
    set_cell(ws1.cell(row=nh_row + 4, column=2),
             "Total permits", bold=True, align="left", bg=LIGHT)
    set_cell(ws1.cell(row=nh_row + 4, column=3),
             len(df), bold=True, num_fmt="#,##0", bg=LIGHT)

    # Priority band table (cols E:F)
    ws1.cell(row=network_y, column=5,
              value="By priority band").font = Font(
                  name="Segoe UI", bold=True, color=NAVY, size=14)
    header_cell(ws1.cell(row=nh_row, column=5), "Band")
    header_cell(ws1.cell(row=nh_row, column=6), "Routes")
    band_rows = [
        ("HP — High Priority",   "HP", GREEN_LT,  GREEN),
        ("MP — Medium Priority", "MP", PURPLE_LT, PURPLE),
        ("LP — Low Priority",    "LP", GOLD_LT,   GOLD),
    ]
    for i, (label, code, bg, fc) in enumerate(band_rows):
        n = int((df["Priority_Band"] == code).sum())
        rr = nh_row + 1 + i
        set_cell(ws1.cell(row=rr, column=5), label, align="left", bg=bg)
        set_cell(ws1.cell(row=rr, column=6), n, bold=True, font_color=fc,
                 num_fmt="#,##0", bg=bg)

    # Headway distribution table (cols B:C, lower block)
    headway_y = nh_row + 6
    ws1.cell(row=headway_y, column=2,
              value="Headway distribution (active only)").font = Font(
                  name="Segoe UI", bold=True, color=NAVY, size=14)
    hd_row = headway_y + 1
    header_cell(ws1.cell(row=hd_row, column=2), "Headway")
    header_cell(ws1.cell(row=hd_row, column=3), "Routes")
    expected_bins = [15, 20, 30, 35, 45, 60]
    bin_counts = {b: int((active["Headway_Min"] == b).sum()) for b in expected_bins}
    other = int(len(active) - sum(bin_counts.values()))
    rows_hw = [(b, bin_counts[b]) for b in expected_bins if bin_counts[b] > 0]
    if other:
        rows_hw.append(("other", other))
    for i, (b, n) in enumerate(rows_hw):
        rr = hd_row + 1 + i
        label = f"{b} min" if isinstance(b, int) else b
        set_cell(ws1.cell(row=rr, column=2), label, align="left",
                 bg=LIGHT if i % 2 == 0 else WHITE)
        set_cell(ws1.cell(row=rr, column=3), n, bold=True, num_fmt="#,##0",
                 bg=LIGHT if i % 2 == 0 else WHITE)
    total_hr = hd_row + 1 + len(rows_hw)
    set_cell(ws1.cell(row=total_hr, column=2), "Total active",
             bold=True, align="left", bg=NAVY, font_color=WHITE)
    set_cell(ws1.cell(row=total_hr, column=3), len(active),
             bold=True, num_fmt="#,##0", bg=NAVY, font_color=WHITE)

    # Route type table (cols E:F, lower block)
    ws1.cell(row=headway_y, column=5,
              value="Route type").font = Font(
                  name="Segoe UI", bold=True, color=NAVY, size=14)
    header_cell(ws1.cell(row=hd_row, column=5), "Type")
    header_cell(ws1.cell(row=hd_row, column=6), "Routes")
    rt_rows = [
        ("Urban", "Urban", TEAL_LT, TEAL),
        ("Peri-urban", "Peri_Urban", GOLD_LT, GOLD),
        ("Regional / District", "Regional_District", PURPLE_LT, PURPLE),
    ]
    for i, (label, code, bg, fc) in enumerate(rt_rows):
        n = int((df["Route_Type"] == code).sum())
        rr = hd_row + 1 + i
        set_cell(ws1.cell(row=rr, column=5), label, align="left", bg=bg)
        set_cell(ws1.cell(row=rr, column=6), n, bold=True, num_fmt="#,##0",
                 font_color=fc, bg=bg)

    # Footer notes
    foot_y = max(total_hr, hd_row + 1 + len(rt_rows)) + 2
    foot1 = ws1.cell(row=foot_y, column=2,
                      value="Action_Taken legend  ·  "
                            "Green = Upgraded to Trunk  ·  "
                            "Teal = Retained as Feeder  ·  "
                            "Red = Merged into Trunk  ·  "
                            "Priority bands: HP/MP/LP set by demand index (population + POI)")
    foot1.font = Font(name="Segoe UI", italic=True, size=9, color=GREY)
    foot1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws1.merge_cells(start_row=foot_y, start_column=2, end_row=foot_y, end_column=7)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2 — Route-Level Plan (styled)
    # ═══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Route-Level Plan")
    hide_gridlines(ws2)

    columns = [
        ("Route_Code",     "Route Code",      16, "left"),
        ("Route_Name",     "Route Name",      40, "left"),
        ("Action_Taken",   "Action",          22, "center"),
        ("Route_KM",       "Length (km)",     12, "center"),
        ("Route_Type",     "Route Type",      16, "center"),
        ("Priority_Band",  "Priority",        12, "center"),
        ("Headway_Min",    "Headway (min)",   13, "center"),
        ("Cycle_Time_Min", "Cycle (min)",     12, "center"),
        ("Fleet_Required", "Fleet",           10, "center"),
        ("HPV_Count",      "HPV",             8,  "center"),
        ("MPV_Count",      "MPV",             8,  "center"),
        ("LPV_Count",      "LPV",             8,  "center"),
    ]

    # Header row
    ws2.row_dimensions[1].height = 28
    for ci, (key, hdr, width, _align) in enumerate(columns, start=1):
        cell = ws2.cell(row=1, column=ci, value=hdr)
        cell.font = Font(name="Segoe UI", bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        cell.border = BORDER
        ws2.column_dimensions[get_column_letter(ci)].width = width

    action_palette = {
        "UPGRADED_TO_TRUNK":  (GREEN_LT,   GREEN,   "Upgraded to Trunk"),
        "RETAINED_AS_FEEDER": (TEAL_LT,    TEAL,    "Retained as Feeder"),
        "MERGED_INTO_TRUNK":  (SAFFRON_LT, SAFFRON, "Merged into Trunk"),
    }
    band_palette = {
        "HP": (GREEN_LT,  GREEN),
        "MP": (PURPLE_LT, PURPLE),
        "LP": (GOLD_LT,   GOLD),
    }

    for row_i, (_, row) in enumerate(df.iterrows(), start=2):
        action      = str(row.get("Action_Taken", ""))
        band        = str(row.get("Priority_Band", ""))
        # alternating row tint
        row_bg = LIGHT if row_i % 2 == 0 else WHITE
        if action == "MERGED_INTO_TRUNK":
            row_bg = SAFFRON_LT

        for ci, (key, _hdr, _w, align) in enumerate(columns, start=1):
            value = row.get(key, "")
            cell = ws2.cell(row=row_i, column=ci, value=value)
            cell.font = Font(name="Segoe UI", size=10, color=DARK)
            cell.alignment = Alignment(horizontal=align, vertical="center",
                                        wrap_text=(key == "Route_Name"))
            cell.border = BORDER
            cell.fill = PatternFill("solid", fgColor=row_bg)

            # Number formats
            if key == "Route_KM":
                cell.number_format = "0.0"
            elif key in ("Cycle_Time_Min",):
                cell.number_format = "0.0"
            elif key in ("Headway_Min", "Fleet_Required",
                         "HPV_Count", "MPV_Count", "LPV_Count"):
                cell.number_format = "#,##0"

            # Action_Taken pill
            if key == "Action_Taken":
                bg, fc, pretty = action_palette.get(action, (LIGHT, DARK, action))
                cell.value = pretty
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Segoe UI", size=9, bold=True, color=fc)

            # Priority_Band pill
            if key == "Priority_Band" and band in band_palette:
                bg, fc = band_palette[band]
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Segoe UI", size=10, bold=True, color=fc)

            # Length cell — value bigger on long inter-district routes
            if key == "Route_KM" and isinstance(value, (int, float)) and value >= 30:
                cell.font = Font(name="Segoe UI", size=10, bold=True,
                                  color=PURPLE)

    # Freeze panes — keep Route Code + Route Name visible while scrolling
    ws2.freeze_panes = "C2"

    # Auto-filter on the whole data range
    last_col = get_column_letter(len(columns))
    ws2.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

    # Print setup
    ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
    ws2.page_setup.paperSize   = ws2.PAPERSIZE_A3
    ws2.page_setup.fitToPage   = True
    ws2.page_setup.fitToWidth  = 1
    ws2.page_setup.fitToHeight = 0
    ws2.print_title_rows = "1:1"
    ws2.print_options.gridLines = False

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 3 — Quick filters (top-fleet, longest, LPV-only)
    # ═══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Quick filters")
    hide_gridlines(ws3)
    for col, w in zip("ABCDEFGH", [3, 16, 38, 18, 12, 12, 12, 3]):
        ws3.column_dimensions[col].width = w

    def write_table(start_row, title, frame, color_bg=NAVY):
        ws3.merge_cells(start_row=start_row, start_column=2,
                         end_row=start_row, end_column=7)
        t = ws3.cell(row=start_row, column=2, value=title)
        t.font = Font(name="Segoe UI", bold=True, color=color_bg, size=14)
        t.alignment = Alignment(horizontal="left", vertical="center")

        headers = ["Route Code", "Route Name", "Action",
                    "Length", "Headway", "Fleet"]
        for ci, h in enumerate(headers, start=2):
            header_cell(ws3.cell(row=start_row + 1, column=ci), h, bg=color_bg)

        for ri, (_, r) in enumerate(frame.iterrows(), start=start_row + 2):
            row_bg = LIGHT if ri % 2 == 0 else WHITE
            set_cell(ws3.cell(row=ri, column=2), r["Route_Code"],
                     align="left", bg=row_bg)
            set_cell(ws3.cell(row=ri, column=3), r["Route_Name"],
                     align="left", bg=row_bg)
            action = str(r["Action_Taken"])
            bg, fc, pretty = action_palette.get(action, (row_bg, DARK, action))
            set_cell(ws3.cell(row=ri, column=4), pretty,
                     bold=True, font_color=fc, bg=bg, size=9)
            set_cell(ws3.cell(row=ri, column=5), float(r["Route_KM"]),
                     num_fmt="0.0", bg=row_bg)
            set_cell(ws3.cell(row=ri, column=6), int(r["Headway_Min"]),
                     bg=row_bg)
            set_cell(ws3.cell(row=ri, column=7), int(r["Fleet_Required"]),
                     bold=True, bg=row_bg)

    # Top 15 fleet-heavy routes
    top_fleet = (active.sort_values("Fleet_Required", ascending=False)
                       .head(15))
    write_table(2, "Top 15 routes by fleet requirement",
                top_fleet, color_bg=SAFFRON)

    # Top 10 longest routes
    longest = (active.sort_values("Route_KM", ascending=False).head(10))
    write_table(2 + 2 + len(top_fleet) + 2,
                "Top 10 longest routes (inter-district lifelines)",
                longest, color_bg=PURPLE)

    # LPV-only routes (narrow-lane operation)
    lpv_only = active[(active["LPV_Count"] > 0) &
                       (active["HPV_Count"] == 0) &
                       (active["MPV_Count"] == 0)].head(15)
    write_table(2 + 2 + len(top_fleet) + 2 + 2 + len(longest) + 2,
                f"LPV-only routes ({len(active[(active['LPV_Count']>0)&(active['HPV_Count']==0)&(active['MPV_Count']==0)])} total — narrow-lane operation)",
                lpv_only, color_bg=TEAL)

    # Save
    wb.save(DST)
    print(f"Wrote {DST}")
    print(f"  Sheet 1  ·  Summary")
    print(f"  Sheet 2  ·  Route-Level Plan  ({len(df)} rows × {len(columns)} cols)")
    print(f"  Sheet 3  ·  Quick filters")


if __name__ == "__main__":
    main()
