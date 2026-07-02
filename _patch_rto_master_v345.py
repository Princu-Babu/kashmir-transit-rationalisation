#!/usr/bin/env python
"""
Patch the copied 9-sheet RTO master to v3.4.5 (measured-cycle corrections).

Updates ONLY: the 5 corrected route rows (cycle/fleet/mix), the fleet totals,
the cover narrative/version, and the calibration version line. Everything else
stays byte-identical. Verified after by re-summing the Route-Level Plan.
"""
import csv, sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from openpyxl import load_workbook

WB = "outputs_v3.4.5/Kashmir_Route_Frequency_Plan_v3.4.5_RTO.xlsx"
CSV = "outputs_v3.4.5/Rationalised_Routes_Kashmir_v3.csv"

rows = list(csv.DictReader(open(CSV, encoding="utf-8-sig")))
act = [r for r in rows if float(r["Fleet_Required"] or 0) > 0]
tot = sum(int(float(r["Fleet_Required"])) for r in act)
hp = sum(int(float(r["HPV_Count"])) for r in act)
mp = sum(int(float(r["MPV_Count"])) for r in act)
lp = sum(int(float(r["LPV_Count"] or 0)) for r in act)
changed = {r["Route_Code"]: r for r in csv.DictReader(open("corrections_applied_v345.csv", encoding="utf-8-sig"))
           if r["Old_Fleet"] != r["New_Fleet"] or r["Old_Cycle"] != r["New_Cycle"]}
csv_by_code_active = {}
for r in act:
    csv_by_code_active.setdefault(r["Route_Code"], r)
print(f"targets: {len(changed)} changed routes; new totals {tot} ({hp}/{mp}/{lp})")

wb = load_workbook(WB)

# ── Route-Level Plan: the changed active rows ──────────────────────
ws = wb["Route-Level Plan"]
hdr = {c.value: i + 1 for i, c in enumerate(ws[1])}
n_patched = 0
for row in ws.iter_rows(min_row=2):
    code = row[hdr["Route_Code"] - 1].value
    fleet_cell = row[hdr["Fleet_Required"] - 1]
    if code in changed and float(fleet_cell.value or 0) > 0:
        src = csv_by_code_active[code]
        row[hdr["Cycle_Time_Min"] - 1].value = float(src["Cycle_Time_Min"])
        fleet_cell.value = int(float(src["Fleet_Required"]))
        row[hdr["HPV_Count"] - 1].value = int(float(src["HPV_Count"]))
        row[hdr["MPV_Count"] - 1].value = int(float(src["MPV_Count"]))
        row[hdr["LPV_Count"] - 1].value = int(float(src["LPV_Count"]))
        n_patched += 1
print(f"Route-Level Plan: patched {n_patched} rows")

# ── Network Summary: fleet composition ─────────────────────────────
ws = wb["Network Summary"]
for row in ws.iter_rows():
    for cell in row:
        if cell.value == "HPV (12m Bus)": ws.cell(cell.row, cell.column + 1).value = hp
        if cell.value == "MPV (9m Bus)": ws.cell(cell.row, cell.column + 1).value = mp
        if cell.value == "LPV (Minibus/Tempo)": ws.cell(cell.row, cell.column + 1).value = lp

# ── Cover: version line, narrative, highlight total ────────────────
ws = wb["Cover & Sign-off"]
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if not isinstance(v, str): continue
        if v.startswith("Version:"):
            cell.value = "Version: v3.4.5  |  Generated: 02 Jul 2026"
        elif "1004 buses" in v:
            cell.value = v.replace("1004 buses", f"{tot} buses")
        elif v.strip() == "1004":
            cell.value = str(tot)

# ── Social Obligation / Tourist sheets: per-route Fleet cells ──────
codes_by_rid = {csv_by_code_active[c]["New_Route_ID"]: c for c in changed if c in csv_by_code_active}
for sheet in ("Social Obligation", "Tourist & Seasonal", "Trunk Detail"):
    ws = wb[sheet]
    hits = 0
    for row in ws.iter_rows(min_row=4):
        rid = row[1].value
        if rid in codes_by_rid:
            src = csv_by_code_active[codes_by_rid[rid]]
            for cell in row:
                # the Fleet column holds the old fleet int for this route
                if isinstance(cell.value, (int, float)) and cell.value == int(float(changed[codes_by_rid[rid]]["Old_Fleet"])):
                    cell.value = int(float(src["Fleet_Required"])); hits += 1; break
    if hits: print(f"{sheet}: patched {hits} fleet cells")

# ── Calibration & Sources: version line ────────────────────────────
ws = wb["Calibration & Sources"]
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and cell.value.startswith("v3.4."):
            cell.value = ("v3.4.5 (Kashmir) — v3.4.4 audited distances + measured-cycle "
                          "corrections on 5 GPS-verified corridors (bus-sathi-trace-intelligence)")

wb.save(WB)

# ── verify: re-open and re-sum ──────────────────────────────────────
wb2 = load_workbook(WB, read_only=True, data_only=True)
ws = wb2["Route-Level Plan"]
hdr2 = {c.value: i for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))}
s = sum(int(v[hdr2["Fleet_Required"]]) for v in ws.iter_rows(min_row=2, values_only=True)
        if v[hdr2["Fleet_Required"]] and float(v[hdr2["Fleet_Required"]]) > 0)
print(f"VERIFY: Route-Level Plan fleet sum = {s} (expect {tot})  {'PASS' if s == tot else 'FAIL'}")
